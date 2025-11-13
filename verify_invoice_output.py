"""
Quick verification script to check if:
1. Static content was written correctly
2. Times New Roman font was applied
"""
from openpyxl import load_workbook
import zipfile
from io import BytesIO

def verify_invoice():
    print("\n" + "="*60)
    print("INVOICE OUTPUT VERIFICATION")
    print("="*60)
    
    # Load the workbook
    wb = load_workbook('result_test2.xlsx')
    
    # Check Invoice sheet
    if 'Invoice' in wb.sheetnames:
        ws = wb['Invoice']
        print("\n[OK] Invoice sheet found")
        
        # Check static content in column A (col 1)
        print("\n--- Static Content Check (Column A) ---")
        static_values = {
            22: "VENDOR#:",
            23: "Des: LEATHER", 
            24: "MADE IN CAMBODIA"
        }
        
        for row, expected_value in static_values.items():
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == expected_value:
                print(f"  [OK] Row {row}: '{cell_value}' (CORRECT)")
            else:
                print(f"  [NOT OK] Row {row}: Expected '{expected_value}', got '{cell_value}'")
        
        # Check font in header (openpyxl limitation - font.name may be None)
        print("\n--- Font Check (openpyxl read - may show None) ---")
        header_cell = ws['A21']  # Header row
        print(f"  Header A21 font.name: {header_cell.font.name}")
        print(f"  Header A21 font.size: {header_cell.font.size}")
        print(f"  Header A21 font.bold: {header_cell.font.bold}")
        
        # Check font in Excel XML (definitive)
        print("\n--- Font Check (Excel XML - definitive) ---")
        with zipfile.ZipFile('result_test2.xlsx', 'r') as z:
            with z.open('xl/styles.xml') as f:
                styles_xml = f.read().decode('utf-8')
                if 'Times New Roman' in styles_xml:
                    print("  [OK] Times New Roman found in styles.xml")
                    # Count occurrences
                    count = styles_xml.count('Times New Roman')
                    print(f"  [OK]OK] Found {count} occurrences of 'Times New Roman'")
                else:
                    print("  [NOT OK] Times New Roman NOT found in styles.xml")
                    if 'Calibri' in styles_xml:
                        print("  ! Found 'Calibri' instead")
    
    else:
        print("\n[NOT OK] Invoice sheet NOT found")
    
    wb.close()
    print("\n" + "="*60)

if __name__ == '__main__':
    verify_invoice()
