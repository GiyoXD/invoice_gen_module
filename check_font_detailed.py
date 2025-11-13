from openpyxl import load_workbook
import openpyxl

print(f"openpyxl version: {openpyxl.__version__}")

wb = load_workbook('result_test2.xlsx')
ws = wb['Invoice']

cell = ws['A21']
print(f"\nCell A21:")
print(f"  Value: {cell.value}")
print(f"  Font object: {cell.font}")
print(f"  Font name: {cell.font.name}")
print(f"  Font size: {cell.font.size}")
print(f"  Font bold: {cell.font.bold}")
print(f"  Font __dict__: {cell.font.__dict__}")
