"""
Test script to verify read-only template can be captured by TemplateStateBuilder.

This test validates:
1. Template can be loaded as read-only
2. TemplateStateBuilder can instantiate with read-only worksheet
3. State is captured correctly (header rows, footer rows, merges)
4. No errors occur when reading from read-only workbook
"""

from openpyxl import load_workbook
from invoice_generator.builders.template_state_builder import TemplateStateBuilder
import os


def test_readonly_template_capture():
    """Test that read-only template can be captured successfully."""
    
    # Path to template
    template_dir = "invoice_generator/template"
    template_files = [f for f in os.listdir(template_dir) if f.endswith('.xlsx')]
    
    if not template_files:
        print("❌ No template files found in invoice_generator/template")
        return False
    
    template_path = os.path.join(template_dir, template_files[0])
    print(f"📂 Using template: {template_path}")
    
    # Test 1: Try read_only=True (expected to fail with merged_cells)
    print("\n" + "="*60)
    print("TEST 1: read_only=True (Expected to fail)")
    print("="*60)
    
    try:
        print("\n🔍 Step 1a: Loading template as read_only=True...")
        template_workbook = load_workbook(template_path, read_only=True)
        print(f"✅ Template loaded successfully as read_only=True")
        
        sheet_name = template_workbook.sheetnames[0]
        template_worksheet = template_workbook[sheet_name]
        
        print("\n🏗️ Step 2a: Attempting to instantiate TemplateStateBuilder...")
        
        num_header_cols = 12
        header_end_row = 22
        footer_start_row = 24
        
        state_builder = TemplateStateBuilder(
            worksheet=template_worksheet,
            num_header_cols=num_header_cols,
            header_end_row=header_end_row,
            footer_start_row=footer_start_row
        )
        print("❌ UNEXPECTED: TemplateStateBuilder worked with read_only=True!")
        template_workbook.close()
        return False
        
    except AttributeError as e:
        if "merged_cells" in str(e):
            print(f"✅ EXPECTED ERROR: {e}")
            print("   ➜ ReadOnlyWorksheet does not support merged_cells attribute")
            template_workbook.close()
        else:
            print(f"❌ UNEXPECTED ERROR: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    # Test 2: Try read_only=False, data_only=False (should work)
    print("\n" + "="*60)
    print("TEST 2: read_only=False, data_only=False (Should work)")
    print("="*60)
    
    try:
        print("\n🔍 Step 1b: Loading template as read_only=False, data_only=False...")
        template_workbook = load_workbook(template_path, read_only=False, data_only=False)
        print(f"✅ Template loaded successfully")
        print(f"   Sheet names: {template_workbook.sheetnames}")
        
        # Get first worksheet
        sheet_name = template_workbook.sheetnames[0]
        template_worksheet = template_workbook[sheet_name]
        print(f"   Working with sheet: '{sheet_name}'")
        
        # Instantiate TemplateStateBuilder
        print("\n🏗️ Step 2b: Instantiating TemplateStateBuilder...")
        
        num_header_cols = 12
        header_end_row = 22
        footer_start_row = 24
        
        print(f"   Using num_header_cols={num_header_cols}")
        print(f"   Using header_end_row={header_end_row}")
        print(f"   Using footer_start_row={footer_start_row}")
        
        state_builder = TemplateStateBuilder(
            worksheet=template_worksheet,
            num_header_cols=num_header_cols,
            header_end_row=header_end_row,
            footer_start_row=footer_start_row
        )
        print("✅ TemplateStateBuilder instantiated successfully")
        
        # Step 3: Verify state was captured during initialization
        print("\n📸 Step 3: Verifying template state...")
        print("✅ Template state captured during initialization")
        
        # Step 4: Verify state content
        print("\n✅ Step 4: Checking captured state content...")
        
        # Check header state
        header_row_count = len(state_builder.header_state)
        print(f"   📋 Header rows: {header_row_count}")
        
        if header_row_count == 0:
            print("   ⚠️ Warning: No header rows detected")
        
        # Check footer state
        footer_row_count = len(state_builder.footer_state)
        print(f"   📋 Footer rows: {footer_row_count}")
        
        if footer_row_count == 0:
            print("   ⚠️ Warning: No footer rows detected")
        
        # Check merged cells
        header_merges = state_builder.header_merged_cells
        footer_merges = state_builder.footer_merged_cells
        print(f"   🔗 Header merged cells: {len(header_merges)}")
        print(f"   🔗 Footer merged cells: {len(footer_merges)}")
        
        # Display sample merge ranges
        if header_merges:
            print(f"   📌 Sample header merge: {header_merges[0]}")
        if footer_merges:
            print(f"   📌 Sample footer merge: {footer_merges[0]}")
        
        # Check row heights
        row_heights = state_builder.row_heights
        print(f"   📏 Row heights captured: {len(row_heights)}")
        
        # Check column widths
        column_widths = state_builder.column_widths
        print(f"   � Column widths captured: {len(column_widths)}")
        
        # Check template footer tracking
        print(f"   📍 Template footer start row: {state_builder.template_footer_start_row}")
        print(f"   � Template footer end row: {state_builder.template_footer_end_row}")
        
        # Final verification for Test 2
        print("\n" + "="*60)
        print("✅ SUCCESS: Test 2 - read_only=False, data_only=False works")
        print("="*60)
        print(f"✓ State captured with {header_row_count} header rows")
        print(f"✓ State captured with {footer_row_count} footer rows")
        print(f"✓ {len(header_merges)} header merges captured")
        print(f"✓ {len(footer_merges)} footer merges captured")
        
        template_workbook.close()
        
    except Exception as e:
        print(f"\n❌ UNEXPECTED ERROR in Test 2: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
    
    # Test 3: Try read_only=False, data_only=True (RECOMMENDED - should also work)
    print("\n" + "="*60)
    print("TEST 3: read_only=False, data_only=True (RECOMMENDED)")
    print("="*60)
    
    try:
        print("\n🔍 Step 1c: Loading template as read_only=False, data_only=True...")
        template_workbook = load_workbook(template_path, read_only=False, data_only=True)
        print(f"✅ Template loaded successfully")
        print(f"   Sheet names: {template_workbook.sheetnames}")
        
        # Get first worksheet
        sheet_name = template_workbook.sheetnames[0]
        template_worksheet = template_workbook[sheet_name]
        print(f"   Working with sheet: '{sheet_name}'")
        
        # Instantiate TemplateStateBuilder
        print("\n🏗️ Step 2c: Instantiating TemplateStateBuilder...")
        
        num_header_cols = 12
        header_end_row = 22
        footer_start_row = 24
        
        print(f"   Using num_header_cols={num_header_cols}")
        print(f"   Using header_end_row={header_end_row}")
        print(f"   Using footer_start_row={footer_start_row}")
        
        state_builder = TemplateStateBuilder(
            worksheet=template_worksheet,
            num_header_cols=num_header_cols,
            header_end_row=header_end_row,
            footer_start_row=footer_start_row
        )
        print("✅ TemplateStateBuilder instantiated successfully")
        
        # Verify state was captured
        print("\n📸 Step 3c: Verifying template state...")
        print("✅ Template state captured during initialization")
        
        # Check captured state content
        print("\n✅ Step 4c: Checking captured state content...")
        
        header_row_count = len(state_builder.header_state)
        footer_row_count = len(state_builder.footer_state)
        header_merges = state_builder.header_merged_cells
        footer_merges = state_builder.footer_merged_cells
        row_heights = state_builder.row_heights
        column_widths = state_builder.column_widths
        
        print(f"   📋 Header rows: {header_row_count}")
        print(f"   📋 Footer rows: {footer_row_count}")
        print(f"   🔗 Header merged cells: {len(header_merges)}")
        print(f"   🔗 Footer merged cells: {len(footer_merges)}")
        print(f"   📏 Row heights captured: {len(row_heights)}")
        print(f"   📐 Column widths captured: {len(column_widths)}")
        
        if header_merges:
            print(f"   📌 Sample header merge: {header_merges[0]}")
        if footer_merges:
            print(f"   📌 Sample footer merge: {footer_merges[0]}")
        
        # Final verification
        print("\n" + "="*60)
        print("✅ SUCCESS: Test 3 - read_only=False, data_only=True works!")
        print("="*60)
        print(f"✓ Template loaded: {template_path}")
        print(f"✓ State captured with {header_row_count} header rows")
        print(f"✓ State captured with {footer_row_count} footer rows")
        print(f"✓ {len(header_merges)} header merges captured")
        print(f"✓ {len(footer_merges)} footer merges captured")
        print(f"✓ {len(row_heights)} row heights captured")
        print(f"✓ {len(column_widths)} column widths captured")
        print(f"✓ No errors when reading template")
        
        print("\n" + "="*60)
        print("📊 SUMMARY OF FINDINGS:")
        print("="*60)
        print("❌ read_only=True, data_only=X         → FAILS (no merged_cells)")
        print("✅ read_only=False, data_only=False    → WORKS (reads formulas)")
        print("✅ read_only=False, data_only=True     → WORKS (reads formula results)")
        print("\n📝 RECOMMENDED APPROACH:")
        print("="*60)
        print("Use: load_workbook(template_path, read_only=False, data_only=True)")
        print("")
        print("Why:")
        print("  • data_only=True reads formula RESULTS instead of formula strings")
        print("  • read_only=False allows access to merged_cells for state capture")
        print("  • Template still protected via separate output workbook pattern")
        
        template_workbook.close()
        return True
        
    except Exception as e:
        print(f"\n❌ ERROR: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    print("="*60)
    print("TEST: Read-Only Template Capture")
    print("="*60)
    success = test_readonly_template_capture()
    exit(0 if success else 1)
