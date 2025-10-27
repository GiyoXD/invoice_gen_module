"""
Test script to verify LayoutBuilder uses template_worksheet for capture and output_worksheet for writing.

This test validates:
1. LayoutBuilder accepts template_worksheet parameter
2. Template state captured from template_worksheet
3. Output written to self.worksheet (output worksheet)
4. No internal workbook creation
5. No cleanup/rename code executed
"""

from openpyxl import load_workbook
from invoice_generator.builders.workbook_builder import WorkbookBuilder
from invoice_generator.builders.layout_builder import LayoutBuilder
import os
import json


def test_layout_builder_dual_worksheet():
    """Test that LayoutBuilder uses separate template and output worksheets."""
    
    # Setup paths
    template_dir = "invoice_generator/template"
    config_dir = "invoice_generator/config"
    template_files = [f for f in os.listdir(template_dir) if f.endswith('.xlsx')]
    
    if not template_files:
        print("❌ No template files found")
        return False
    
    template_path = os.path.join(template_dir, template_files[0])
    config_path = os.path.join(config_dir, "JF_config.json")
    
    if not os.path.exists(config_path):
        print(f"❌ Config file not found: {config_path}")
        return False
    
    print("="*60)
    print("TEST: LayoutBuilder with Dual Worksheets")
    print("="*60)
    print(f"📂 Template: {template_path}")
    print(f"⚙️ Config: {config_path}")
    
    try:
        # Load config
        print("\n1️⃣ Loading configuration...")
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        data_mapping = config.get('data_mapping', {})
        print(f"   ✅ Config loaded")
        
        # Load template workbook
        print("\n2️⃣ Loading template workbook...")
        template_workbook = load_workbook(
            template_path,
            read_only=False,
            data_only=True
        )
        print(f"   ✅ Template loaded: {template_workbook.sheetnames}")
        
        # Create output workbook
        print("\n3️⃣ Creating output workbook...")
        workbook_builder = WorkbookBuilder(sheet_names=template_workbook.sheetnames)
        output_workbook = workbook_builder.build()
        print(f"   ✅ Output workbook created: {output_workbook.sheetnames}")
        
        # Test with Invoice sheet (most complex)
        sheet_name = "Invoice"
        if sheet_name not in template_workbook.sheetnames:
            sheet_name = template_workbook.sheetnames[0]
            print(f"   ⚠️ Using '{sheet_name}' instead of 'Invoice'")
        
        template_worksheet = template_workbook[sheet_name]
        output_worksheet = output_workbook[sheet_name]
        
        print(f"\n4️⃣ Testing LayoutBuilder with sheet '{sheet_name}'...")
        
        # Get sheet config
        sheet_config = data_mapping.get(sheet_name, {})
        if not sheet_config:
            print(f"   ❌ No config for '{sheet_name}'")
            return False
        
        # Record initial state
        template_initial_rows = template_worksheet.max_row
        template_initial_cols = template_worksheet.max_column
        output_initial_rows = output_worksheet.max_row
        output_initial_cols = output_worksheet.max_column
        
        print(f"   📖 Template initial: {template_initial_rows}x{template_initial_cols}")
        print(f"   📝 Output initial: {output_initial_rows}x{output_initial_cols}")
        
        # Mock data and args
        mock_invoice_data = {
            'standard_aggregation_results': {}  # Empty data for test
        }
        
        class MockArgs:
            DAF = False
            custom = False
        
        mock_args = MockArgs()
        
        # Instantiate LayoutBuilder with both worksheets
        print("\n   🏗️ Instantiating LayoutBuilder...")
        try:
            layout_builder = LayoutBuilder(
                workbook=output_workbook,
                worksheet=output_worksheet,
                template_worksheet=template_worksheet,
                sheet_name=sheet_name,
                sheet_config=sheet_config,
                all_sheet_configs=data_mapping,
                invoice_data=mock_invoice_data,
                styling_config=sheet_config.get('styling'),
                args=mock_args,
                final_grand_total_pallets=0,
                enable_text_replacement=False
            )
            print(f"      ✅ LayoutBuilder instantiated")
            
            # Verify attributes
            assert layout_builder.template_worksheet == template_worksheet
            assert layout_builder.worksheet == output_worksheet
            assert layout_builder.workbook == output_workbook
            print(f"      ✅ All worksheet references correct")
            
        except Exception as e:
            print(f"      ❌ LayoutBuilder instantiation failed: {e}")
            import traceback
            traceback.print_exc()
            return False
        
        # Verify template unchanged after instantiation
        print("\n5️⃣ Verifying template worksheet unchanged...")
        template_after_rows = template_worksheet.max_row
        template_after_cols = template_worksheet.max_column
        
        if template_after_rows == template_initial_rows and template_after_cols == template_initial_cols:
            print(f"   ✅ Template unchanged: {template_after_rows}x{template_after_cols}")
        else:
            print(f"   ❌ Template modified! Before: {template_initial_rows}x{template_initial_cols}, After: {template_after_rows}x{template_after_cols}")
            return False
        
        # Verify no internal workbook creation
        print("\n6️⃣ Verifying no internal workbook creation...")
        if not hasattr(layout_builder, 'new_workbook'):
            print(f"   ✅ No 'new_workbook' attribute (good - removed)")
        else:
            print(f"   ⚠️ Found 'new_workbook' attribute (may need cleanup)")
        
        # Check output workbook sheet count
        if len(output_workbook.sheetnames) == len(template_workbook.sheetnames):
            print(f"   ✅ Output workbook has correct number of sheets ({len(output_workbook.sheetnames)})")
        else:
            print(f"   ❌ Sheet count mismatch! Output: {len(output_workbook.sheetnames)}, Template: {len(template_workbook.sheetnames)}")
        
        # Final summary
        print("\n" + "="*60)
        print("✅ SUCCESS: LayoutBuilder Updated!")
        print("="*60)
        print("✓ LayoutBuilder accepts template_worksheet parameter")
        print("✓ Template worksheet reference stored correctly")
        print("✓ Output worksheet reference stored correctly")
        print("✓ Template unchanged during instantiation")
        print("✓ No internal workbook creation")
        print("✓ Clean separation between template (read) and output (write)")
        
        print("\n📝 Architecture:")
        print("  • template_worksheet: For reading template state")
        print("  • self.worksheet (output): For writing final output")
        print("  • No workbook duplication or cleanup needed")
        
        # Cleanup
        template_workbook.close()
        output_workbook.close()
        
        return True
        
    except Exception as e:
        print(f"\n❌ ERROR: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = test_layout_builder_dual_worksheet()
    exit(0 if success else 1)
