"""
Test script to verify processor interface accepts both template and output workbooks.

This test validates:
1. Template and output workbooks can be passed to processors
2. Both SingleTableProcessor and MultiTableProcessor accept new interface
3. Processors can access both template and output worksheets
4. Backward compatibility maintained (self.workbook and self.worksheet still work)
"""

from openpyxl import load_workbook
from invoice_generator.builders.workbook_builder import WorkbookBuilder
from invoice_generator.processors.single_table_processor import SingleTableProcessor
from invoice_generator.processors.multi_table_processor import MultiTableProcessor
import os
import json


def test_processor_interface():
    """Test that processors accept new interface with both workbooks."""
    
    # Setup paths
    template_dir = "invoice_generator/template"
    config_dir = "invoice_generator/config"
    template_files = [f for f in os.listdir(template_dir) if f.endswith('.xlsx')]
    
    if not template_files:
        print("❌ No template files found")
        return False
    
    template_path = os.path.join(template_dir, template_files[0])
    config_path = os.path.join(config_dir, "JF_config.json")
    
    if not os.path.exists(config_path):
        print(f"❌ Config file not found: {config_path}")
        return False
    
    print("="*60)
    print("TEST: Processor Interface with Dual Workbooks")
    print("="*60)
    print(f"📂 Template: {template_path}")
    print(f"⚙️ Config: {config_path}")
    
    try:
        # Load config
        print("\n1️⃣ Loading configuration...")
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        data_mapping = config.get('data_mapping', {})
        print(f"   ✅ Config loaded")
        
        # Load template workbook
        print("\n2️⃣ Loading template workbook...")
        template_workbook = load_workbook(
            template_path,
            read_only=False,
            data_only=True
        )
        print(f"   ✅ Template loaded: {template_workbook.sheetnames}")
        
        # Create output workbook
        print("\n3️⃣ Creating output workbook...")
        workbook_builder = WorkbookBuilder(sheet_names=template_workbook.sheetnames)
        output_workbook = workbook_builder.build()
        print(f"   ✅ Output workbook created: {output_workbook.sheetnames}")
        
        # Test with first sheet
        sheet_name = template_workbook.sheetnames[0]
        template_worksheet = template_workbook[sheet_name]
        output_worksheet = output_workbook[sheet_name]
        
        print(f"\n4️⃣ Testing processor instantiation with sheet '{sheet_name}'...")
        
        # Get sheet config
        sheet_config = data_mapping.get(sheet_name, {})
        if not sheet_config:
            print(f"   ⚠️ No config for '{sheet_name}', using minimal config")
            sheet_config = {
                'start_row': 21,
                'header_to_write': []
            }
        
        # Mock data and args
        mock_invoice_data = {}
        
        class MockArgs:
            DAF = False
            custom = False
        
        mock_args = MockArgs()
        
        # Test SingleTableProcessor
        print("\n   📋 Testing SingleTableProcessor...")
        try:
            single_processor = SingleTableProcessor(
                template_workbook=template_workbook,
                output_workbook=output_workbook,
                template_worksheet=template_worksheet,
                output_worksheet=output_worksheet,
                sheet_name=sheet_name,
                sheet_config=sheet_config,
                data_mapping_config=data_mapping,
                data_source_indicator="aggregation",
                invoice_data=mock_invoice_data,
                cli_args=mock_args,
                final_grand_total_pallets=0
            )
            print(f"      ✅ SingleTableProcessor instantiated")
            
            # Verify attributes
            assert single_processor.template_workbook == template_workbook
            assert single_processor.output_workbook == output_workbook
            assert single_processor.template_worksheet == template_worksheet
            assert single_processor.output_worksheet == output_worksheet
            print(f"      ✅ All new attributes accessible")
            
            # Verify backward compatibility
            assert single_processor.workbook == output_workbook
            assert single_processor.worksheet == output_worksheet
            print(f"      ✅ Backward compatibility maintained")
            
        except Exception as e:
            print(f"      ❌ SingleTableProcessor failed: {e}")
            raise
        
        # Test MultiTableProcessor
        print("\n   📋 Testing MultiTableProcessor...")
        try:
            multi_processor = MultiTableProcessor(
                template_workbook=template_workbook,
                output_workbook=output_workbook,
                template_worksheet=template_worksheet,
                output_worksheet=output_worksheet,
                sheet_name=sheet_name,
                sheet_config=sheet_config,
                data_mapping_config=data_mapping,
                data_source_indicator="processed_tables_multi",
                invoice_data=mock_invoice_data,
                cli_args=mock_args,
                final_grand_total_pallets=0
            )
            print(f"      ✅ MultiTableProcessor instantiated")
            
            # Verify attributes
            assert multi_processor.template_workbook == template_workbook
            assert multi_processor.output_workbook == output_workbook
            assert multi_processor.template_worksheet == template_worksheet
            assert multi_processor.output_worksheet == output_worksheet
            print(f"      ✅ All new attributes accessible")
            
            # Verify backward compatibility
            assert multi_processor.workbook == output_workbook
            assert multi_processor.worksheet == output_worksheet
            print(f"      ✅ Backward compatibility maintained")
            
        except Exception as e:
            print(f"      ❌ MultiTableProcessor failed: {e}")
            raise
        
        # Verify worksheet access
        print("\n5️⃣ Verifying worksheet access...")
        print(f"   📖 Template worksheet dimensions: {template_worksheet.max_row}x{template_worksheet.max_column}")
        print(f"   📝 Output worksheet dimensions: {output_worksheet.max_row}x{output_worksheet.max_column}")
        
        if template_worksheet.max_row > 1:
            print(f"   ✅ Template has content (as expected)")
        if output_worksheet.max_row == 1 and output_worksheet.max_column == 1:
            print(f"   ✅ Output is empty (as expected)")
        
        # Final summary
        print("\n" + "="*60)
        print("✅ SUCCESS: Processor Interface Updated!")
        print("="*60)
        print("✓ Both processors accept new interface")
        print("✓ template_workbook and output_workbook parameters work")
        print("✓ template_worksheet and output_worksheet parameters work")
        print("✓ All new attributes accessible in processors")
        print("✓ Backward compatibility maintained (self.workbook, self.worksheet)")
        print("✓ Ready for LayoutBuilder integration")
        
        # Cleanup
        template_workbook.close()
        output_workbook.close()
        
        return True
        
    except Exception as e:
        print(f"\n❌ ERROR: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = test_processor_interface()
    exit(0 if success else 1)
