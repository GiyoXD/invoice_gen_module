"""
Test script to build an actual workbook using WorkbookBuilder and save it.
This allows manual verification that the workbook opens correctly in Excel.
"""

from openpyxl import load_workbook
from invoice_generator.builders.workbook_builder import WorkbookBuilder
import os


def test_build_and_save_workbook():
    """Build a workbook using WorkbookBuilder and save it for manual inspection."""
    
    # Path to template
    template_dir = "invoice_generator/template"
    template_files = [f for f in os.listdir(template_dir) if f.endswith('.xlsx')]
    
    if not template_files:
        print("❌ No template files found")
        return False
    
    template_path = os.path.join(template_dir, template_files[0])
    output_path = "test_workbook_builder_output.xlsx"
    
    print("="*60)
    print("TEST: Build and Save Workbook with WorkbookBuilder")
    print("="*60)
    print(f"📂 Template: {template_path}")
    print(f"💾 Output: {output_path}")
    
    try:
        # Step 1: Load template to get sheet names
        print("\n1️⃣ Loading template to get sheet names...")
        template_workbook = load_workbook(
            template_path,
            read_only=False,
            data_only=True
        )
        sheet_names = template_workbook.sheetnames
        print(f"   Found {len(sheet_names)} sheets: {sheet_names}")
        
        # Step 2: Create WorkbookBuilder
        print("\n2️⃣ Creating WorkbookBuilder...")
        workbook_builder = WorkbookBuilder(sheet_names=sheet_names)
        print(f"   WorkbookBuilder initialized with {len(sheet_names)} sheets")
        
        # Step 3: Build the workbook
        print("\n3️⃣ Building new workbook...")
        new_workbook = workbook_builder.build()
        print(f"   ✅ Workbook built successfully")
        print(f"   Sheets in new workbook: {new_workbook.sheetnames}")
        
        # Step 4: Add some test content so we can verify it opens
        print("\n4️⃣ Adding test content to verify functionality...")
        for i, sheet_name in enumerate(new_workbook.sheetnames, 1):
            ws = new_workbook[sheet_name]
            ws['A1'] = f"Sheet {i}: {sheet_name}"
            ws['A2'] = "This workbook was created by WorkbookBuilder"
            ws['A3'] = "✅ If you can read this, the workbook works!"
            
            # Add some formatting
            from openpyxl.styles import Font, PatternFill
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            print(f"   ✅ Added test content to '{sheet_name}'")
        
        # Step 5: Save the workbook
        print(f"\n5️⃣ Saving workbook to: {output_path}")
        new_workbook.save(output_path)
        print(f"   ✅ Workbook saved successfully")
        
        # Step 6: Verify file exists
        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            print(f"\n6️⃣ Verifying output file...")
            print(f"   ✅ File exists: {output_path}")
            print(f"   📊 File size: {file_size:,} bytes")
        else:
            print(f"\n   ❌ ERROR: File not found after save!")
            return False
        
        # Cleanup
        template_workbook.close()
        new_workbook.close()
        
        # Final summary
        print("\n" + "="*60)
        print("✅ SUCCESS: Workbook created and saved!")
        print("="*60)
        print(f"📁 Output file: {output_path}")
        print(f"📋 Sheets: {len(sheet_names)}")
        for i, name in enumerate(sheet_names, 1):
            print(f"   {i}. {name}")
        print("\n🔍 Please open the file to verify:")
        print(f"   • File opens without errors")
        print(f"   • All {len(sheet_names)} sheets are present")
        print(f"   • Test content is visible in each sheet")
        print(f"   • No corruption warnings")
        
        return True
        
    except Exception as e:
        print(f"\n❌ ERROR: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = test_build_and_save_workbook()
    
    if success:
        print("\n" + "="*60)
        print("✅ TEST PASSED - Ready to open file!")
        print("="*60)
    else:
        print("\n" + "="*60)
        print("❌ TEST FAILED")
        print("="*60)
    
    exit(0 if success else 1)
