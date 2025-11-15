"""
Tests for TableDataAdapter (Adapter Pattern)

Tests the table data preparation logic that transforms raw invoice data
into table-ready row dictionaries.

Run these tests:
    python -m pytest tests/config/test_table_data_resolver.py -v
    python -m pytest tests/config/test_table_data_resolver.py::TestTableDataAdapter::test_resolve_with_real_invoice_config -v
"""
import unittest
from unittest.mock import Mock
from pathlib import Path

from invoice_generator.config.config_loader import BundledConfigLoader
from invoice_generator.config.builder_config_resolver import BuilderConfigResolver
from invoice_generator.config.table_value_adapter import TableDataAdapter


class TestTableDataAdapter(unittest.TestCase):
    """Test suite for TableDataAdapter class."""
    
    @classmethod
    def setUpClass(cls):
        """Load the real JF config once."""
        config_path = Path(__file__).parent.parent.parent / "invoice_generator" / "config_bundled" / "JF_config" / "JF_config.json"
        cls.config_loader = BundledConfigLoader(config_path)
    
    def setUp(self):
        """Set up test fixtures."""
        # Sample invoice data with different data source types
        self.invoice_data = {
            'standard_aggregation_results': {
                ('PO123', 'ITEM001', 5.25, 'LEATHER TYPE A'): {
                    'sqft_sum': 100.50,
                    'amount_sum': 527.625
                },
                ('PO123', 'ITEM002', 6.50, 'LEATHER TYPE B'): {
                    'sqft_sum': 200.75,
                    'amount_sum': 1304.875
                }
            },
            'processed_tables_data': {
                '1': {
                    'po': ['PO123', 'PO123'],
                    'item': ['ITEM001', 'ITEM002'],
                    'description': ['LEATHER', 'LEATHER'],
                    'pcs': [10, 20],
                    'sqft': [100.50, 200.75],
                    'net': [50.0, 100.0],
                    'gross': [52.5, 105.0],
                    'cbm': [0.15, 0.30],
                    'pallet_count': [2, 3]
                },
                '2': {
                    'po': ['PO456'],
                    'item': ['ITEM003'],
                    'description': ['LEATHER'],
                    'pcs': [15],
                    'sqft': [150.25],
                    'net': [75.0],
                    'gross': [78.75],
                    'cbm': [0.22],
                    'pallet_count': [1]
                }
            }
        }
        
        self.args = Mock(DAF=False, custom=False)
    
    # ========== Initialization Tests ==========
    
    def test_initialization_with_standard_aggregation(self):
        """Test initializing resolver with standard aggregation data."""
        header_info = {
            'column_map': {'PO': 1, 'Item': 2, 'Description': 3},
            'column_id_map': {'col_po': 1, 'col_item': 2, 'col_desc': 3}
        }
        
        resolver = TableDataAdapter(
            data_source_type='aggregation',
            data_source=self.invoice_data['standard_aggregation_results'],
            mapping_rules={},
            header_info=header_info,
            DAF_mode=False
        )
        
        self.assertEqual(resolver.data_source_type, 'aggregation')
        self.assertFalse(resolver.DAF_mode)
        self.assertIsNone(resolver.table_key)
    
    def test_initialization_with_table_key(self):
        """Test initializing resolver with table key for multi-table data."""
        header_info = {
            'column_map': {},
            'column_id_map': {}
        }
        
        resolver = TableDataAdapter(
            data_source_type='processed_tables',
            data_source=self.invoice_data['processed_tables_data']['1'],
            mapping_rules={},
            header_info=header_info,
            DAF_mode=False,
            table_key='1'
        )
        
        self.assertEqual(resolver.table_key, '1')
    
    # ========== Factory Method Tests ==========
    
    def test_create_from_bundles(self):
        """Test creating resolver from bundle configs."""
        data_config = {
            'data_source_type': 'aggregation',
            'data_source': {'test': 'data'},
            'mapping_rules': {},
            'header_info': {
                'column_map': {},
                'column_id_map': {}
            },
            'table_key': None
        }
        
        context_config = {
            'args': self.args
        }
        
        resolver = TableDataAdapter.create_from_bundles(
            data_config=data_config,
            context_config=context_config
        )
        
        self.assertIsInstance(resolver, TableDataAdapter)
        self.assertEqual(resolver.data_source_type, 'aggregation')
        self.assertFalse(resolver.DAF_mode)
    
    def test_create_from_bundles_with_daf_mode(self):
        """Test factory method detects DAF mode from args."""
        data_config = {
            'data_source_type': 'DAF_aggregation',
            'data_source': {},
            'mapping_rules': {},
            'header_info': {'column_map': {}, 'column_id_map': {}}
        }
        
        args = Mock(DAF=True)
        context_config = {'args': args}
        
        resolver = TableDataAdapter.create_from_bundles(
            data_config=data_config,
            context_config=context_config
        )
        
        self.assertTrue(resolver.DAF_mode)
    
    # ========== Integration with BuilderConfigResolver ==========
    
    def test_get_table_data_resolver_from_builder_resolver(self):
        """Test getting TableDataAdapter from BuilderConfigResolver."""
        from openpyxl import Workbook
        workbook = Workbook()
        worksheet = workbook.active
        
        builder_resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Invoice',
            worksheet=worksheet,
            args=self.args,
            invoice_data=self.invoice_data
        )
        
        table_resolver = builder_resolver.get_table_data_resolver()
        
        self.assertIsInstance(table_resolver, TableDataAdapter)
        self.assertEqual(table_resolver.data_source_type, 'aggregation')
    
    def test_get_table_data_resolver_with_table_key(self):
        """Test getting TableDataAdapter for specific table."""
        from openpyxl import Workbook
        workbook = Workbook()
        worksheet = workbook.active
        
        builder_resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Packing list',
            worksheet=worksheet,
            args=self.args,
            invoice_data=self.invoice_data
        )
        
        table_resolver = builder_resolver.get_table_data_resolver(table_key='1')
        
        self.assertIsInstance(table_resolver, TableDataAdapter)
        self.assertEqual(table_resolver.table_key, '1')
    
    # ========== Helper Method Tests ==========
    
    def test_idx_to_header_map_generation(self):
        """Test that reverse header map is generated correctly."""
        header_info = {
            'column_map': {'PO': 1, 'Item': 2, 'Description': 3},
            'column_id_map': {'col_po': 1, 'col_item': 2, 'col_desc': 3}
        }
        
        resolver = TableDataAdapter(
            data_source_type='aggregation',
            data_source={},
            mapping_rules={},
            header_info=header_info,
            DAF_mode=False
        )
        
        self.assertEqual(resolver.idx_to_header_map[1], 'PO')
        self.assertEqual(resolver.idx_to_header_map[2], 'Item')
        self.assertEqual(resolver.idx_to_header_map[3], 'Description')
    
    def test_get_desc_col_idx_finds_description_column(self):
        """Test that description column index is found correctly."""
        header_info = {
            'column_map': {},
            'column_id_map': {'col_po': 1, 'col_desc': 3, 'col_item': 2}
        }
        
        resolver = TableDataAdapter(
            data_source_type='aggregation',
            data_source={},
            mapping_rules={},
            header_info=header_info,
            DAF_mode=False
        )
        
        desc_idx = resolver._get_desc_col_idx()
        self.assertEqual(desc_idx, 3)
    
    def test_extract_table_data_for_multi_table(self):
        """Test extracting specific table data."""
        header_info = {
            'column_map': {},
            'column_id_map': {}
        }
        
        # When BuilderConfigResolver extracts a table, it passes the extracted dict directly
        # So table_1_data is already the table data (dict with column arrays), not a dict of tables
        table_1_data = self.invoice_data['processed_tables_data']['1']
        
        resolver = TableDataAdapter(
            data_source_type='processed_tables_multi',
            data_source=table_1_data,
            mapping_rules={},
            header_info=header_info,
            DAF_mode=False,
            table_key='1'  # This tells resolver which table this data is for
        )
        
        extracted = resolver._extract_table_data()
        # data_source is already a dict with column arrays (already extracted by BuilderConfigResolver)
        # So _extract_table_data should return it as-is
        self.assertIsInstance(extracted, dict)
        self.assertIn('po', extracted)
        self.assertEqual(len(extracted['po']), 2)  # Table 1 has 2 rows
    
    # ========== Resolution Tests (Integration with real config) ==========
    
    def test_resolve_with_real_invoice_config(self):
        """Test full resolution with real Invoice sheet config."""
        from openpyxl import Workbook
        workbook = Workbook()
        worksheet = workbook.active
        
        builder_resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Invoice',
            worksheet=worksheet,
            args=self.args,
            invoice_data=self.invoice_data
        )
        
        table_resolver = builder_resolver.get_table_data_resolver()
        result = table_resolver.resolve()
        
        # Check result structure
        self.assertIn('data_rows', result)
        self.assertIn('pallet_counts', result)
        self.assertIn('dynamic_desc_used', result)
        self.assertIn('num_data_rows', result)
        self.assertIn('static_info', result)
        self.assertIn('formula_rules', result)
        
        # Check that data rows were prepared
        self.assertIsInstance(result['data_rows'], list)
        self.assertIsInstance(result['pallet_counts'], list)
        self.assertIsInstance(result['dynamic_desc_used'], bool)
    
    def test_resolve_with_packing_list_config(self):
        """Test full resolution with real Packing list sheet config."""
        from openpyxl import Workbook
        workbook = Workbook()
        worksheet = workbook.active
        
        builder_resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Packing list',
            worksheet=worksheet,
            args=self.args,
            invoice_data=self.invoice_data
        )
        
        table_resolver = builder_resolver.get_table_data_resolver(table_key='1')
        result = table_resolver.resolve()
        
        # Should have data for table 1
        self.assertIn('data_rows', result)
        self.assertEqual(result['num_data_rows'], 2)  # Table 1 has 2 rows
    
    def test_static_info_extraction(self):
        """Test that static info is correctly extracted."""
        from openpyxl import Workbook
        workbook = Workbook()
        worksheet = workbook.active
        
        builder_resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Invoice',
            worksheet=worksheet,
            args=self.args,
            invoice_data=self.invoice_data
        )
        
        table_resolver = builder_resolver.get_table_data_resolver()
        result = table_resolver.resolve()
        
        static_info = result['static_info']
        self.assertIn('col1_index', static_info)
        self.assertIn('num_static_labels', static_info)
        self.assertIn('initial_static_col1_values', static_info)
        self.assertIn('static_column_header_name', static_info)
        self.assertIn('apply_special_border_rule', static_info)


class TestTableDataAdapterEdgeCases(unittest.TestCase):
    """Test edge cases and error handling."""
    
    def test_resolver_with_empty_data_source(self):
        """Test resolver handles empty data source."""
        header_info = {
            'column_map': {},
            'column_id_map': {}
        }
        
        resolver = TableDataAdapter(
            data_source_type='aggregation',
            data_source=None,
            mapping_rules={},
            header_info=header_info,
            DAF_mode=False
        )
        
        result = resolver.resolve()
        self.assertEqual(result['num_data_rows'], 0)
        self.assertEqual(len(result['data_rows']), 0)
    
    def test_resolver_with_missing_header_info(self):
        """Test resolver handles missing header info gracefully."""
        resolver = TableDataAdapter(
            data_source_type='aggregation',
            data_source={},
            mapping_rules={},
            header_info={},  # Empty header info
            DAF_mode=False
        )
        
        # Should not crash
        result = resolver.resolve()
        self.assertIsInstance(result, dict)


if __name__ == '__main__':
    unittest.main()


# ========== Additional Isolated Unit Tests ==========

class TestTableDataAdapterIsolated(unittest.TestCase):
    """Isolated unit tests that don't require real config files."""
    
    def test_bundled_to_legacy_mapping_conversion(self):
        """Test conversion from bundled config mapping format to legacy format."""
        bundled_mappings = {
            'po': {
                'column': 'col_po',
                'source_key': 0
            },
            'item': {
                'column': 'col_item',
                'source_key': 1,
                'fallback': 'N/A'
            },
            'total': {
                'column': 'col_total',
                'formula_template': 'SUM({sqft},{amount})',
                'inputs': ['sqft', 'amount']
            }
        }
        
        header_info = {
            'column_map': {},
            'column_id_map': {
                'col_po': 1,
                'col_item': 2,
                'col_total': 5
            }
        }
        
        resolver = TableDataAdapter(
            data_source_type='aggregation',
            data_source={},
            mapping_rules=bundled_mappings,
            header_info=header_info,
            DAF_mode=False
        )
        
        # Call the conversion method
        legacy = resolver._convert_bundled_mappings_to_legacy(bundled_mappings)
        
        # Verify conversions
        self.assertEqual(legacy['po']['id'], 'col_po')
        self.assertEqual(legacy['po']['key_index'], 0)
        
        self.assertEqual(legacy['item']['id'], 'col_item')
        self.assertEqual(legacy['item']['key_index'], 1)
        self.assertEqual(legacy['item']['fallback'], 'N/A')
        
        self.assertEqual(legacy['total']['id'], 'col_total')
        self.assertEqual(legacy['total']['formula_template'], 'SUM({sqft},{amount})')
    
    def test_all_three_data_source_types(self):
        """Test that all 3 supported data source types work."""
        header_info = {
            'column_map': {},
            'column_id_map': {'col_po': 1}
        }
        
        # Test aggregation
        resolver1 = TableDataAdapter(
            data_source_type='aggregation',
            data_source={},
            mapping_rules={},
            header_info=header_info,
            DAF_mode=False
        )
        self.assertEqual(resolver1.data_source_type, 'aggregation')
        
        # Test DAF_aggregation
        resolver2 = TableDataAdapter(
            data_source_type='DAF_aggregation',
            data_source={},
            mapping_rules={},
            header_info=header_info,
            DAF_mode=True
        )
        self.assertEqual(resolver2.data_source_type, 'DAF_aggregation')
        self.assertTrue(resolver2.DAF_mode)
        
        # Test processed_tables_multi
        resolver3 = TableDataAdapter(
            data_source_type='processed_tables_multi',
            data_source={'po': ['PO123']},
            mapping_rules={},
            header_info=header_info,
            DAF_mode=False,
            table_key='1'
        )
        self.assertEqual(resolver3.data_source_type, 'processed_tables_multi')
        self.assertEqual(resolver3.table_key, '1')
    
    def test_resolver_caches_parsed_rules(self):
        """Test that parsed mapping rules are cached after first parse."""
        header_info = {
            'column_map': {},
            'column_id_map': {'col_po': 1}
        }
        
        resolver = TableDataAdapter(
            data_source_type='aggregation',
            data_source={},
            mapping_rules={'po': {'column': 'col_po', 'source_key': 0}},
            header_info=header_info,
            DAF_mode=False
        )
        
        # First call should parse
        self.assertIsNone(resolver._parsed_rules)
        parsed1 = resolver._parse_mapping_rules()
        self.assertIsNotNone(resolver._parsed_rules)
        
        # Second call should return cached
        parsed2 = resolver._parse_mapping_rules()
        self.assertIs(parsed1, parsed2)  # Same object reference
