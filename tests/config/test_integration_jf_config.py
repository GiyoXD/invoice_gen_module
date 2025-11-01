"""
Integration tests using real JF_config.json

Tests the full chain: BundledConfigLoader → BuilderConfigResolver → Bundle extraction
This ensures the actual production config works correctly with our resolver.
"""
import unittest
from pathlib import Path
from unittest.mock import Mock
from openpyxl import Workbook

from invoice_generator.config.config_loader import BundledConfigLoader
from invoice_generator.config.builder_config_resolver import BuilderConfigResolver


class TestJFConfigIntegration(unittest.TestCase):
    """Integration tests using the real JF configuration file."""
    
    @classmethod
    def setUpClass(cls):
        """Load the real JF config once for all tests."""
        config_path = Path(__file__).parent.parent.parent / "invoice_generator" / "config_bundled" / "JF_config" / "JF_config.json"
        if not config_path.exists():
            raise FileNotFoundError(f"JF config not found at: {config_path}")
        
        cls.config_loader = BundledConfigLoader(config_path)
        cls.config_path = config_path
    
    def setUp(self):
        """Set up test fixtures before each test."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
        # Mock invoice data matching JF structure
        self.invoice_data = {
            'standard_aggregation_results': {
                'table_1': [
                    {'po': 'PO123', 'item': 'ITEM001', 'description': 'LEATHER TYPE A', 'sqft': 100.50, 'unit_price': 5.25},
                    {'po': 'PO123', 'item': 'ITEM002', 'description': 'LEATHER TYPE B', 'sqft': 200.75, 'unit_price': 6.50}
                ]
            },
            'processed_tables_data': {
                '1': [
                    {'po': 'PO123', 'item': 'ITEM001', 'description': 'LEATHER', 'pcs': 10, 'sqft': 100.50, 'net': 50.0, 'gross': 52.5, 'cbm': 0.15},
                    {'po': 'PO123', 'item': 'ITEM002', 'description': 'LEATHER', 'pcs': 20, 'sqft': 200.75, 'net': 100.0, 'gross': 105.0, 'cbm': 0.30}
                ],
                '2': [
                    {'po': 'PO456', 'item': 'ITEM003', 'description': 'LEATHER', 'pcs': 15, 'sqft': 150.25, 'net': 75.0, 'gross': 78.75, 'cbm': 0.22}
                ]
            }
        }
        
        self.args = Mock(DAF=False, custom=False)
    
    # ========== Config Loader Tests ==========
    
    def test_config_loader_version(self):
        """Test that JF config version is correctly loaded."""
        self.assertEqual(self.config_loader.version, "2.1_developer_optimized")
    
    def test_config_loader_customer(self):
        """Test that customer name is correctly loaded."""
        self.assertEqual(self.config_loader.customer, "JF")
    
    def test_config_loader_sheets_to_process(self):
        """Test that sheets to process are correctly extracted."""
        sheets = self.config_loader.get_sheets_to_process()
        self.assertEqual(sheets, ["Invoice", "Contract", "Packing list"])
    
    def test_config_loader_data_sources(self):
        """Test that data sources are correctly mapped."""
        self.assertEqual(self.config_loader.get_data_source_type("Invoice"), "aggregation")
        self.assertEqual(self.config_loader.get_data_source_type("Contract"), "aggregation")
        self.assertEqual(self.config_loader.get_data_source_type("Packing list"), "processed_tables_multi")
    
    def test_config_loader_features(self):
        """Test that feature flags are accessible."""
        features = self.config_loader.get_features()
        self.assertFalse(features.get('enable_text_replacement'))
        self.assertTrue(features.get('enable_auto_calculations'))
    
    # ========== Invoice Sheet Tests ==========
    
    def test_invoice_sheet_resolver_initialization(self):
        """Test resolver initialization with Invoice sheet."""
        resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Invoice',
            worksheet=self.worksheet,
            args=self.args,
            invoice_data=self.invoice_data,
            pallets=31
        )
        
        self.assertEqual(resolver.sheet_name, 'Invoice')
        self.assertIsNotNone(resolver._sheet_config)
    
    def test_invoice_sheet_structure(self):
        """Test Invoice sheet structure configuration."""
        sheet_config = self.config_loader.get_layout_config('Invoice')
        structure = sheet_config.get('structure', {})
        
        self.assertEqual(structure['header_row'], 21)
        
        columns = structure.get('columns', [])
        self.assertEqual(len(columns), 7)
        
        # Check specific columns
        col_ids = [col['id'] for col in columns]
        self.assertIn('col_static', col_ids)
        self.assertIn('col_po', col_ids)
        self.assertIn('col_amount', col_ids)
    
    def test_invoice_sheet_styling(self):
        """Test Invoice sheet styling configuration."""
        styling = self.config_loader.get_styling_config('Invoice')
        
        # Check header styling
        header = styling.get('header', {})
        self.assertEqual(header['font']['size'], 12)
        self.assertTrue(header['font']['bold'])
        self.assertEqual(header['row_height'], 35)
        
        # Check data styling
        data = styling.get('data', {})
        self.assertEqual(data['font']['size'], 12)
        self.assertEqual(data['row_height'], 35)
    
    def test_invoice_sheet_data_flow(self):
        """Test Invoice sheet data flow mappings."""
        sheet_config = self.config_loader.get_layout_config('Invoice')
        mappings = sheet_config.get('data_flow', {}).get('mappings', {})
        
        self.assertIn('po', mappings)
        self.assertIn('item', mappings)
        self.assertIn('amount', mappings)
        
        # Check specific mapping
        po_mapping = mappings['po']
        self.assertEqual(po_mapping['column'], 'col_po')
        self.assertEqual(po_mapping['source_key'], 0)
        
        # Check formula mapping
        amount_mapping = mappings['amount']
        self.assertEqual(amount_mapping['column'], 'col_amount')
        self.assertIn('formula', amount_mapping)
    
    def test_invoice_sheet_header_bundles(self):
        """Test header bundles for Invoice sheet."""
        resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Invoice',
            worksheet=self.worksheet,
            args=self.args,
            invoice_data=self.invoice_data
        )
        
        style_config, context_config, layout_config = resolver.get_header_bundles()
        
        # Check style bundle
        self.assertIn('styling_config', style_config)
        styling = style_config['styling_config']
        self.assertIn('header', styling)
        
        # Check context bundle
        self.assertEqual(context_config['sheet_name'], 'Invoice')
        self.assertIn('all_sheet_configs', context_config)
        
        # Check layout bundle
        self.assertIn('sheet_config', layout_config)
    
    def test_invoice_sheet_data_bundles(self):
        """Test data bundles for Invoice sheet."""
        resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Invoice',
            worksheet=self.worksheet,
            args=self.args,
            invoice_data=self.invoice_data
        )
        
        style_config, context_config, layout_config, data_config = resolver.get_datatable_bundles()
        
        # Check data bundle
        self.assertIn('data_source', data_config)
        self.assertIn('header_info', data_config)
        self.assertIn('mapping_rules', data_config)
        
        # Check header_info construction
        header_info = data_config['header_info']
        self.assertIn('column_map', header_info)
        self.assertIn('column_id_map', header_info)
        self.assertEqual(header_info['second_row_index'], 22)  # header_row 21 + 1
        self.assertEqual(header_info['num_columns'], 7)
        
        # Check column maps
        self.assertIn('Mark & Nº', header_info['column_map'])
        self.assertIn('col_static', header_info['column_id_map'])
        self.assertIn('col_amount', header_info['column_id_map'])
    
    def test_invoice_sheet_footer_bundles(self):
        """Test footer bundles for Invoice sheet."""
        resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Invoice',
            worksheet=self.worksheet,
            args=self.args,
            invoice_data=self.invoice_data,
            pallets=25
        )
        
        sum_ranges = [('E', 23, 30), ('F', 23, 30), ('G', 23, 30)]
        style_config, context_config, data_config = resolver.get_footer_bundles(
            sum_ranges=sum_ranges,
            pallet_count=25,
            is_last_table=True
        )
        
        # Check footer data
        self.assertEqual(data_config['sum_ranges'], sum_ranges)
        self.assertIn('footer_config', data_config)
        self.assertFalse(data_config['DAF_mode'])
        
        # Check context
        self.assertEqual(context_config['pallet_count'], 25)
        self.assertTrue(context_config['is_last_table'])
    
    # ========== Contract Sheet Tests ==========
    
    def test_contract_sheet_structure(self):
        """Test Contract sheet structure configuration."""
        sheet_config = self.config_loader.get_layout_config('Contract')
        structure = sheet_config.get('structure', {})
        
        self.assertEqual(structure['header_row'], 15)
        
        columns = structure.get('columns', [])
        self.assertEqual(len(columns), 5)
        
        # Check specific columns
        col_ids = [col['id'] for col in columns]
        self.assertIn('col_no', col_ids)
        self.assertIn('col_item', col_ids)
        self.assertIn('col_amount', col_ids)
    
    def test_contract_sheet_styling(self):
        """Test Contract sheet has different styling than Invoice."""
        contract_styling = self.config_loader.get_styling_config('Contract')
        invoice_styling = self.config_loader.get_styling_config('Invoice')
        
        # Contract should have larger font than Invoice
        self.assertEqual(contract_styling['header']['font']['size'], 16)
        self.assertEqual(invoice_styling['header']['font']['size'], 12)
        
        # Different row heights
        self.assertEqual(contract_styling['header']['row_height'], 36)
        self.assertEqual(invoice_styling['header']['row_height'], 35)
    
    def test_contract_sheet_footer_config(self):
        """Test Contract sheet footer configuration."""
        sheet_config = self.config_loader.get_layout_config('Contract')
        footer = sheet_config.get('footer', {})
        
        self.assertEqual(footer['total_column'], 0)
        self.assertIn('merge_cells', footer)
        self.assertEqual(footer['merge_cells']['start'], 0)
        self.assertEqual(footer['merge_cells']['span'], 2)
    
    # ========== Packing List Sheet Tests ==========
    
    def test_packing_list_structure(self):
        """Test Packing list sheet structure with complex headers."""
        sheet_config = self.config_loader.get_layout_config('Packing list')
        structure = sheet_config.get('structure', {})
        
        self.assertEqual(structure['header_row'], 21)
        
        columns = structure.get('columns', [])
        # 8 columns: static, po, item, desc, qty_header (with children), net, gross, cbm
        self.assertEqual(len(columns), 8)
        
        # Check for column with children (colspan)
        qty_header = next(col for col in columns if col['id'] == 'col_qty_header')
        self.assertIn('colspan', qty_header)
        self.assertEqual(qty_header['colspan'], 2)
        self.assertIn('children', qty_header)
        self.assertEqual(len(qty_header['children']), 2)
    
    def test_packing_list_styling(self):
        """Test Packing list sheet styling configuration."""
        styling = self.config_loader.get_styling_config('Packing list')
        
        # Check dimensions
        dimensions = styling.get('dimensions', {})
        col_widths = dimensions.get('column_widths', {})
        
        self.assertIn('col_static', col_widths)
        self.assertEqual(col_widths['col_static'], 24.71)
        self.assertEqual(col_widths['col_po'], 17)
        
        # Check column-specific styling
        col_specific = styling.get('column_specific', {})
        self.assertIn('col_static', col_specific)
        static_align = col_specific['col_static']['alignment']
        self.assertEqual(static_align['horizontal'], 'left')
        self.assertEqual(static_align['vertical'], 'top')
    
    def test_packing_list_data_flow(self):
        """Test Packing list sheet data flow mappings."""
        sheet_config = self.config_loader.get_layout_config('Packing list')
        mappings = sheet_config.get('data_flow', {}).get('mappings', {})
        
        # Packing list has more fields
        self.assertIn('po', mappings)
        self.assertIn('pcs', mappings)
        self.assertIn('net', mappings)
        self.assertIn('gross', mappings)
        self.assertIn('cbm', mappings)
    
    def test_packing_list_static_content(self):
        """Test Packing list sheet static content configuration."""
        sheet_config = self.config_loader.get_layout_config('Packing list')
        content = sheet_config.get('content', {})
        static = content.get('static', {})
        
        # Check static column content
        self.assertIn('col_static', static)
        static_values = static['col_static']
        self.assertIsInstance(static_values, list)
        self.assertIn('VENDOR#:', static_values)
        self.assertIn('MADE IN CAMBODIA', static_values)
        
        # Check before_footer content
        before_footer = static.get('before_footer', {})
        self.assertEqual(before_footer['col'], 2)
        self.assertIn('LEATHER', before_footer['text'])
        self.assertEqual(before_footer['merge'], 2)
    
    def test_packing_list_footer_config(self):
        """Test Packing list sheet footer configuration."""
        sheet_config = self.config_loader.get_layout_config('Packing list')
        footer = sheet_config.get('footer', {})
        
        self.assertEqual(footer['total_column'], 1)
        self.assertEqual(footer['pallet_column'], 2)
        self.assertTrue(footer.get('add_blank_before', False))
    
    def test_packing_list_multi_table_data(self):
        """Test Packing list resolver with multi-table data."""
        resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Packing list',
            worksheet=self.worksheet,
            args=self.args,
            invoice_data=self.invoice_data
        )
        
        # Get data for table 1
        style_config, context_config, layout_config, data_config = resolver.get_datatable_bundles(table_key='1')
        
        self.assertEqual(data_config['table_key'], '1')
        self.assertIsInstance(data_config['data_source'], list)
        self.assertEqual(len(data_config['data_source']), 2)
        
        # Get data for table 2
        style_config, context_config, layout_config, data_config = resolver.get_datatable_bundles(table_key='2')
        
        self.assertEqual(data_config['table_key'], '2')
        self.assertEqual(len(data_config['data_source']), 1)
    
    # ========== Cross-Sheet Tests ==========
    
    def test_all_sheets_have_required_sections(self):
        """Test that all sheets have required configuration sections."""
        sheets = self.config_loader.get_sheets_to_process()
        
        for sheet_name in sheets:
            with self.subTest(sheet=sheet_name):
                # Layout config
                layout = self.config_loader.get_layout_config(sheet_name)
                self.assertIn('structure', layout)
                self.assertIn('data_flow', layout)
                
                # Styling config
                styling = self.config_loader.get_styling_config(sheet_name)
                self.assertIn('header', styling)
                self.assertIn('data', styling)
                self.assertIn('footer', styling)
    
    def test_all_sheets_resolver_instantiation(self):
        """Test that resolver can be instantiated for all sheets."""
        sheets = self.config_loader.get_sheets_to_process()
        
        for sheet_name in sheets:
            with self.subTest(sheet=sheet_name):
                resolver = BuilderConfigResolver(
                    config_loader=self.config_loader,
                    sheet_name=sheet_name,
                    worksheet=self.worksheet,
                    args=self.args,
                    invoice_data=self.invoice_data
                )
                
                self.assertEqual(resolver.sheet_name, sheet_name)
                self.assertIsNotNone(resolver._sheet_config)
    
    def test_all_sheets_can_generate_bundles(self):
        """Test that all sheets can generate all bundle types."""
        sheets = self.config_loader.get_sheets_to_process()
        
        for sheet_name in sheets:
            with self.subTest(sheet=sheet_name):
                resolver = BuilderConfigResolver(
                    config_loader=self.config_loader,
                    sheet_name=sheet_name,
                    worksheet=self.worksheet,
                    args=self.args,
                    invoice_data=self.invoice_data
                )
                
                # Test all bundle types
                style_bundle = resolver.get_style_bundle()
                self.assertIn('styling_config', style_bundle)
                
                context_bundle = resolver.get_context_bundle()
                self.assertEqual(context_bundle['sheet_name'], sheet_name)
                
                layout_bundle = resolver.get_layout_bundle()
                self.assertIn('sheet_config', layout_bundle)
                
                data_bundle = resolver.get_data_bundle()
                self.assertIn('data_source', data_bundle)
                self.assertIn('header_info', data_bundle)
    
    # ========== Defaults Tests ==========
    
    def test_defaults_section(self):
        """Test that defaults section is accessible."""
        raw_config = self.config_loader.get_raw_config()
        defaults = raw_config.get('defaults', {})
        
        self.assertIn('footer', defaults)
        footer_defaults = defaults['footer']
        
        self.assertTrue(footer_defaults.get('show_total'))
        self.assertTrue(footer_defaults.get('show_pallet_count'))
        self.assertEqual(footer_defaults.get('total_text'), 'TOTAL:')
    
    # ========== Data Preparation Module Hint Tests ==========
    
    def test_data_preparation_hint(self):
        """Test that data preparation module hint is present."""
        raw_config = self.config_loader.get_raw_config()
        hint = raw_config.get('data_preparation_module_hint', {})
        
        self.assertIn('prority', hint)
        self.assertEqual(hint['prority'], ['po'])
        self.assertEqual(hint['numbers_per_group_by_po'], 7)
    
    # ========== Column Format Tests ==========
    
    def test_column_formats_extraction(self):
        """Test that column formats are correctly extracted."""
        resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Invoice',
            worksheet=self.worksheet,
            invoice_data=self.invoice_data
        )
        
        data_bundle = resolver.get_data_bundle()
        header_info = data_bundle['header_info']
        formats = header_info.get('column_formats', {})
        
        # Check specific formats from config
        self.assertEqual(formats.get('col_po'), '@')  # Text format
        self.assertEqual(formats.get('col_qty_sf'), '#,##0.00')  # Number format
        self.assertEqual(formats.get('col_amount'), '#,##0.00')


class TestJFConfigColumnMaps(unittest.TestCase):
    """Detailed tests for column map construction from JF config."""
    
    @classmethod
    def setUpClass(cls):
        """Load the real JF config once."""
        config_path = Path(__file__).parent.parent.parent / "invoice_generator" / "config_bundled" / "JF_config" / "JF_config.json"
        cls.config_loader = BundledConfigLoader(config_path)
    
    def setUp(self):
        """Set up test fixtures."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
    
    def test_invoice_column_map(self):
        """Test Invoice sheet column map construction."""
        resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Invoice',
            worksheet=self.worksheet
        )
        
        data_bundle = resolver.get_data_bundle()
        column_map = data_bundle['header_info']['column_map']
        
        # Check all expected headers
        expected_headers = {
            'Mark & Nº': 1,
            'P.O. Nº': 2,
            'ITEM Nº': 3,
            'Description': 4,
            'Quantity(SF)': 5,
            'Unit price (USD)': 6,
            'Amount (USD)': 7
        }
        
        for header, expected_col in expected_headers.items():
            self.assertEqual(column_map[header], expected_col, f"Header '{header}' column mismatch")
    
    def test_invoice_column_id_map(self):
        """Test Invoice sheet column ID map construction."""
        resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Invoice',
            worksheet=self.worksheet
        )
        
        data_bundle = resolver.get_data_bundle()
        column_id_map = data_bundle['header_info']['column_id_map']
        
        # Check all expected column IDs
        expected_ids = {
            'col_static': 1,
            'col_po': 2,
            'col_item': 3,
            'col_desc': 4,
            'col_qty_sf': 5,
            'col_unit_price': 6,
            'col_amount': 7
        }
        
        for col_id, expected_col in expected_ids.items():
            self.assertEqual(column_id_map[col_id], expected_col, f"Column ID '{col_id}' column mismatch")
    
    def test_packing_list_column_map_with_complex_headers(self):
        """Test Packing list column map handles complex headers with rowspan/colspan."""
        resolver = BuilderConfigResolver(
            config_loader=self.config_loader,
            sheet_name='Packing list',
            worksheet=self.worksheet
        )
        
        data_bundle = resolver.get_data_bundle()
        column_id_map = data_bundle['header_info']['column_id_map']
        
        # The resolver treats parent column as single entity
        # Children are not expanded in the column_id_map at the resolver level
        # This is correct - the HeaderBuilder will handle expanding children during actual header creation
        
        # Check that parent columns are indexed correctly
        self.assertIn('col_qty_header', column_id_map)
        
        # Verify column indices for top-level columns
        # Based on structure: static, po, item, desc, qty_header, net, gross, cbm
        self.assertEqual(column_id_map['col_static'], 1)
        self.assertEqual(column_id_map['col_po'], 2)
        self.assertEqual(column_id_map['col_item'], 3)
        self.assertEqual(column_id_map['col_desc'], 4)
        self.assertEqual(column_id_map['col_qty_header'], 5)
        self.assertEqual(column_id_map['col_net'], 6)
        self.assertEqual(column_id_map['col_gross'], 7)
        self.assertEqual(column_id_map['col_cbm'], 8)


if __name__ == '__main__':
    unittest.main()
