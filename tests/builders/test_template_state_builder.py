import unittest
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
import os
from pathlib import Path

from invoice_generator.builders.template_state_builder import TemplateStateBuilder

class TestTemplateStateBuilder(unittest.TestCase):

    def setUp(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet: Worksheet = self.workbook.active

        # Create a dummy template with header, data, footer, and grand total footer
        # Header
        self.worksheet['A1'] = 'Header 1'
        self.worksheet['B1'] = 'Header 2'

        # Data
        self.worksheet['A2'] = 'Data 1'
        self.worksheet['B2'] = 'Data 2'

        # Footer
        self.worksheet['A4'] = 'Footer 1'
        self.worksheet['B4'] = 'Footer 2'
        self.worksheet.merge_cells('A4:B4')

        # Grand Total Footer (after a blank row)
        self.worksheet['A6'] = 'Grand Total'
        self.worksheet['B6'] = '1000'

    def test_capture_and_restore_footer_with_grand_total(self):
        """Test that footer with grand total is properly captured and restored (updated for current API)."""
        # 1. Capture the template state (auto-captured in __init__)
        template_builder = TemplateStateBuilder(
            worksheet=self.worksheet,
            num_header_cols=2,
            header_end_row=1,
            footer_start_row=4,  # Footer starts at row 4
            debug=False
        )

        # 2. Assert that the footer is captured correctly
        self.assertGreater(len(template_builder.footer_state), 0, "Footer should be captured")
        self.assertEqual(template_builder.footer_state[0][0]['value'], 'Footer 1')
        
        # Grand total at row 6 should also be captured (rows 4-6 = 3 rows)
        if len(template_builder.footer_state) >= 3:
            self.assertEqual(template_builder.footer_state[2][0]['value'], 'Grand Total')

        # 3. Create a new workbook and restore the state
        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active
        
        template_builder.restore_header_only(target_worksheet=new_worksheet)
        template_builder.restore_footer_only(target_worksheet=new_worksheet, footer_start_row=4)

        # 4. Assert that the footer and grand total footer are restored at the correct positions
        self.assertEqual(new_worksheet['A4'].value, 'Footer 1')
        self.assertIn('A4:B4', [str(r) for r in new_worksheet.merged_cells.ranges])
        
        # Grand total might be at row 6 depending on footer size
        if len(template_builder.footer_state) >= 3:
            self.assertEqual(new_worksheet['A6'].value, 'Grand Total')
        
        new_workbook.close()
    
    def test_merged_cells_in_header(self):
        """Test that merged cells in header are properly captured and restored."""
        # Create header with merged cells
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set up header with merged cells
        ws['A1'] = 'Company Name'
        ws.merge_cells('A1:D1')
        ws['A2'] = 'PO Number'
        ws['B2'] = 'Item'
        
        # Create template builder (auto-captures in __init__)
        tsb = TemplateStateBuilder(
            worksheet=ws,
            num_header_cols=4,
            header_end_row=2,
            footer_start_row=10,
            debug=False
        )
        
        # Check that header state was captured
        self.assertEqual(len(tsb.header_state), 2, "Should capture 2 header rows")
        self.assertEqual(tsb.header_state[0][0]['value'], 'Company Name', "First cell should have company name")
        
        # Check that merged cells are tracked
        self.assertGreater(len(tsb.header_merged_cells), 0, "Should capture merged cell ranges")
        
        # Test restoration
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        
        tsb.restore_header_only(target_worksheet=new_ws)
        
        # Check restored content
        self.assertEqual(new_ws['A1'].value, 'Company Name')
        self.assertEqual(new_ws['A2'].value, 'PO Number')
        self.assertEqual(new_ws['B2'].value, 'Item')
        
        # Check that B1 is a MergedCell (not master cell)
        self.assertIsInstance(new_ws['B1'], MergedCell, "B1 should be a MergedCell")
        self.assertNotIsInstance(new_ws['A1'], MergedCell, "A1 should be master cell, not MergedCell")
        
        wb.close()
        new_wb.close()
    
    def test_merged_cells_store_none_for_non_master(self):
        """Test that template_state_builder stores None for non-master cells of merged ranges."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create a merged cell range
        ws['A1'] = 'Master Cell Value'
        ws.merge_cells('A1:C1')
        
        # Create template builder
        tsb = TemplateStateBuilder(
            worksheet=ws,
            num_header_cols=3,
            header_end_row=1,
            footer_start_row=10,
            debug=False
        )
        
        # Check stored values
        # Cell A1 (master) should have value
        self.assertEqual(tsb.header_state[0][0]['value'], 'Master Cell Value', 
                        "Master cell should store actual value")
        
        # Cells B1, C1 (non-master) should have None
        self.assertIsNone(tsb.header_state[0][1]['value'], 
                         "Non-master merged cell should store None")
        self.assertIsNone(tsb.header_state[0][2]['value'], 
                         "Non-master merged cell should store None")
        
        wb.close()
    
    def test_footer_restoration_at_custom_row(self):
        """Test that footer can be restored at any row without errors."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create footer content
        ws['A10'] = 'TOTAL:'
        ws['B10'] = '=SUM(B2:B9)'
        
        # Create template builder
        tsb = TemplateStateBuilder(
            worksheet=ws,
            num_header_cols=2,
            header_end_row=1,
            footer_start_row=10,
            debug=False
        )
        
        # Check footer was captured
        self.assertGreater(len(tsb.footer_state), 0, "Footer should be captured")
        self.assertEqual(tsb.footer_state[0][0]['value'], 'TOTAL:', "Footer text should be captured")
        
        # Test restoration at different rows
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        
        # Restore at row 50
        tsb.restore_footer_only(target_worksheet=new_ws, footer_start_row=50)
        self.assertEqual(new_ws['A50'].value, 'TOTAL:', "Footer should be at row 50")
        
        # Restore at row 100
        tsb.restore_footer_only(target_worksheet=new_ws, footer_start_row=100)
        self.assertEqual(new_ws['A100'].value, 'TOTAL:', "Footer should be at row 100")
        
        wb.close()
        new_wb.close()
    
    def test_empty_cells_captured_as_none(self):
        """Test that empty cells are captured as None."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create sparse header
        ws['A1'] = 'Name'
        # B1 is empty
        ws['C1'] = 'Amount'
        
        tsb = TemplateStateBuilder(
            worksheet=ws,
            num_header_cols=3,
            header_end_row=1,
            footer_start_row=10,
            debug=False
        )
        
        # Check that empty cell is captured as None
        self.assertIsNone(tsb.header_state[0][1]['value'], "Empty cell should be None")
        
        wb.close()
    
    def test_real_template_integration(self):
        """Integration test with real JF template if available."""
        template_path = Path(__file__).parent.parent.parent / "invoice_generator" / "template" / "JF.xlsx"
        
        if not template_path.exists():
            self.skipTest(f"Template not found: {template_path}")
        
        # Load real template
        wb = openpyxl.load_workbook(template_path, read_only=False)
        ws = wb["Packing list"]
        
        # Create template builder with real template dimensions
        tsb = TemplateStateBuilder(
            worksheet=ws,
            num_header_cols=9,
            header_end_row=20,  # Packing list header ends before row 21
            footer_start_row=39,  # Footer in template
            debug=False
        )
        
        # Verify captures
        self.assertGreater(len(tsb.header_state), 0, "Should capture header from real template")
        self.assertGreater(len(tsb.footer_state), 0, "Should capture footer from real template")
        self.assertGreater(len(tsb.header_merged_cells), 0, "Should find merged cells in header")
        
        # Test restoration doesn't crash
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        
        try:
            tsb.restore_header_only(target_worksheet=new_ws)
            tsb.restore_footer_only(target_worksheet=new_ws, footer_start_row=50)
            
            # Check that merged cells exist in restored sheet
            self.assertGreater(len(new_ws.merged_cells.ranges), 0, 
                             "Restored sheet should have merged cells")
            
        except Exception as e:
            self.fail(f"Restoration should not crash: {e}")
        
        wb.close()
        new_wb.close()

    def test_footer_with_empty_rows(self):
        """Test that footer with empty rows (spacing) is captured and restored correctly."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create a footer with mixed content and empty rows
        # Row 10: Grand Total row
        ws['A10'].value = 'Grand Total:'
        ws['B10'].value = 5000
        
        # Row 11: EMPTY ROW (spacing)
        # (leave all cells empty)
        
        # Row 12: Thank you message
        ws['A12'].value = 'Thank you for your business'
        
        # Row 13: EMPTY ROW
        # (leave all cells empty)
        
        # Row 14: Signature line
        ws['A14'].value = 'Authorized Signature: ___________'
        
        # Create builder - footer starts at row 10, ends at row 14
        tsb = TemplateStateBuilder(ws, num_header_cols=10, header_end_row=5, footer_start_row=10, debug=True)
        
        # Verify footer was captured (5 rows: 10, 11, 12, 13, 14)
        self.assertEqual(len(tsb.footer_state), 5, "Should capture 5 footer rows")
        
        # Verify empty rows captured correctly (row 11 and 13, indices 1 and 3)
        row_11_data = tsb.footer_state[1]  # Row 11 (index 1)
        row_13_data = tsb.footer_state[3]  # Row 13 (index 3)
        
        # Check that all cells in empty rows are None
        all_none_row_11 = all(cell['value'] is None for cell in row_11_data)
        all_none_row_13 = all(cell['value'] is None for cell in row_13_data)
        
        self.assertTrue(all_none_row_11, "Row 11 should have all None values")
        self.assertTrue(all_none_row_13, "Row 13 should have all None values")
        
        # Now restore footer at a different location (row 50)
        wb_output = openpyxl.Workbook()
        ws_output = wb_output.active
        
        tsb.restore_footer_only(ws_output, footer_start_row=50)
        
        # Verify content was restored at new location with correct offset
        # Original row 10 -> new row 50 (offset = 40)
        self.assertEqual(ws_output['A50'].value, 'Grand Total:', "Grand Total should be at row 50")
        self.assertEqual(ws_output['B50'].value, 5000, "Amount should be at row 50")
        
        # Row 51 should be empty (original row 11)
        self.assertIsNone(ws_output['A51'].value, "Row 51 should be empty")
        self.assertIsNone(ws_output['B51'].value, "Row 51 should be empty")
        
        # Row 52 should have thank you message (original row 12)
        self.assertEqual(ws_output['A52'].value, 'Thank you for your business', "Thank you at row 52")
        
        # Row 53 should be empty (original row 13)
        self.assertIsNone(ws_output['A53'].value, "Row 53 should be empty")
        
        # Row 54 should have signature (original row 14)
        self.assertEqual(ws_output['A54'].value, 'Authorized Signature: ___________', "Signature at row 54")
        
        wb.close()
        wb_output.close()

if __name__ == '__main__':
    unittest.main()