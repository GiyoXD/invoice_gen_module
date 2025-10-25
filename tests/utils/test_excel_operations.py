import unittest
from openpyxl import Workbook
from invoice_generator.utils.layout import unmerge_row

class TestExcelOperations(unittest.TestCase):

    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def test_unmerge_row(self):
        # Merge cells for testing
        self.ws.merge_cells('A1:C1')
        self.ws.merge_cells('D2:E3')

        # Test unmerging a row that intersects with a merge
        unmerge_row(self.ws, 1, 5)
        self.assertNotIn('A1:C1', [str(r) for r in self.ws.merged_cells.ranges])

        # Test unmerging a row that is inside a multi-row merge
        unmerge_row(self.ws, 2, 5)
        self.assertNotIn('D2:E3', [str(r) for r in self.ws.merged_cells.ranges])

if __name__ == '__main__':
    unittest.main()
