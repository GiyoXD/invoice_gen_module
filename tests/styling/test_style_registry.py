"""
Test StyleRegistry and CellStyler - ID-driven styling system
"""

import pytest
from invoice_generator.styling.style_registry import StyleRegistry, ColumnStyle, RowContextStyle
from invoice_generator.styling.cell_styler import CellStyler
from openpyxl import Workbook


def test_column_style_creation():
    """Test ColumnStyle dataclass."""
    col = ColumnStyle(
        col_id='col_cbm',
        format='0.00',
        alignment='center',
        width=12
    )
    
    assert col.col_id == 'col_cbm'
    assert col.format == '0.00'
    assert col.alignment == 'center'
    assert col.width == 12
    
    col_dict = col.to_dict()
    assert col_dict['format'] == '0.00'
    assert col_dict['alignment'] == 'center'


def test_row_context_style_creation():
    """Test RowContextStyle dataclass."""
    context = RowContextStyle(
        context='header',
        bold=True,
        fill_color='CCCCCC',
        font_size=12
    )
    
    assert context.context == 'header'
    assert context.bold is True
    assert context.fill_color == 'CCCCCC'
    
    context_dict = context.to_dict()
    assert context_dict['bold'] is True
    assert context_dict['fill_color'] == 'CCCCCC'


def test_style_registry_initialization():
    """Test StyleRegistry loads columns and contexts."""
    config = {
        'columns': {
            'col_po': {
                'format': '@',
                'alignment': 'center',
                'width': 28
            },
            'col_cbm': {
                'format': '0.00',
                'alignment': 'center',
                'width': 12
            }
        },
        'row_contexts': {
            'header': {
                'bold': True,
                'fill_color': 'CCCCCC',
                'font_size': 12
            },
            'data': {
                'bold': False,
                'font_size': 11
            },
            'footer': {
                'bold': True,
                'fill_color': 'FFFFCC'
            }
        }
    }
    
    registry = StyleRegistry(config)
    
    # Check columns loaded
    assert registry.has_column('col_po')
    assert registry.has_column('col_cbm')
    assert len(registry.columns) == 2
    
    # Check contexts loaded
    assert registry.has_context('header')
    assert registry.has_context('data')
    assert registry.has_context('footer')
    assert len(registry.row_contexts) == 3


def test_style_merging():
    """Test column + context style merging."""
    config = {
        'columns': {
            'col_cbm': {
                'format': '0.00',
                'alignment': 'center',
                'width': 12
            }
        },
        'row_contexts': {
            'header': {
                'bold': True,
                'fill_color': 'CCCCCC'
            },
            'data': {
                'bold': False
            }
        }
    }
    
    registry = StyleRegistry(config)
    
    # Get header style
    header_style = registry.get_style('col_cbm', context='header')
    assert header_style['format'] == '0.00'  # From column
    assert header_style['alignment'] == 'center'  # From column
    assert header_style['bold'] is True  # From context
    assert header_style['fill_color'] == 'CCCCCC'  # From context
    
    # Get data style
    data_style = registry.get_style('col_cbm', context='data')
    assert data_style['format'] == '0.00'  # From column
    assert data_style['alignment'] == 'center'  # From column
    assert data_style['bold'] is False  # From context
    assert 'fill_color' not in data_style  # Not in context


def test_column_owned_properties_not_overridden():
    """Test that column-owned properties (format, alignment, width, wrap_text) are NEVER overridden by context."""
    config = {
        'columns': {
            'col_desc': {
                'format': '@',
                'alignment': 'left',  # Column says left-align
                'width': 20,
                'wrap_text': True
            }
        },
        'row_contexts': {
            'header': {
                'bold': True,
                'alignment': 'center',  # Context tries to override - should be IGNORED
                'format': '0.00',        # Context tries to override - should be IGNORED
                'fill_color': 'CCCCCC'
            }
        }
    }
    
    registry = StyleRegistry(config)
    style = registry.get_style('col_desc', context='header')
    
    # Column-owned properties should NOT be overridden
    assert style['format'] == '@', "Column format should not be overridden by context"
    assert style['alignment'] == 'left', "Column alignment should not be overridden by context"
    assert style['width'] == 20, "Column width should not be overridden by context"
    assert style['wrap_text'] is True, "Column wrap_text should not be overridden by context"
    
    # Context properties should be added
    assert style['bold'] is True, "Context bold should be applied"
    assert style['fill_color'] == 'CCCCCC', "Context fill_color should be applied"



def test_style_overrides():
    """Test style overrides for special cases."""
    config = {
        'columns': {
            'col_total': {
                'format': '#,##0.00',
                'alignment': 'right'
            }
        },
        'row_contexts': {
            'footer': {
                'bold': True
            }
        }
    }
    
    registry = StyleRegistry(config)
    
    # Normal footer style
    normal_style = registry.get_style('col_total', context='footer')
    assert normal_style['bold'] is True
    assert normal_style['alignment'] == 'right'
    
    # Grand total with override
    grand_total_style = registry.get_style(
        'col_total',
        context='footer',
        overrides={'fill_color': 'FFFF00', 'font_size': 14}
    )
    assert grand_total_style['bold'] is True  # From context
    assert grand_total_style['alignment'] == 'right'  # From column
    assert grand_total_style['fill_color'] == 'FFFF00'  # Override
    assert grand_total_style['font_size'] == 14  # Override


def test_cell_styler_font_application():
    """Test CellStyler applies font properties."""
    wb = Workbook()
    ws = wb.active
    cell = ws['A1']
    
    styler = CellStyler()
    style = {
        'bold': True,
        'font_size': 14,
        'font_name': 'Arial'
    }
    
    styler.apply(cell, style)
    
    assert cell.font.bold is True
    assert cell.font.size == 14
    assert cell.font.name == 'Arial'


def test_cell_styler_alignment_application():
    """Test CellStyler applies alignment."""
    wb = Workbook()
    ws = wb.active
    cell = ws['A1']
    
    styler = CellStyler()
    style = {
        'alignment': 'center',
        'wrap_text': True
    }
    
    styler.apply(cell, style)
    
    assert cell.alignment.horizontal == 'center'
    assert cell.alignment.wrap_text is True


def test_cell_styler_fill_application():
    """Test CellStyler applies fill color."""
    wb = Workbook()
    ws = wb.active
    cell = ws['A1']
    
    styler = CellStyler()
    style = {
        'fill_color': 'CCCCCC'
    }
    
    styler.apply(cell, style)
    
    # openpyxl prepends '00' for alpha transparency
    assert cell.fill.start_color.rgb == '00CCCCCC'
    assert cell.fill.fill_type == 'solid'


def test_cell_styler_number_format():
    """Test CellStyler applies number format."""
    wb = Workbook()
    ws = wb.active
    cell = ws['A1']
    
    styler = CellStyler()
    style = {
        'format': '0.00'
    }
    
    styler.apply(cell, style)
    
    assert cell.number_format == '0.00'


def test_complete_workflow():
    """Test complete workflow: Registry → Style → Cell."""
    # Setup config
    config = {
        'columns': {
            'col_cbm': {
                'format': '0.00',
                'alignment': 'center',
                'width': 12
            }
        },
        'row_contexts': {
            'header': {
                'bold': True,
                'fill_color': 'CCCCCC',
                'font_size': 12
            }
        }
    }
    
    # Create registry and styler
    registry = StyleRegistry(config)
    styler = CellStyler()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    cell = ws['A1']
    
    # Get style and apply
    style = registry.get_style('col_cbm', context='header')
    styler.apply(cell, style)
    
    # Verify all properties applied
    assert cell.number_format == '0.00'  # From column
    assert cell.alignment.horizontal == 'center'  # From column
    assert cell.font.bold is True  # From context
    assert cell.font.size == 12  # From context
    # openpyxl prepends '00' for alpha transparency
    assert cell.fill.start_color.rgb == '00CCCCCC'  # From context


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
