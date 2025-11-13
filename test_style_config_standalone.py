"""
Standalone test to verify config transformation and style application work correctly.
Tests the entire pipeline: JSON config → transformation → StylingConfigModel → Font application
"""
import sys
from pathlib import Path

# Add invoice_generator to path
sys.path.insert(0, str(Path(__file__).parent / 'invoice_generator'))

from openpyxl import Workbook
from openpyxl.styles import Font
from invoice_generator.config.config_loader import BundledConfigLoader
from invoice_generator.styling.models import StylingConfigModel
from invoice_generator.styling.style_applier import apply_header_style

def test_config_transformation():
    """Test that config is loaded and transformed correctly"""
    print("=" * 60)
    print("TEST 1: Config Transformation")
    print("=" * 60)
    
    config_path = Path('invoice_generator/config_bundled/JF_config/JF_config.json')
    loader = BundledConfigLoader(config_path)
    
    # Get transformed styling config for Invoice sheet
    styling_dict = loader.get_styling_config('Invoice')
    
    print(f"\n[OK] Config loaded from: {config_path}")
    print(f"ok Transformed styling config keys: {list(styling_dict.keys())}")
    
    # Check critical properties
    if 'header_font' in styling_dict:
        print(f"ok header_font present: {styling_dict['header_font']}")
    else:
        print(f"[NOT OK] MISSING header_font!")
        return False
    
    if 'default_font' in styling_dict:
        print(f"ok default_font present: {styling_dict['default_font']}")
    else:
        print(f"[NOT OK] MISSING default_font!")
        return False
    
    return True

def test_styling_model_creation():
    """Test that StylingConfigModel can be created from transformed config"""
    print("\n" + "=" * 60)
    print("TEST 2: StylingConfigModel Creation")
    print("=" * 60)
    
    config_path = Path('invoice_generator/config_bundled/JF_config/JF_config.json')
    loader = BundledConfigLoader(config_path)
    styling_dict = loader.get_styling_config('Invoice')
    
    try:
        styling_model = StylingConfigModel(**styling_dict)
        print(f"\nok StylingConfigModel created successfully")
        
        # Check headerFont
        if styling_model.headerFont:
            font_dict = styling_model.headerFont.model_dump()
            print(f"[OK] headerFont exists:")
            print(f"    name: {font_dict.get('name')}")
            print(f"    size: {font_dict.get('size')}")
            print(f"    bold: {font_dict.get('bold')}")
        else:
            print(f"[NOT OK] headerFont is None!")
            return False, None
        
        # Check defaultFont
        if styling_model.defaultFont:
            font_dict = styling_model.defaultFont.model_dump()
            print(f"[OK] defaultFont exists:")
            print(f"    name: {font_dict.get('name')}")
            print(f"    size: {font_dict.get('size')}")
        else:
            print(f"[NOT OK] defaultFont is None!")
            return False, None
        
        return True, styling_model
    except Exception as e:
        print(f"[NOT OK] Failed to create StylingConfigModel: {e}")
        return False, None

def test_font_application():
    """Test that fonts are actually applied to cells"""
    print("\n" + "=" * 60)
    print("TEST 3: Font Application to Cell")
    print("=" * 60)
    
    # Get styling model
    config_path = Path('invoice_generator/config_bundled/JF_config/JF_config.json')
    loader = BundledConfigLoader(config_path)
    styling_dict = loader.get_styling_config('Invoice')
    styling_model = StylingConfigModel(**styling_dict)
    
    # Create a workbook and apply styling
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    
    # Create a cell and apply header style
    cell = ws['A1']
    cell.value = "Test Header"
    
    print(f"\n[OK] Created test workbook and cell A1")
    print(f"  Before styling - Font: name={cell.font.name}, size={cell.font.size}, bold={cell.font.bold}")
    
    # Apply header style
    apply_header_style(cell, styling_model)
    
    print(f"  After apply_header_style - Font: name={cell.font.name}, size={cell.font.size}, bold={cell.font.bold}")
    
    # Verify font was applied
    if cell.font.name != 'Times New Roman':
        print(f"[NOT OK] FAILED: Font name is '{cell.font.name}', expected 'Times New Roman'")
        return False
    
    if cell.font.size != 12:
        print(f"[NOT OK] FAILED: Font size is {cell.font.size}, expected 12")
        return False
    
    if cell.font.bold != True:
        print(f"[NOT OK] FAILED: Font bold is {cell.font.bold}, expected True")
        return False
    
    print(f"[OK] Font correctly applied: Times New Roman, size 12, bold")
    
    return True

def test_save_and_verify():
    """Test that fonts persist through save/load cycle"""
    print("\n" + "=" * 60)
    print("TEST 4: Save and Verify Font Persistence")
    print("=" * 60)
    
    # Get styling model
    config_path = Path('invoice_generator/config_bundled/JF_config/JF_config.json')
    loader = BundledConfigLoader(config_path)
    styling_dict = loader.get_styling_config('Invoice')
    styling_model = StylingConfigModel(**styling_dict)
    
    # Create workbook with styled cells
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    
    # Apply styles to multiple cells
    for row in range(1, 4):
        cell = ws.cell(row=row, column=1)
        cell.value = f"Header {row}"
        apply_header_style(cell, styling_model)
    
    print(f"\n[OK] Created workbook with 3 styled header cells")
    
    # Save
    test_file = Path('test_font_standalone.xlsx')
    wb.save(test_file)
    print(f"[OK]OK] Saved to: {test_file}")
    
    # Verify XML contains Times New Roman
    import zipfile
    import xml.etree.ElementTree as ET
    
    with zipfile.ZipFile(test_file, 'r') as zip_ref:
        styles_xml = zip_ref.read('xl/styles.xml').decode('utf-8')
        root = ET.fromstring(styles_xml)
        
        # Find fonts
        ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        fonts = root.find('.//fonts', ns)
        
        times_new_roman_found = False
        if fonts:
            for font in fonts.findall('font', ns):
                name = font.find('name', ns)
                size = font.find('sz', ns)
                bold = font.find('b', ns)
                
                if name is not None and name.get('val') == 'Times New Roman':
                    if size is not None and float(size.get('val')) == 12.0:
                        if bold is not None:
                            times_new_roman_found = True
                            print(f"[OK] Found in XML: Times New Roman, size 12, bold")
                            break
        
        if not times_new_roman_found:
            print(f"[NOT OK] FAILED: Times New Roman font not found in saved Excel file")
            return False
    
    print(f"[OK] Font successfully persisted to Excel file")
    print(f"\n📊 Open '{test_file}' in Excel to verify Times New Roman is displayed")
    
    return True

def main():
    """Run all tests"""
    print("\n" + "🧪" * 30)
    print("STANDALONE CONFIG & STYLING TEST")
    print("🧪" * 30)
    
    results = []
    
    # Test 1: Config transformation
    results.append(("Config Transformation", test_config_transformation()))
    
    # Test 2: StylingConfigModel creation
    success, _ = test_styling_model_creation()
    results.append(("StylingConfigModel Creation", success))
    
    # Test 3: Font application
    results.append(("Font Application", test_font_application()))
    
    # Test 4: Save and verify
    results.append(("Save & Verify", test_save_and_verify()))
    
    # Summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)
    
    all_passed = True
    for test_name, passed in results:
        status = "✅ PASS" if passed else "❌ FAIL"
        print(f"{status} - {test_name}")
        if not passed:
            all_passed = False
    
    if all_passed:
        print("\n🎉 ALL TESTS PASSED! Config system is working correctly.")
        print("   Your config changes ARE taking effect.")
        print("   Open 'test_font_standalone.xlsx' in Excel to see Times New Roman.")
    else:
        print("\n⚠️  SOME TESTS FAILED. Check output above for details.")
    
    return 0 if all_passed else 1

if __name__ == "__main__":
    sys.exit(main())
