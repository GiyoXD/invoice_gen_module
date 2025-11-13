from openpyxl import load_workbook

wb = load_workbook('result_test2.xlsx')
ws = wb['Invoice']

print("Checking cell A21:")
cell = ws['A21']
print(f"  Value: {cell.value}")
print(f"  Font name: {cell.font.name}")
print(f"  Font size: {cell.font.size}")
print(f"  Font bold: {cell.font.bold}")

print("\nChecking if A21 is in a merged cell:")
from openpyxl.worksheet.cell_range import CellRange
for merged in ws.merged_cells.ranges:
    if 'A21' in merged:
        print(f"  YES - A21 is part of merge: {merged}")
        break
else:
    print("  NO - A21 is not merged")

print("\nChecking cells B21, C21:")
for addr in ['B21', 'C21']:
    cell = ws[addr]
    print(f"{addr}: value={cell.value}, font={cell.font.name}")
