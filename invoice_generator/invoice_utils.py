from pickle import NONE
import openpyxl
import re
import traceback
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, NamedStyle
from openpyxl.utils import column_index_from_string, get_column_letter
from typing import List, Dict, Any, Optional, Tuple, Union
from decimal import Decimal
from decimal import Decimal, InvalidOperation

from .data.data_preparer import prepare_data_rows, parse_mapping_rules
from .styling.style_applier import apply_cell_style
from .utils.layout import unmerge_row, unmerge_block, safe_unmerge_block, apply_column_widths, apply_row_heights, calculate_header_dimensions

# --- Constants for Styling ---
thin_side = Side(border_style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side) # Full grid border
no_border = Border(left=None, right=None, top=None, bottom=None)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
bold_font = Font(bold=True)

# --- Constants for Number Formats ---
FORMAT_GENERAL = 'General'
FORMAT_TEXT = '@'
FORMAT_NUMBER_COMMA_SEPARATED1 = '#,##0'
FORMAT_NUMBER_COMMA_SEPARATED2 = '#,##0.00'

# --- Utility Functions ---





























def write_footer_row(
    worksheet: Worksheet,
    footer_row_num: int,
    header_info: Dict[str, Any],
    sum_ranges: List[Tuple[int, int]],
    footer_config: Dict[str, Any],
    pallet_count: int,
    override_total_text: Optional[str] = None,
    DAF_mode: bool = False,
    grand_total_flag: bool = False,
    sheet_styling_config: Optional[Dict[str, Any]] = None
) -> int:
    """
    Writes a fully configured footer row, including styling, borders, merges,
    summed totals with number formatting, and a pallet count summary.

    This function is driven entirely by a footer configuration object and can sum
    over multiple, non-contiguous data ranges.

    Args:
        worksheet: The openpyxl worksheet to modify.
        footer_row_num: The 1-based row index for the footer.
        header_info: The header dictionary containing 'column_id_map' and 'num_columns'.
        sum_ranges: A list of tuples, where each tuple is a (start_row, end_row) of data to sum.
        footer_config: The footer configuration object from the JSON. The start_column_id field now supports:
                      - Column ID strings: {"start_column_id": "col_po", "colspan": 3}
                      - Raw column indices: {"start_column_id": 2, "colspan": 3}
                      - Raw indices as strings: {"start_column_id": "2", "colspan": 3}
        pallet_count: The total number of pallets to display in the footer.
        override_total_text: Optional text to use instead of the one in the config.
        DAF_mode: Whether DAF mode is enabled.
        grand_total_flag: Whether this is a grand total footer.

    Returns:
        The row index (footer_row_num) on success, or -1 on failure.
    """
    if not footer_config or footer_row_num <= 0:
        return -1

    try:
        # --- 1. Parse Configs and Prepare Style Objects ---
        num_columns = header_info.get('num_columns', 1)
        column_map_by_id = header_info.get('column_id_map', {})

        # Get style configurations with sensible defaults
        style_config = footer_config.get('style', {})
        font_config = style_config.get('font', {'bold': True})
        align_config = style_config.get('alignment', {'horizontal': 'center', 'vertical': 'center'})
        border_config = style_config.get('border', {'apply': True})
        
        # Get number format configuration
        number_format_config = footer_config.get("number_formats", {})

        # Create openpyxl style objects
        font_to_apply = Font(**font_config)
        align_to_apply = Alignment(**align_config)
        border_to_apply = None
        if border_config.get('apply'):
            side = Side(border_style=border_config.get('style', 'thin'), color=border_config.get('color', '000000'))
            border_to_apply = Border(left=side, right=side, top=side, bottom=side)

        unmerge_row(worksheet, footer_row_num, num_columns)

        # --- 2. Write Content (Labels, Formulas, and Pallet Count) ---
        total_text = override_total_text if override_total_text is not None else footer_config.get("total_text", "TOTAL:")
        total_text_col_id = footer_config.get("total_text_column_id")
        
        # Enhanced total_text_column_id: supports both column IDs and raw indices (0-based)
        total_text_col_idx = None
        if total_text_col_id is not None:
            if isinstance(total_text_col_id, int):
                # Raw column index (0-based, like programming arrays)
                # Convert to Excel's 1-based indexing by adding 1
                total_text_col_idx = total_text_col_id + 1
                print(f"Using raw column index {total_text_col_id} (0-based) for total text -> Excel column {total_text_col_idx} (1-based)")
            elif isinstance(total_text_col_id, str):
                # Try to parse as integer first (raw index as string)
                try:
                    raw_index = int(total_text_col_id)
                    # Convert to Excel's 1-based indexing by adding 1
                    total_text_col_idx = raw_index + 1
                    print(f"Using raw column index '{total_text_col_id}' (0-based string) for total text -> Excel column {total_text_col_idx} (1-based)")
                except ValueError:
                    # Not a number, treat as column ID and look up in map
                    total_text_col_idx = column_map_by_id.get(total_text_col_id)
                    print(f"Using column ID '{total_text_col_id}' for total text -> Excel column {total_text_col_idx} (1-based)")
        
        if total_text_col_idx:
            cell = worksheet.cell(row=footer_row_num, column=total_text_col_idx, value=total_text)
            cell.font = font_to_apply
            cell.alignment = align_to_apply

        # Write Pallet Count Text
        pallet_col_id = footer_config.get("pallet_count_column_id")
        
        # Enhanced pallet_count_column_id: supports both column IDs and raw indices (0-based)
        pallet_col_idx = None
        if pallet_col_id is not None and pallet_count > 0:
            if isinstance(pallet_col_id, int):
                # Raw column index (0-based, like programming arrays)
                # Convert to Excel's 1-based indexing by adding 1
                pallet_col_idx = pallet_col_id + 1
                print(f"Using raw column index {pallet_col_id} (0-based) for pallet count -> Excel column {pallet_col_idx} (1-based)")
            elif isinstance(pallet_col_id, str):
                # Try to parse as integer first (raw index as string)
                try:
                    raw_index = int(pallet_col_id)
                    # Convert to Excel's 1-based indexing by adding 1
                    pallet_col_idx = raw_index + 1
                    print(f"Using raw column index '{pallet_col_id}' (0-based string) for pallet count -> Excel column {pallet_col_idx} (1-based)")
                except ValueError:
                    # Not a number, treat as column ID and look up in map
                    pallet_col_idx = column_map_by_id.get(pallet_col_id)
                    print(f"Using column ID '{pallet_col_id}' for pallet count -> Excel column {pallet_col_idx} (1-based)")
        
        if pallet_col_idx:
            pallet_text = f"{pallet_count} PALLET{'S' if pallet_count != 1 else ''}"
            cell = worksheet.cell(row=footer_row_num, column=pallet_col_idx, value=pallet_text)
            cell.font = font_to_apply
            cell.alignment = align_to_apply

        sum_column_ids = footer_config.get("sum_column_ids", [])
        if sum_ranges:
            for col_id in sum_column_ids:
                col_idx = column_map_by_id.get(col_id)
                if col_idx:
                    col_letter = get_column_letter(col_idx)
                    sum_parts = [f"{col_letter}{start}:{col_letter}{end}" for start, end in sum_ranges]
                    formula = f"=SUM({','.join(sum_parts)})"
                    cell = worksheet.cell(row=footer_row_num, column=col_idx, value=formula)
                    cell.font = font_to_apply
                    cell.alignment = align_to_apply
                    
                    # Apply Number Formatting from Config if DAF
                    number_format_str = number_format_config.get(col_id)
                    if number_format_str and DAF_mode and col_id not in ['col_pcs', 'col_qty_pcs']:
                        cell.number_format = "##,00.00"
                    elif number_format_str:
                        cell.number_format = number_format_str["number_format"]
        # --- 3. Apply Border and Final Styling to the Whole Row ---
        if grand_total_flag != True:
            for c_idx in range(1, num_columns + 1):
                cell = worksheet.cell(row=footer_row_num, column=c_idx)
                if cell.font != font_to_apply: cell.font = font_to_apply
                if cell.alignment != align_to_apply: cell.alignment = align_to_apply
                if border_to_apply:
                    cell.border = border_to_apply

        # --- 4. Apply Merges ---
        merge_rules = footer_config.get("merge_rules", [])
        for rule in merge_rules:
            start_column_id = rule.get("start_column_id")
            colspan = rule.get("colspan")
            
            resolved_start_col = None
            
            # Enhanced start_column_id: supports both column IDs and raw indices
            if start_column_id is not None:
                if isinstance(start_column_id, int):
                    # Raw column index (0-based, like programming arrays)
                    # Convert to Excel's 1-based indexing by adding 1
                    resolved_start_col = start_column_id + 1
                    print(f"Using raw column index {start_column_id} (0-based) -> Excel column {resolved_start_col} (1-based)")
                elif isinstance(start_column_id, str):
                    # Try to parse as integer first (raw index as string)
                    try:
                        raw_index = int(start_column_id)
                        # Convert to Excel's 1-based indexing by adding 1
                        resolved_start_col = raw_index + 1
                        print(f"Using raw column index '{start_column_id}' (0-based string) -> Excel column {resolved_start_col} (1-based)")
                    except ValueError:
                        # Not a number, treat as column ID and look up in map
                        resolved_start_col = column_map_by_id.get(start_column_id)
                        print(f"Using column ID '{start_column_id}' -> Excel column {resolved_start_col} (1-based)")
            
            if resolved_start_col and colspan:
                end_col = min(resolved_start_col + colspan - 1, num_columns)
                worksheet.merge_cells(start_row=footer_row_num, start_column=resolved_start_col, end_row=footer_row_num, end_column=end_col)
        # --- 5. Apply Per-Column Alignment and Font from column_id_styles (if available) ---
        if sheet_styling_config:
            column_id_styles = sheet_styling_config.get("column_id_styles", {})
            idx_to_id_map = {v: k for k, v in column_map_by_id.items()}
            for c_idx in range(1, num_columns + 1):
                cell = worksheet.cell(row=footer_row_num, column=c_idx)
                column_id = idx_to_id_map.get(c_idx)
                if column_id and column_id in column_id_styles:
                    col_style = column_id_styles[column_id]
                    # Apply alignment if specified
                    if 'alignment' in col_style:
                        cell.alignment = Alignment(**{k: v for k, v in col_style['alignment'].items() if v is not None})
                    # Apply font if specified
                    if 'font' in col_style:
                        cell.font = Font(**{k: v for k, v in col_style['font'].items() if v is not None})

        return footer_row_num

    except Exception as e:
        print(f"ERROR: An error occurred during footer generation on row {footer_row_num}: {e}")
        return -1

    except Exception as e:
        print(f"ERROR: An error occurred during footer generation on row {footer_row_num}: {e}")
        # On failure, return -1
        return -1






def fill_invoice_data(
    worksheet: Worksheet,
    sheet_name: str,
    sheet_config: Dict[str, Any], # Keep current sheet config param
    all_sheet_configs: Dict[str, Any], # <--- Add param for all sheet configs
    data_source: Union[Dict[str, List[Any]], Dict[Tuple, Dict[str, Any]]],
    data_source_type: str,
    header_info: Dict[str, Any],
    mapping_rules: Dict[str, Any],
    sheet_styling_config: Optional[Dict[str, Any]] = None,
    add_blank_after_header: bool = False,
    static_content_after_header: Optional[Dict[str, Any]] = None,
    add_blank_before_footer: bool = False,
    static_content_before_footer: Optional[Dict[str, Any]] = None,
    merge_rules_after_header: Optional[Dict[str, int]] = None,
    merge_rules_before_footer: Optional[Dict[str, int]] = None,
    merge_rules_footer: Optional[Dict[str, int]] = None, # Added footer merge rules
    footer_info: Optional[Dict[str, Any]] = None, # Currently unused
    max_rows_to_fill: Optional[int] = None,
    grand_total_pallets: int = 0, # RE-ADDED parameter
    custom_flag: bool = False, # Added custom flag parameter
    data_cell_merging_rules: Optional[Dict[str, Any]] = None, # Added data cell merging rules 29/05/2025
    DAF_mode: Optional[bool] = False,
    ) -> Tuple[bool, int, int, int, int]: # Still 5 return values
    """
    REVISED LOGIC V13: Added merge_rules_footer parameter.
    Footer pallet count uses local_chunk_pallets for processed_tables,
    and grand_total_pallets for aggregation/DAF_aggregation.
    """

    # --- Initialize Variables --- (Keep existing initializations)
    actual_rows_to_process = 0; data_rows_prepared = []; col1_index = 1; num_static_labels = 0

    columns_to_grid = []
    desc_col_idx = None
    local_chunk_pallets = 0
    dynamic_desc_used = False



    # get data source pallet count and handle null/conversion errors
    for pallet_count in data_source.get("pallet_count", []):
        if pallet_count is not None:
            try:
                # Convert to float first to handle decimal strings, then to int
                numeric_pallet_count = float(str(pallet_count).strip())
                local_chunk_pallets += int(numeric_pallet_count)
            except (ValueError, TypeError) as e:
                # Log the conversion error but continue processing
                print(f"Warning: Could not convert pallet_count '{pallet_count}' to number: {e}")
                continue

    # --- Row Index Tracking --- (Keep existing)
    row_after_header_idx = -1
    data_start_row = -1
    data_end_row = -1
    row_before_footer_idx = -1
    footer_row_final = -1

    # Ensure dictionaries/lists are initialized (Keep existing)
    static_content_after_header = static_content_after_header or {}
    static_content_before_footer = static_content_before_footer or {}
    merge_rules_after_header = merge_rules_after_header or {}
    merge_rules_before_footer = merge_rules_before_footer or {}
    merge_rules_footer = merge_rules_footer or {} # Initialize footer merge rules
    mapping_rules = mapping_rules or {}
    col_id_map = header_info.get('column_id_map', {})
    column_map = header_info.get('column_map', {})
    idx_to_header_map = {v: k for k, v in column_map.items()}


    try:
        data_cell_merging_rules = data_cell_merging_rules or {}
        # --- Validate Header Info ---
        if not header_info or 'second_row_index' not in header_info or 'column_map' not in header_info or 'num_columns' not in header_info:
            print("Error: Invalid header_info provided.")
            return False, -1, -1, -1, 0

        # --- FIX: Extract num_columns and other values from header_info ---
        num_columns = header_info['num_columns']
        data_writing_start_row = header_info['second_row_index'] + 1
 
        # --- Find Description & Pallet Info Column Indices --- (Keep existing)
        desc_col_idx = col_id_map.get("col_desc")
        pallet_info_col_idx = col_id_map.get("col_pallet")
        if pallet_info_col_idx is None: print("Warning: Header 'Pallet Info' not found.")

        # --- ADD/MODIFY THIS PART FOR PALLET INFO INDEX ---
        if pallet_info_col_idx is None:
            print("Warning: Could not find a 'Pallet Info' (e.g., 'Pallet\\nNo') column header.")
        # --- END OF ADDITION/MODIFICATION FOR PALLET INFO INDEX ---

        # --- Get Styling Config --- (Keep existing)
        force_text_headers = []
        effective_header_font = bold_font # Start with default
        effective_header_align = center_alignment # Start with default

        if sheet_styling_config:
            columns_to_grid = sheet_styling_config.get("column_ids_with_full_grid", [])
            if not isinstance(columns_to_grid, list): columns_to_grid = []

            force_text_headers = sheet_styling_config.get("force_text_format_ids", [])
            if not isinstance(force_text_headers, list): force_text_headers = []

            header_font_cfg = sheet_styling_config.get("header_font")
            if header_font_cfg and isinstance(header_font_cfg, dict):
                 font_params = {k: v for k, v in header_font_cfg.items() if v is not None}
                 if font_params:
                     try: # Expanded try block
                         effective_header_font = Font(**font_params)
                     except TypeError:
                         print(f"Warning: Invalid parameters in header_font config: {font_params}. Using default.")
                         pass # Keep default font on error
                     except Exception as font_err: # Catch other potential errors
                         print(f"Warning: Error applying header_font config: {font_err}. Using default.")
                         pass # Keep default font on error

            header_align_cfg = sheet_styling_config.get("header_alignment")
            if header_align_cfg and isinstance(header_align_cfg, dict):
                 align_params = {k: v for k, v in header_align_cfg.items() if v is not None}
                 if align_params:
                     try: # Expanded try block
                         effective_header_align = Alignment(**align_params)
                     except TypeError:
                         print(f"Warning: Invalid parameters in header_alignment config: {align_params}. Using default.")
                         pass # Keep default alignment on error
                     except Exception as align_err: # Catch other potential errors
                          print(f"Warning: Error applying header_alignment config: {align_err}. Using default.")
                          pass # Keep default alignment on error
        parsed_rules = parse_mapping_rules(
            mapping_rules=mapping_rules,
            column_id_map=col_id_map,
            idx_to_header_map=idx_to_header_map
        )

        # Unpack the results into local variables for the rest of the function to use
        static_value_map = parsed_rules["static_value_map"]
        initial_static_col1_values = parsed_rules["initial_static_col1_values"]
        dynamic_mapping_rules = parsed_rules["dynamic_mapping_rules"]
        formula_rules = parsed_rules["formula_rules"]
        col1_index = parsed_rules["col1_index"]
        num_static_labels = parsed_rules["num_static_labels"]
        static_column_header_name = parsed_rules["static_column_header_name"]
        apply_special_border_rule = parsed_rules["apply_special_border_rule"]
        fallback_on_none = parsed_rules.get("dynamic_mapping_rules", {}).get("description", {}).get("fallback_on_none")

        # --- Prepare Data Rows for Writing (Determine number of rows needed from source) ---
        # This section remains largely the same, preparing the `data_rows_prepared` list
        # which holds the *input* data, not the calculated formulas.
        desc_col_idx = col_id_map.get("col_desc") # Get the description column index
        data_rows_prepared, pallet_counts_for_rows, dynamic_desc_used, num_data_rows_from_source = prepare_data_rows(
            data_source_type=data_source_type,
            data_source=data_source,
            dynamic_mapping_rules=dynamic_mapping_rules,
            column_id_map=col_id_map,
            idx_to_header_map=idx_to_header_map,
            desc_col_idx=desc_col_idx,
            num_static_labels=num_static_labels,
            static_value_map=static_value_map,
            DAF_mode=DAF_mode,
        )
# --- Determine Final Number of Data Rows ---
# The number of rows to process is the greater of the number of data rows or static labels.
        actual_rows_to_process = max(len(data_rows_prepared), num_static_labels)

        # Optional: Apply max_rows_to_fill constraint if it exists
        if max_rows_to_fill is not None and max_rows_to_fill >= 0:
            actual_rows_to_process = min(actual_rows_to_process, max_rows_to_fill)

        # Ensure pallet counts list matches the number of rows we intend to process
        if len(pallet_counts_for_rows) < actual_rows_to_process: pallet_counts_for_rows.extend([0] * (actual_rows_to_process - len(pallet_counts_for_rows)))
        elif len(pallet_counts_for_rows) > actual_rows_to_process: pallet_counts_for_rows = pallet_counts_for_rows[:actual_rows_to_process]

        # --- Calculate Total Rows to Insert and Row Indices ---
        total_rows_to_insert = 0
        current_row_offset = 0

        # Row after header (static/blank)
        if add_blank_after_header:
            row_after_header_idx = data_writing_start_row + current_row_offset
            total_rows_to_insert += 1
            current_row_offset += 1
        else:
            row_after_header_idx = -1 # Indicate no blank row

        # Data rows
        data_start_row = data_writing_start_row + current_row_offset
        if actual_rows_to_process > 0:
            data_end_row = data_start_row + actual_rows_to_process - 1
            total_rows_to_insert += actual_rows_to_process
            current_row_offset += actual_rows_to_process
        else:
            # No data rows to process (can happen if source is empty)
            data_end_row = data_start_row - 1 # Indicate no data rows

        # Row before footer (static/blank)
        if add_blank_before_footer:
            row_before_footer_idx = data_writing_start_row + current_row_offset
            total_rows_to_insert += 1
            current_row_offset += 1
        else:
            row_before_footer_idx = -1 # Indicate no blank row

        # Calculate final footer row index relative to where this chunk starts
        footer_row_final = data_writing_start_row + total_rows_to_insert
        total_rows_to_insert += 1 # Add 1 for the footer itself

        # --- Bulk Insert Rows --- # V11: Only insert if NOT pre-inserted by caller (i.e., for single-table modes)
        if data_source_type in ['aggregation', 'DAF_aggregation', "custom_aggregation"]:
            if total_rows_to_insert > 0:
                try:
                    worksheet.insert_rows(data_writing_start_row, amount=total_rows_to_insert)
                    # Unmerge the block covering the inserted rows *before* the footer starts
                    safe_unmerge_block(worksheet, data_writing_start_row, footer_row_final - 1, num_columns)
                    print("Rows inserted and unmerged successfully.")
                except Exception as bulk_insert_err:
                    print(f"Error during single-table bulk row insert/unmerge: {bulk_insert_err}")
                    # Adjust fallback row calculation
                    fallback_row = max(header_info.get('second_row_index', 0) + 1, footer_row_final)
                    return False, fallback_row, -1, -1, 0

        # --- Fill Row After Header (if applicable) --- 

        # --- Prepare DAF Data Dictionary (inside loop now, safer) ---
        # Removed the premature preparation block here.
        # DAF data dict will be prepared inside the loop if data_source_type is DAF_aggregation.

        # --- Fill Data Rows Loop ---
        if actual_rows_to_process > 0:
            print(f"--- DEBUG START LOOP (Sheet: {worksheet.title}) ---")
            print(f"  data_start_row: {data_start_row}")
            print(f"  actual_rows_to_process: {actual_rows_to_process}")
            print(f"  num_static_labels: {num_static_labels}")
            print(f"  col1_index: {col1_index}")
            print(f"  initial_static_col1_values: {initial_static_col1_values}")
            print(f"  data_source_type: {data_source_type}")
            # --- END DEBUG START LOOP ---
        try:
            # --- Create a reverse map from index to ID for easy lookups inside the loop ---
            idx_to_id_map = {v: k for k, v in col_id_map.items()}

            data_row_indices_written = []
        except Exception as fill_data_err:
            print(f"Error during data filling loop: {fill_data_err}\n{traceback.format_exc()}")
            return False, footer_row_final + 1, data_start_row, data_end_row, 0

    # Merge Description Column if the layout used fallback/static data
        if not dynamic_desc_used and data_start_row > 0 and data_end_row > data_start_row:
            desc_col_id = "col_desc" 
            if col_id_map.get(desc_col_id):
                merge_contiguous_cells_by_id(
                    worksheet=worksheet,
                    start_row=data_start_row,
                    end_row=data_end_row,
                    col_id_to_merge=desc_col_id,
                    column_id_map=col_id_map
                )

        # Always try to merge the Pallet Info Column if it exists
        if data_start_row > 0 and data_end_row > data_start_row:
            pallet_col_id = "col_pallet" 
            if col_id_map.get(pallet_col_id):
                merge_contiguous_cells_by_id(
                    worksheet=worksheet,
                    start_row=data_start_row,
                    end_row=data_end_row,
                    col_id_to_merge=pallet_col_id,
                    column_id_map=col_id_map
                )
        if data_start_row > 0 and data_end_row > data_start_row:
            pallet_col_id = "col_hs" 
            if col_id_map.get(pallet_col_id):
                merge_contiguous_cells_by_id(
                    worksheet=worksheet,
                    start_row=data_start_row,
                    end_row=data_end_row,
                    col_id_to_merge=pallet_col_id,
                    column_id_map=col_id_map
                )

# --- Fill Row Before Footer ---
        if add_blank_before_footer and row_before_footer_idx > 0:
            try:
                # Step 1: Fill the row with content (this applies default styles)
                fill_static_row(worksheet, row_before_footer_idx, num_columns, static_content_before_footer, sheet_styling_config)
                
                # Step 2: Apply the special styling and borders for this specific row
                for c_idx in range(1, num_columns + 1):
                    cell = worksheet.cell(row=row_before_footer_idx, column=c_idx)
                    current_col_id = idx_to_id_map.get(c_idx)
                    context = {
                        "col_id": current_col_id,
                        "col_idx": c_idx,
                        "static_col_idx": col1_index,
                        "is_pre_footer": True,
                        "DAF_mode": DAF_mode
                    }
                    apply_cell_style(cell, sheet_styling_config, context)
            except Exception as fill_bf_err:
                print(f"Warning: Error filling/styling row before footer: {fill_bf_err}")
        

        # --- Fill Footer Row --- (Keep existing logic)
        # The SUM formulas here should correctly sum the results of the formulas
        # written in the data rows above.
        if footer_row_final > 0:
            # Get the footer configuration object from the main sheet config
            footer_config = sheet_config.get("footer_configurations", {})
            data_range_to_sum = [(data_start_row, data_end_row)]

            pallet_count = 0
            if data_source_type == "processed_tables":
                pallet_count = local_chunk_pallets
            else:
                pallet_count = grand_total_pallets

            write_footer_row(
                worksheet=worksheet,
                footer_row_num=footer_row_final,
                header_info=header_info,
                sum_ranges=data_range_to_sum,
                footer_config=footer_config,
                pallet_count=pallet_count,
                DAF_mode=data_source_type == "DAF_aggregation",
                sheet_styling_config=sheet_styling_config
            )
    # No need to pass font, alignment, num_columns, etc. as the
    # function gets this info from header_info and footer_config.
        # --- Apply Merges ---
        # Apply merges to row after header (if applicable)
        if add_blank_after_header and row_after_header_idx > 0 and merge_rules_after_header:
            apply_row_merges(worksheet, row_after_header_idx, num_columns, merge_rules_after_header)

        # Apply merges to row before footer (if applicable)
        target_row_for_bf_merge = row_before_footer_idx if add_blank_before_footer and row_before_footer_idx > 0 else -1
        if target_row_for_bf_merge > 0 and merge_rules_before_footer:
            apply_row_merges(worksheet, target_row_for_bf_merge, num_columns, merge_rules_before_footer)

        # Apply merges to the footer row itself (if applicable)
        if footer_row_final > 0 and merge_rules_footer:
            print(f"Applying footer merges to row {footer_row_final} with rules: {merge_rules_footer}") # Optional Debug
            try:
                apply_row_merges(worksheet, footer_row_final, num_columns, merge_rules_footer)
            except Exception as footer_merge_err:
                 print(f"Warning: Error applying footer merges: {footer_merge_err}")

        # --- Apply Row Heights --- (Keep existing)
        apply_row_heights(worksheet=worksheet, sheet_styling_config=sheet_styling_config, header_info=header_info, data_row_indices=data_row_indices_written, footer_row_index=footer_row_final, row_after_header_idx=row_after_header_idx, row_before_footer_idx=row_before_footer_idx)

        # --- Finalization --- (Keep existing)
        next_available_row_final = footer_row_final + 1
        if actual_rows_to_process == 0: data_start_row = -1; data_end_row = -1
        return True, next_available_row_final, data_start_row, data_end_row, local_chunk_pallets

    except Exception as e:
        # --- Error Handling --- (Keep existing)
        print(f"Critical error in fill_invoice_data: {e}\n{traceback.format_exc()}")
        fallback_row = header_info.get('second_row_index', 0) + 1; frf_local = locals().get('footer_row_final', -1)
        if frf_local > 0: fallback_row = max(fallback_row, frf_local + 1)
        else: est_footer = locals().get('initial_insert_point', fallback_row) + locals().get('total_rows_to_insert', 0); fallback_row = max(fallback_row, est_footer)
        return False, fallback_row, -1, -1, 0



