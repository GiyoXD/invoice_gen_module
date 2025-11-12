# This module contains low-level, generic Excel operations using openpyxl.
# These functions are not specific to any part of the invoice generation process
# and can be reused in various contexts.

import openpyxl
import traceback
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
from typing import Dict, List, Optional, Tuple, Any
import logging
logger = logging.getLogger(__name__)

center_alignment = Alignment(horizontal='center', vertical='center')

def store_original_merges(workbook: openpyxl.Workbook, sheet_names: List[str]) -> Dict[str, List[Tuple[int, Any, Optional[float]]]]:
    """
    Stores the HORIZONTAL span (colspan), the value of the top-left cell,
    and the height of the starting row for merged ranges in specified sheets,
    ASSUMING all merges are only 1 row high AND **start at row 16 or below**.
    Merges starting above row 16 (row < 16) are ignored.
    WARNING: Does NOT store starting coordinates... (rest of docstring unchanged)
    """
    original_merges = {}
    logger.debug("\nStoring original merge horizontal spans, top-left values, and row heights (NO coordinates)...")
    logger.debug("  (Ignoring merges that start above row 16)") # Updated filter info
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            worksheet: Worksheet = workbook[sheet_name] # Type hint for clarity
            merges_data = []
            merged_ranges_copy = list(worksheet.merged_cells.ranges)
            skipped_above_16_count = 0 # Counter for this filter

            for merged_range in merged_ranges_copy:
                min_col, min_row, max_col, max_row = merged_range.bounds

                if max_row != min_row:
                    continue

                if min_row < 16:
                    skipped_above_16_count += 1
                    continue

                col_span = max_col - min_col + 1
                row_height = None # Default to None
                try:
                    row_dim = worksheet.row_dimensions[min_row]
                    row_height = row_dim.height

                    top_left_value = worksheet.cell(row=min_row, column=min_col).value
                    merges_data.append((col_span, top_left_value, row_height))

                except KeyError:
                     logger.warning(f"    Warning: Could not find row dimension for row {min_row} on sheet '{sheet_name}' while getting height. Storing height as None.")
                     try:
                         top_left_value = worksheet.cell(row=min_row, column=min_col).value
                     except Exception as val_e:
                         logger.warning(f"    Warning: Also failed to get value for merge at ({min_row},{min_col}) on sheet '{sheet_name}'. Storing value as None. Error: {val_e}")
                         top_left_value = None
                     merges_data.append((col_span, top_left_value, None))

                except Exception as e:
                    logger.warning(f"    Warning: Could not get value/height for merge starting at ({min_row},{min_col}) on sheet '{sheet_name}'. Storing value/height as None. Error: {e}")
                    merges_data.append((col_span, None, None))

            original_merges[sheet_name] = merges_data
            logger.debug(f"  Stored {len(original_merges[sheet_name])} horizontal merge span/value/height entries for sheet '{sheet_name}'.")
            if skipped_above_16_count > 0:
                logger.debug(f"    (Skipped {skipped_above_16_count} merges starting above row 16)")
        else:
             logger.warning(f"  Warning: Sheet '{sheet_name}' specified but not found during merge storage.")
             original_merges[sheet_name] = []
    return original_merges

def find_and_restore_merges_heuristic(workbook: openpyxl.Workbook,
                                      stored_merges: Dict[str, List[Tuple[int, Any, Optional[float]]]],
                                      processed_sheet_names: List[str],
                                      search_range_str: str = "A16:H200"):
    """
    Attempts to restore merges based on stored HORIZONTAL spans, values, and row heights
    by searching for the value within a specified range (default A16:H200).
    This version is silent, with no detailed logging.
    """
    logger.debug("Starting merge restoration process...")

    restored_count = 0
    failed_count = 0
    skipped_count = 0
    skipped_duplicate_value_count = 0

    try:
        search_min_col, search_min_row, search_max_col, search_max_row = range_boundaries(search_range_str)
    except Exception as e:
        logger.error(f"Error: Invalid search range string '{search_range_str}'. Cannot proceed with restoration. Error: {e}")
        return

    for sheet_name in processed_sheet_names:
        if sheet_name in workbook.sheetnames and sheet_name in stored_merges:
            worksheet: Worksheet = workbook[sheet_name]
            original_merges_data = stored_merges[sheet_name]
            successfully_restored_values_on_sheet = set()

            for col_span, stored_value, stored_height in original_merges_data:

                if col_span <= 1:
                    skipped_count += 1
                    continue

                if stored_value in successfully_restored_values_on_sheet:
                    skipped_duplicate_value_count += 1
                    continue

                found = False
                for r in range(search_max_row, search_min_row - 1, -1):
                    for c in range(search_min_col, search_max_col + 1):
                        current_cell = worksheet.cell(row=r, column=c)
                        current_val = current_cell.value

                        if current_val == stored_value:
                            start_row, start_col = r, c
                            end_row = start_row
                            end_col = start_col + col_span - 1

                            try:
                                worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

                                if stored_height is not None:
                                    try:
                                        worksheet.row_dimensions[start_row].height = stored_height
                                    except Exception:
                                        pass

                                top_left_cell_to_set = worksheet.cell(row=start_row, column=start_col)
                                top_left_cell_to_set.value = stored_value

                                successfully_restored_values_on_sheet.add(stored_value)
                                restored_count += 1
                                found = True
                                break

                            except Exception:
                                failed_count += 1
                                found = True
                                break

                    if found:
                        break

                if not found:
                    if stored_value not in successfully_restored_values_on_sheet:
                        failed_count += 1

    logger.debug("Merge restoration process finished.")
def apply_row_merges(worksheet: Worksheet, row_num: int, num_cols: int, merge_rules: Optional[Dict[str, int]]):
    """
    Applies horizontal merges to a specific row based on a dictionary of rules.
    """
    if not merge_rules:
        return

    logger.debug(f"  Applying custom merge rules for row {row_num}...")
    for start_col_str, colspan_val in merge_rules.items():
        try:
            start_col = int(start_col_str)
            colspan = int(colspan_val)

            if start_col < 1 or colspan <= 1:
                continue

            end_col = start_col + colspan - 1
            if end_col > num_cols:
                end_col = num_cols

            worksheet.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
            cell = worksheet.cell(row=row_num, column=start_col)
            cell.alignment = center_alignment
            logger.debug(f"    - Merged row {row_num} from column {start_col} to {end_col}.")

        except (ValueError, TypeError):
            continue


def merge_vertical_cells_in_range(worksheet: Worksheet, scan_col: int, start_row: int, end_row: int):
    """
    Scans a single column and merges adjacent cells that have the same value.
    """
    if not all(isinstance(i, int) and i > 0 for i in [scan_col, start_row, end_row]) or start_row >= end_row:
        return

    row_idx = start_row
    while row_idx < end_row:
        start_of_merge_row = row_idx
        cell_to_match = worksheet.cell(row=start_of_merge_row, column=scan_col)
        value_to_match = cell_to_match.value

        if value_to_match is None:
            row_idx += 1
            continue

        end_of_merge_row = start_of_merge_row
        for next_row_idx in range(start_of_merge_row + 1, end_row + 1):
            next_cell = worksheet.cell(row=next_row_idx, column=scan_col)
            if next_cell.value == value_to_match:
                end_of_merge_row = next_row_idx
            else:
                break

        if end_of_merge_row > start_of_merge_row:
            try:
                worksheet.merge_cells(
                    start_row=start_of_merge_row,
                    start_column=scan_col,
                    end_row=end_of_merge_row,
                    end_column=scan_col
                )
                cell_to_match.alignment = center_alignment
            except Exception:
                pass

        row_idx = end_of_merge_row + 1