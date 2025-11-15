from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Any, Optional, Tuple
from openpyxl.utils import get_column_letter
from ..styling.models import StylingConfigModel
from ..styling.style_applier import apply_cell_style, apply_header_style
from ..styling.style_config import THIN_BORDER, NO_BORDER, CENTER_ALIGNMENT, LEFT_ALIGNMENT, BOLD_FONT
from decimal import Decimal, InvalidOperation
import re
import traceback
import logging

logger = logging.getLogger(__name__)




def apply_column_widths(worksheet: Worksheet, sheet_styling_config: Optional[StylingConfigModel], header_map: Optional[Dict[str, int]]):
    """
    Sets column widths based on the configuration.

    Args:
        worksheet: The openpyxl Worksheet object.
        sheet_styling_config: Styling configuration containing the 'column_widths' dictionary.
        header_map: Dictionary mapping header text to column index (1-based).
    """
    if not sheet_styling_config or not header_map: return
    column_widths_cfg = sheet_styling_config.columnIdWidths
    if not column_widths_cfg or not isinstance(column_widths_cfg, dict): return
    for header_text, width in column_widths_cfg.items():
        col_idx = header_map.get(header_text)
        if col_idx:
            col_letter = get_column_letter(col_idx)
            try:
                width_val = float(width)
                if width_val > 0: worksheet.column_dimensions[col_letter].width = width_val
                else: pass # Ignore non-positive widths
            except (ValueError, TypeError): pass # Ignore invalid width values
            except Exception as width_err: pass # Log other errors?
        else: pass # Header text not found in map

def calculate_header_dimensions(header_layout: List[Dict[str, Any]]) -> Tuple[int, int]:
    """
    Calculates the total number of rows and columns a header will occupy.
    """
    if not header_layout:
        return (0, 0)
    num_rows = max(cell.get('row', 0) + cell.get('rowspan', 1) for cell in header_layout)
    num_cols = max(cell.get('col', 0) + cell.get('colspan', 1) for cell in header_layout)
    return (num_rows, num_cols)

def merge_contiguous_cells_by_id(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    col_id_to_merge: str,
    column_id_map: Dict[str, int]
):
    """
    Finds and merges contiguous vertical cells within a column that have the same value.
    This is called AFTER all data has been written to the sheet.
    """
    col_idx = column_id_map.get(col_id_to_merge)
    if not col_idx or start_row >= end_row:
        return

    current_merge_start_row = start_row
    value_to_match = worksheet.cell(row=start_row, column=col_idx).value

    for row_idx in range(start_row + 1, end_row + 2):
        cell_value = worksheet.cell(row=row_idx, column=col_idx).value if row_idx <= end_row else object()
        if cell_value != value_to_match:
            if row_idx - 1 > current_merge_start_row:
                if value_to_match is not None and str(value_to_match).strip():
                    try:
                        worksheet.merge_cells(
                            start_row=current_merge_start_row,
                            start_column=col_idx,
                            end_row=row_idx - 1, end_column=col_idx
                        )
                    except Exception as e:
                        logger.error(f"Could not merge cells for ID {col_id_to_merge} from row {current_merge_start_row} to {row_idx - 1}. Error: {e}")
            current_merge_start_row = row_idx
            if row_idx <= end_row:
                value_to_match = cell_value


def write_summary_rows(
    worksheet: Worksheet,
    start_row: int,
    header_info: Dict[str, Any],
    all_tables_data: Dict[str, Any],
    table_keys: List[str],
    footer_config: Dict[str, Any],
    mapping_rules: Dict[str, Any],
    styling_config: Optional[StylingConfigModel] = None,
    DAF_mode: Optional[bool] = False,
    grand_total_pallets: int = 0,
    style_registry=None,
    cell_styler=None
) -> int:
    """
    Calculates and writes ID-driven summary rows for different leather types.
    This function is specifically designed to handle "BUFFALO LEATHER" and "COW LEATHER"
    summaries, calculating totals for specified numeric columns and pallet counts.
    It applies styling based on StyleRegistry (modern) or falls back to legacy styling.
    
    Args:
        style_registry: StyleRegistry instance for modern ID-driven styling
        cell_styler: CellStyler instance for applying styles
    """
    buffalo_summary_row = start_row
    leather_summary_row = start_row + 1
    next_available_row = start_row + 2

    try:
        # --- Configuration and Data Extraction ---
        column_id_map = header_info.get('column_id_map', {})
        idx_to_id_map = {v: k for k, v in column_id_map.items()}
        data_map = mapping_rules.get('data_map', {})
        
        # Define which column IDs should be summed up
        numeric_ids_to_sum = ["col_qty_pcs", "col_qty_sf", "col_net", "col_gross", "col_cbm"]
        id_to_data_key_map = {v['id']: k for k, v in data_map.items() if v.get('id') in numeric_ids_to_sum}
        ids_to_sum = list(id_to_data_key_map.keys())

        # --- Totals Initialization ---
        grand_totals = {col_id: 0 for col_id in ids_to_sum}
        buffalo_totals = {col_id: 0 for col_id in ids_to_sum}
        cow_totals = {col_id: 0 for col_id in ids_to_sum}
        grand_pallet_total = grand_total_pallets # Initialize with the passed parameter
        buffalo_pallet_total = 0
        cow_pallet_total = 0

        for table_key in table_keys:
            table_data = all_tables_data.get(str(table_key), {})
            descriptions = table_data.get("description", [])
            pallet_counts = table_data.get("pallet_count", [])

            for i in range(len(descriptions)):
                raw_val = descriptions[i]
                desc_val = str(raw_val)
                is_buffalo = "BUFFALO" in desc_val.upper()

                logger.debug(f"DEBUG: desc_val: {desc_val}, is_buffalo: {is_buffalo}")

                try:
                    pallet_val = int(pallet_counts[i]) if i < len(pallet_counts) else 0
                    # REMOVED: grand_pallet_total += pallet_val
                    if is_buffalo:
                        buffalo_pallet_total += pallet_val
                    else: # If not buffalo, it's cow leather
                        cow_pallet_total += pallet_val
                    logger.debug(f"DEBUG: Pallet - grand: {grand_pallet_total}, buffalo: {buffalo_pallet_total}, cow: {cow_pallet_total}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid pallet count value at index {i} for description '{desc_val}': {e}")
                    pass

                for col_id in ids_to_sum:
                    data_key = id_to_data_key_map.get(col_id)
                    if not data_key: continue
                    
                    data_list = table_data.get(data_key, [])
                    if i < len(data_list):
                        try:
                            value_to_add = data_list[i]
                            numeric_value = 0
                            if isinstance(value_to_add, (int, float)):
                                numeric_value = float(value_to_add)
                            elif isinstance(value_to_add, str) and value_to_add.strip():
                                numeric_value = float(value_to_add.replace(',', ''))
                            
                            grand_totals[col_id] += numeric_value
                            if is_buffalo:
                                buffalo_totals[col_id] += numeric_value
                            else: # If not buffalo, it's cow leather
                                cow_totals[col_id] += numeric_value
                            logger.debug(f"DEBUG: {col_id} - grand: {grand_totals[col_id]}, buffalo: {buffalo_totals[col_id]}, cow: {cow_totals[col_id]}") # Changed to logging.debug
                        except (ValueError, TypeError, IndexError) as e:
                            logger.warning(f"Failed to process numeric value for {col_id} at index {i} (description '{desc_val}'): {e}")
                            pass

        # --- Writing to Worksheet ---
        num_columns = header_info['num_columns']
        desc_col_idx = column_id_map.get("col_desc")
        label_col_idx = column_id_map.get("col_pallet") or 2
        
        # Helper function to apply styling without borders
        def apply_summary_style(cell, col_id):
            """Apply styling without borders for summary rows"""
            if not style_registry or not cell_styler or not col_id:
                logger.warning(f"Cannot apply summary style: style_registry={style_registry is not None}, cell_styler={cell_styler is not None}, col_id={col_id}")
                return
            
            style = style_registry.get_style(col_id, context='footer')
            # Remove borders by setting border_style to None
            from copy import deepcopy
            style_no_border = deepcopy(style)
            style_no_border['border_style'] = None
            cell_styler.apply(cell, style_no_border)

        # Write Buffalo Summary Row
        cell = worksheet.cell(row=buffalo_summary_row, column=label_col_idx, value="TOTAL OF:")
        apply_summary_style(cell, "col_pallet")
        cell = worksheet.cell(row=buffalo_summary_row, column=label_col_idx + 1, value="BUFFALO LEATHER")
        apply_summary_style(cell, "col_pallet")
        if desc_col_idx:
            cell = worksheet.cell(row=buffalo_summary_row, column=desc_col_idx, value=f"{buffalo_pallet_total} PALLETS")
            apply_summary_style(cell, "col_desc")
        for col_id, total_value in buffalo_totals.items():
            col_idx = column_id_map.get(col_id)
            if col_idx:
                cell = worksheet.cell(row=buffalo_summary_row, column=col_idx, value=total_value)
                apply_summary_style(cell, col_id)

        # Write Cow Leather Summary Row
        cell = worksheet.cell(row=leather_summary_row, column=label_col_idx, value="TOTAL OF:")
        apply_summary_style(cell, "col_pallet")
        cell = worksheet.cell(row=leather_summary_row, column=label_col_idx + 1, value="COW LEATHER")
        apply_summary_style(cell, "col_pallet")
        if desc_col_idx:
            cell = worksheet.cell(row=leather_summary_row, column=desc_col_idx, value=f"{cow_pallet_total} PALLETS")
            apply_summary_style(cell, "col_desc")
        for col_id, total_value in cow_totals.items():
            col_idx = column_id_map.get(col_id)
            if col_idx:
                cell = worksheet.cell(row=leather_summary_row, column=col_idx, value=total_value)
                apply_summary_style(cell, col_id)

        return next_available_row

    except Exception as summary_err:
        logger.error(f"Warning: Failed processing summary rows: {summary_err}") # Changed to logging.error
        traceback.print_exc()
        return start_row + 2
