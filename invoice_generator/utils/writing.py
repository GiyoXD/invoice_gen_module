try:
    import sys
except ImportError as e:
    print(f"ImportError for sys: {e}")
print(f"Loading module: {__file__}")
try:
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:
    print(f"ImportError for openpyxl.worksheet.worksheet: {e}")
try:
    from typing import List, Dict, Any, Optional, Tuple
except ImportError as e:
    print(f"ImportError for typing: {e}")
try:
    from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
except ImportError as e:
    print(f"ImportError for openpyxl.styles: {e}")
try:
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"ImportError for openpyxl.utils: {e}")
try:
    from .layout import unmerge_row, unmerge_block, safe_unmerge_block, calculate_header_dimensions
except ImportError as e:
    print(f"ImportError for .layout: {e}")
try:
    import re
except ImportError as e:
    print(f"ImportError for re: {e}")
try:
    import traceback
except ImportError as e:
    print(f"ImportError for traceback: {e}")
try:
    from decimal import Decimal, InvalidOperation
except ImportError as e:
    print(f"ImportError for decimal: {e}")
try:
    from . import merge_utils
except ImportError as e:
    print(f"ImportError for .merge_utils: {e}")

# --- Constants for Styling ---
thin_side = Side(border_style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side) # Full grid border
no_border = Border(left=None, right=None, top=None, bottom=None)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
bold_font = Font(bold=True)

# --- Constants for Number Formats ---
FORMAT_GENERAL = 'General'
FORMAT_TEXT = '@'
FORMAT_NUMBER_COMMA_SEPARATED1 = '#,##0'
FORMAT_NUMBER_COMMA_SEPARATED2 = '#,##0.00'

def fill_static_row(worksheet: Worksheet, row_num: int, num_cols: int, static_content_dict: Dict[str, Any]):
    """
    Fills a specific row with static content defined in a dictionary.
    Applies default alignment and no border.

    Args:
        worksheet: The openpyxl Worksheet object.
        row_num: The 1-based row index to fill.
        num_cols: The total number of columns in the table context (for bounds checking).
        static_content_dict: Dictionary where keys are column indices (as strings or ints)
                             and values are the static content to write.
    """
    if not static_content_dict:
        return # Nothing to do
    if row_num <= 0:
        return

    for col_key, value in static_content_dict.items():
        target_col_index = None
        try:
            # Attempt to convert key to integer column index
            target_col_index = int(col_key)
            # Check if the column index is within the valid range
            if 1 <= target_col_index <= num_cols:
                cell = worksheet.cell(row=row_num, column=target_col_index)
                cell.value = value
                # Apply default styling for static rows
                cell.alignment = center_alignment # Default alignment
                cell.border = no_border # Default: no border for static rows
                # Apply basic number formatting
                if isinstance(value, (int, float)):
                    cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED2 if isinstance(value, float) else FORMAT_NUMBER_COMMA_SEPARATED1
                else:
                    cell.number_format = FORMAT_TEXT # Treat as text otherwise
            else:
                # Column index out of range, log warning?
                pass
        except (ValueError, TypeError) as e:
            # Invalid column key, log warning?
            pass
        except Exception as cell_err:
            # Error accessing cell, log warning?
            pass

def apply_row_merges(worksheet: Worksheet, row_num: int, num_cols: int, merge_rules: Optional[Dict[str, int]]):
    """
    Applies horizontal merges to a specific row based on rules.

    Args:
        worksheet: The openpyxl Worksheet object.
        row_num: The 1-based row index to apply merges to.
        num_cols: The total number of columns in the table context.
        merge_rules: Dictionary where keys are starting column indices (as strings or ints)
                     and values are the number of columns to span (colspan).
    """
    if not merge_rules or row_num <= 0:
        return # No rules or invalid row

    try:
        # Convert string keys to integers and sort for predictable application order
        rules_with_int_keys = {int(k): v for k, v in merge_rules.items()}
        sorted_keys = sorted(rules_with_int_keys.keys())
    except (ValueError, TypeError) as e:
        # Invalid key format in merge_rules
        return

    for start_col in sorted_keys:
        colspan_val = rules_with_int_keys[start_col]
        try:
            # Ensure colspan is an integer
            colspan = int(colspan_val)
        except (ValueError, TypeError):
            # Invalid colspan value
            continue

        # Basic validation for start column and colspan
        if not isinstance(start_col, int) or not isinstance(colspan, int) or start_col < 1 or colspan < 1:
            continue

        # Calculate end column, ensuring it doesn't exceed the table width
        end_col = start_col + colspan - 1
        if end_col > num_cols:
            end_col = num_cols
            # Check if clamping made the range invalid (start > end)
            if start_col > end_col:
                continue

        merge_range_str = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
        try:
            # --- Pre-Unmerge Overlapping Cells ---
            merges_to_clear = []
            current_merged_ranges = list(worksheet.merged_cells.ranges) # Work on a copy
            for merged_range in current_merged_ranges:
                # Check if the existing merge overlaps with the target row and column range
                if merged_range.min_row <= row_num <= merged_range.max_row:
                    if max(merged_range.min_col, start_col) <= min(merged_range.max_col, end_col):
                        range_to_remove_str = str(merged_range)
                        if range_to_remove_str not in merges_to_clear:
                            merges_to_clear.append(range_to_remove_str)
            if merges_to_clear:
                for r_str in merges_to_clear:
                    try: worksheet.unmerge_cells(r_str)
                    except KeyError: pass
                    except Exception as unmerge_err_inner: pass # Log?
            # --- End Pre-Unmerge ---

            worksheet.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
            # Apply alignment to the top-left cell of the merged range
            top_left_cell = worksheet.cell(row=row_num, column=start_col)
            if not top_left_cell.alignment or top_left_cell.alignment.horizontal is None:
                top_left_cell.alignment = center_alignment # Apply center alignment if none exists
        except ValueError as ve:
            # This can happen if trying to merge over an existing merged cell that wasn't properly unmerged
            pass
        except Exception as merge_err:
            # Log or handle other merge errors
            pass

def write_grand_total_weight_summary(
    worksheet: Worksheet,
    start_row: int,
    header_info: Dict[str, Any],
    processed_tables_data: Dict[str, Dict[str, List[Any]]],
    weight_config: Dict[str, Any],
    styling_config: Optional[Dict[str, Any]] = None
) -> int:
    """
    Calculates GRAND TOTAL of Net/Gross weights, inserts two new rows,
    and writes a styled two-row summary using the main footer's style.

    Args:
        worksheet: The openpyxl worksheet to modify.
        start_row: The row index to start writing from.
        header_info: The header dictionary containing 'column_id_map' and 'num_columns'.
        processed_tables_data: The dictionary containing all table data.
        weight_config: The configuration object for the weight summary.
        footer_config: The main footer configuration for the sheet, used for styling.
    """
    footer_row_height = styling_config.get("styling", {}).get("row_heights", {}).get("footer", None)
    footer_config = styling_config.get("footer_configurations", {})

    if not weight_config.get("enabled"):
        return start_row

    print(f"--- Calculating and writing GRAND TOTAL Net/Gross Weight summary ---")

    # --- Calculation Logic (no changes here) ---
    grand_total_net = Decimal('0')
    grand_total_gross = Decimal('0')

    for table_data in processed_tables_data.values():
        net_weights = table_data.get("net", [])
        gross_weights = table_data.get("gross", [])
        for weight in net_weights:
            try:
                grand_total_net += Decimal(str(weight))
            except (InvalidOperation, TypeError, ValueError):
                continue
        for weight in gross_weights:
            try:
                grand_total_gross += Decimal(str(weight))
            except (InvalidOperation, TypeError, ValueError):
                continue

    # --- Get Column Indices and Dimensions (no changes here) ---
    col_id_map = header_info.get("column_id_map", {})
    num_columns = header_info.get("num_columns", 1)
    label_col_idx = col_id_map.get(weight_config.get("label_col_id"))
    value_col_idx = col_id_map.get(weight_config.get("value_col_id"))

    if not all([label_col_idx, value_col_idx]):
        print("Warning: Could not write grand total weight summary. Label/Value column ID not found.")
        return start_row

    # --- MODIFICATION: Parse Styling from the main footer_config ---
    # It now uses the new 'footer_config' parameter
    style_config = footer_config.get('style', {})
    font_to_apply = Font(**style_config.get('font', {'bold': True}))
    align_to_apply = Alignment(**style_config.get('alignment', {'horizontal': 'right', 'vertical': 'center'}))

    # --- Insert and unmerge rows (no changes here) ---
    try:
        worksheet.insert_rows(start_row, amount=2)
        unmerge_row(worksheet, start_row, num_columns)
        unmerge_row(worksheet, start_row + 1, num_columns)
    except Exception as insert_err:
        print(f"Error inserting/unmerging rows for weight summary: {insert_err}")
        return start_row

    # --- Write the final rows and apply styles (no changes here) ---
    net_weight_row = start_row
    gross_weight_row = start_row + 1

    try:
        cell_net_label = worksheet.cell(row=net_weight_row, column=label_col_idx, value="NW(KGS)")
        cell_net_value = worksheet.cell(row=net_weight_row, column=value_col_idx, value=float(grand_total_net))
        cell_net_value.number_format = FORMAT_NUMBER_COMMA_SEPARATED2

        cell_gross_label = worksheet.cell(row=gross_weight_row, column=label_col_idx, value="GW(KGS):")
        cell_gross_value = worksheet.cell(row=gross_weight_row, column=value_col_idx, value=float(grand_total_gross))
        cell_gross_value.number_format = FORMAT_NUMBER_COMMA_SEPARATED2

        for cell in [cell_net_label, cell_net_value, cell_gross_label, cell_gross_value]:
            cell.font = font_to_apply
            cell.alignment = align_to_apply

        # --- Apply Per-Column Alignment and Font from column_id_styles (if available) ---
        if styling_config:
            column_id_styles = styling_config.get("column_id_styles", {})
            idx_to_id_map = {v: k for k, v in col_id_map.items()}
            for cell in [cell_net_label, cell_net_value, cell_gross_label, cell_gross_value]:
                col_idx = cell.column
                column_id = idx_to_id_map.get(col_idx)
                if column_id and column_id in column_id_styles:
                    col_style = column_id_styles[column_id]
                    # Apply alignment if specified
                    if 'alignment' in col_style:
                        cell.alignment = Alignment(**{k: v for k, v in col_style['alignment'].items() if v is not None})
                    # Apply font if specified
                    if 'font' in col_style:
                        cell.font = Font(**{k: v for k, v in col_style['font'].items() if v is not None})
        
        # ADD THIS LINE
        last_mapped_col_idx = max(col_id_map.values()) if col_id_map else 1

        # --- ADD THIS BLOCK TO APPLY BORDERS ---
        border_to_apply = thin_border
        for row_idx in [net_weight_row, gross_weight_row]:
            for col_idx in range(1, last_mapped_col_idx + 1):
                worksheet.cell(row=row_idx, column=col_idx).border = border_to_apply
        # --- END OF BLOCK ---

        if footer_row_height:
            worksheet.row_dimensions[net_weight_row].height = footer_row_height
            worksheet.row_dimensions[gross_weight_row].height = footer_row_height
        

        print("--- Finished writing grand total weight summary. ---")
        return start_row + 2

    except Exception as e:
        print(f"Error writing grand total weight summary content: {e}")
        return start_row

def write_header(worksheet: Worksheet, start_row: int, header_layout_config: List[Dict[str, Any]],
                 sheet_styling_config: Optional[Dict[str, Any]] = None
                 ) -> Optional[Dict[str, Any]]:
    if not header_layout_config or start_row <= 0:
        return None

    num_header_rows, num_header_cols = calculate_header_dimensions(header_layout_config)
    if num_header_rows > 0 and num_header_cols > 0:
        unmerge_block(worksheet, start_row, start_row + num_header_rows - 1, num_header_cols)

    first_row_index = start_row
    last_row_index = start_row
    max_col = 0
    column_map = {}
    column_id_map = {}

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if sheet_styling_config:
        header_font_config = sheet_styling_config.get('header_font')
        if header_font_config:
            header_font = Font(**header_font_config)
        header_alignment_config = sheet_styling_config.get('header_alignment')
        if header_alignment_config:
            header_alignment = Alignment(**header_alignment_config)

    for cell_config in header_layout_config:
        row_offset = cell_config.get('row', 0)
        col_offset = cell_config.get('col', 0)
        text = cell_config.get('text', '')
        cell_id = cell_config.get('id')
        rowspan = cell_config.get('rowspan', 1)
        colspan = cell_config.get('colspan', 1)

        cell_row = start_row + row_offset
        cell_col = 1 + col_offset

        last_row_index = max(last_row_index, cell_row + rowspan - 1)
        max_col = max(max_col, cell_col + colspan - 1)

        cell = worksheet.cell(row=cell_row, column=cell_col, value=text)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

        if cell_id:
            column_map[text] = get_column_letter(cell_col)
            column_id_map[cell_id] = cell_col

        if rowspan > 1 or colspan > 1:
            worksheet.merge_cells(start_row=cell_row, start_column=cell_col,
                                  end_row=cell_row + rowspan - 1, end_column=cell_col + colspan - 1)

    return {
        'first_row_index': first_row_index,
        'second_row_index': last_row_index,
        'column_map': column_map,
        'column_id_map': column_id_map,
        'num_columns': max_col
    }


def merge_contiguous_cells_by_id(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    col_id_to_merge: str,
    column_id_map: Dict[str, int]
):
    """
    Finds and merges contiguous vertical cells within a column that have the same value.
    This is called AFTER all data has been written to the sheet.
    """
    col_idx = column_id_map.get(col_id_to_merge)
    if not col_idx or start_row >= end_row:
        return

    current_merge_start_row = start_row
    value_to_match = worksheet.cell(row=start_row, column=col_idx).value

    for row_idx in range(start_row + 1, end_row + 2):
        cell_value = worksheet.cell(row=row_idx, column=col_idx).value if row_idx <= end_row else object()
        if cell_value != value_to_match:
            if row_idx - 1 > current_merge_start_row:
                if value_to_match is not None and str(value_to_match).strip():
                    try:
                        worksheet.merge_cells(
                            start_row=current_merge_start_row,
                            start_column=col_idx,
                            end_row=row_idx - 1,
                            end_column=col_idx
                        )
                    except Exception as e:
                        print(f"Could not merge cells for ID {col_id_to_merge} from row {current_merge_start_row} to {row_idx - 1}. Error: {e}")
            
            current_merge_start_row = row_idx
            if row_idx <= end_row:
                value_to_match = cell_value

def find_footer(worksheet: Worksheet, footer_rules: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Finds the footer row based on marker text and rules.

    Args:
        worksheet: The openpyxl Worksheet object.
        footer_rules: Dictionary defining how to find the footer (marker text, columns, direction, etc.).

    Returns:
        A dictionary containing the footer start row {'start_row': index} or None if not found.
    """
    marker_text = footer_rules.get('marker_text'); search_type = footer_rules.get('search_type', 'exact'); case_sensitive = footer_rules.get('case_sensitive', True)
    search_columns = footer_rules.get('search_columns', [1]); search_direction = footer_rules.get('search_direction', 'down').lower()
    min_row_offset = footer_rules.get('min_row_offset', 1); max_row_search = footer_rules.get('max_row_to_search', worksheet.max_row)
    max_row_search = min(max_row_search, worksheet.max_row) # Ensure max_row_search doesn't exceed actual max row
    if not marker_text: return None
    if not isinstance(search_columns, list) or not search_columns: search_columns = [1]
    if min_row_offset <= 0: min_row_offset = 1

    try:
        # Determine Row Iteration Order
        row_iterator = None
        if search_direction == 'up': row_iterator = range(max_row_search, min_row_offset - 1, -1)
        else: row_iterator = range(min_row_offset, max_row_search + 1)
        marker_text_str = str(marker_text)

        # Search for Marker
        for r_idx in row_iterator:
            for c_idx in search_columns:
                if not (1 <= c_idx <= worksheet.max_column): continue # Skip invalid column index
                try:
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    # If it's a merged cell, only check the top-left origin cell of the merge range
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        is_origin = False
                        for merged_range in worksheet.merged_cells.ranges:
                            if merged_range.min_row == r_idx and merged_range.min_col == c_idx:
                                is_origin = True; break
                        if not is_origin: continue # Skip if not the top-left cell
                    cell_value_str = str(cell.value) if cell.value is not None else ""
                except IndexError: continue # Should not happen with max_column check, but safety first
                found = False
                if search_type == 'substring':
                    pattern = re.escape(marker_text_str); flags = 0 if case_sensitive else re.IGNORECASE
                    if re.search(pattern, cell_value_str, flags): found = True
                elif case_sensitive and cell_value_str == marker_text_str: found = True
                elif not case_sensitive and cell_value_str.lower() == marker_text_str.lower(): found = True
                if found: return {'start_row': r_idx} # Return immediately when found

        return None # Marker not found
    except Exception as e: return None # Error during search

def write_configured_rows(
    worksheet: Worksheet,
    start_row_index: int,
    num_columns: int,
    rows_config_list: List[Dict[str, Any]], # Primary configuration for each row
    calculated_totals: Dict[str, Any],     # Data values to be inserted
    default_style_config: Optional[Dict[str, Any]] = None # Default styles from sheet config
):
    """
    Writes one or more rows with specified content (labels + dynamic values),
    styling, and merges based on configuration.
    Assumes the rows have already been inserted by the caller.
    Number formats are applied ONLY if specified in the cell's config,
    or set to Text ('@') for labels/non-numeric values.
    """
    if not rows_config_list or start_row_index <= 0:
        return

    print(f"--- Writing {len(rows_config_list)} configured rows starting at row {start_row_index} ---")
    calculated_totals = calculated_totals or {} # Ensure it's a dict

    # --- Get overall default styles from the sheet's styling configuration ---
    # These will be used if a row doesn't specify its own font/alignment.
    overall_default_font = Font() # Basic Openpyxl default
    overall_default_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False) # Basic Openpyxl default

    if default_style_config:
        # Use 'default_font' and 'default_alignment' from the sheet's styling config if available
        sheet_default_font_cfg = default_style_config.get("default_font")
        if sheet_default_font_cfg and isinstance(sheet_default_font_cfg, dict):
            try:
                overall_default_font = Font(**{k: v for k, v in sheet_default_font_cfg.items() if v is not None})
            except TypeError:
                print("Warning: Invalid parameters in sheet's default_font config. Using basic default font.")
        
        sheet_default_align_cfg = default_style_config.get("default_alignment")
        if sheet_default_align_cfg and isinstance(sheet_default_align_cfg, dict):
            try:
                overall_default_alignment = Alignment(**{k: v for k, v in sheet_default_align_cfg.items() if v is not None})
            except TypeError:
                print("Warning: Invalid parameters in sheet's default_alignment config. Using basic default alignment.")

    # Iterate through each row's configuration object
    for i, row_config_item in enumerate(rows_config_list):
        current_row_idx = start_row_index + i
        print(f"  Processing configured row {i+1} (Sheet Row: {current_row_idx})")

        # --- Get ROW-LEVEL configurations from the current row_config_item ---
        row_cell_definitions = row_config_item.get("content", []) # List of cell configs for this row
        
        row_specific_height = row_config_item.get("height")
        row_specific_font_config = row_config_item.get("font")      # Font for the whole row
        row_specific_align_config = row_config_item.get("alignment") # Alignment for the whole row
        row_specific_merge_rules = row_config_item.get("merge_rules") # Merges for this specific row
        row_specific_apply_border = row_config_item.get("apply_default_border", True) # Border for the whole row

        # --- Determine effective font and alignment FOR THIS ENTIRE ROW ---
        # Start with the overall defaults, then apply row-level overrides if they exist.
        effective_row_font = overall_default_font
        if row_specific_font_config and isinstance(row_specific_font_config, dict):
            font_params = {k: v for k, v in row_specific_font_config.items() if v is not None}
            if font_params:
                try:
                    effective_row_font = Font(**font_params)
                except TypeError:
                    print(f"Warning: Invalid font config for row {current_row_idx}. Using sheet/basic default.")

        effective_row_alignment = overall_default_alignment
        if row_specific_align_config and isinstance(row_specific_align_config, dict):
            align_params = {k: v for k, v in row_specific_align_config.items() if v is not None}
            if align_params:
                try:
                    effective_row_alignment = Alignment(**align_params)
                except TypeError:
                    print(f"Warning: Invalid alignment config for row {current_row_idx}. Using sheet/basic default.")

        # --- Write Content Items (Cells) for the current row and Apply Styles ---
        written_columns_in_row = set() # Keep track of columns explicitly written to in this row
        
        if isinstance(row_cell_definitions, list):
            for cell_config_item in row_cell_definitions: # Each item in 'content' array from your JSON
                if not isinstance(cell_config_item, dict):
                    print(f"Warning: Invalid cell config item in row {current_row_idx}: {cell_config_item}")
                    continue

                try:
                    target_col_idx = int(cell_config_item.get("col"))
                    if not (1 <= target_col_idx <= num_columns):
                        print(f"Warning: Column index {target_col_idx} out of range for row {current_row_idx}.")
                        continue

                    cell = worksheet.cell(row=current_row_idx, column=target_col_idx)
                    written_columns_in_row.add(target_col_idx)
                    
                    value_to_write = None
                    # Cell-specific number format, font, and alignment from its own config
                    cell_specific_number_format = cell_config_item.get("number_format")
                    # Note: Cell-specific font/alignment could also be added to JSON if needed,
                    # otherwise, the effective_row_font/alignment will be used.

                    if "label" in cell_config_item:
                        value_to_write = cell_config_item["label"]
                        cell.number_format = cell_specific_number_format or FORMAT_TEXT # Use provided or default to Text
                    elif "value_key" in cell_config_item:
                        value_key = cell_config_item["value_key"]
                        raw_value = calculated_totals.get(value_key)
                        suffix = cell_config_item.get("suffix", "")
                        
                        numeric_value = None
                        if isinstance(raw_value, (int, float)):
                            numeric_value = float(raw_value)
                        elif isinstance(raw_value, str):
                            try:
                                # Attempt to convert if it looks like a number, handling commas
                                cleaned_raw_value = raw_value.replace(',', '')
                                if cleaned_raw_value.strip(): # Avoid empty strings
                                    numeric_value = float(cleaned_raw_value)
                            except (ValueError, TypeError):
                                pass # Keep as None if conversion fails

                        if numeric_value is not None:
                            # If there's a suffix, the value becomes a string.
                            # If no suffix, keep it as a number for Excel to handle.
                            value_to_write = f"{numeric_value}{suffix}" if suffix else numeric_value
                            
                            if cell_specific_number_format:
                                cell.number_format = cell_specific_number_format
                            elif suffix: # If suffix is present, it's text
                                cell.number_format = FORMAT_TEXT
                            # Else (numeric, no suffix, no specific format): Let Excel use default number format
                            
                        else: # Value is not numeric or not found, treat as text
                            value_to_write = f"{str(raw_value or '')}{suffix}" # Use empty string if raw_value is None
                            cell.number_format = cell_specific_number_format or FORMAT_TEXT
                    
                    elif "value" in cell_config_item: # Direct static value
                        value_to_write = cell_config_item.get("value")
                        # Assume direct static values are text unless a number_format is given
                        cell.number_format = cell_specific_number_format or FORMAT_TEXT
                    else:
                        # No label, value_key, or value; cell might be intended to be blank but styled
                        pass


                    cell.value = value_to_write
                    cell.font = effective_row_font # Apply the determined row font
                    cell.alignment = effective_row_alignment # Apply the determined row alignment

                    # Apply border based on row-level setting
                    if row_specific_apply_border:
                        cell.border = thin_border
                    else:
                        cell.border = no_border

                except (ValueError, TypeError) as e:
                    print(f"Warning: Invalid data in cell config for row {current_row_idx}: {cell_config_item}. Error: {e}")
                except Exception as cell_err:
                    print(f"Warning: Error writing cell (Row: {current_row_idx}, Col: {cell_config_item.get('col', 'N/A')}): {cell_err}")

        # --- Ensure remaining (unwritten) cells in the row get default row styling (border) ---
        for c_idx_fill in range(1, num_columns + 1):
            if c_idx_fill not in written_columns_in_row: # Only touch columns not explicitly defined
                try:
                    cell = worksheet.cell(row=current_row_idx, column=c_idx_fill)
                    # Apply row's effective font and alignment to blank cells if desired (optional)
                    # cell.font = effective_row_font
                    # cell.alignment = effective_row_alignment
                    if row_specific_apply_border:
                        cell.border = thin_border
                    else:
                        # Only remove border if cell is truly blank and no border is intended for the row
                        if cell.value is None: # Check if cell is actually empty
                            cell.border = no_border
                except Exception as blank_cell_err:
                    print(f"Warning: Error styling blank cell ({current_row_idx},{c_idx_fill}): {blank_cell_err}")


        # --- Apply Merges for this entire row (using row-level merge rules) ---
        if row_specific_merge_rules and isinstance(row_specific_merge_rules, dict):
            apply_row_merges(worksheet, current_row_idx, num_columns, row_specific_merge_rules)
            # Re-apply style/border to the top-left cell of any merged ranges
            # to ensure consistent appearance, as merging can sometimes affect the primary cell's style.
            for start_col_str_merge in row_specific_merge_rules.keys():
                try:
                    start_col_idx_merge = int(start_col_str_merge)
                    merged_cell_anchor = worksheet.cell(row=current_row_idx, column=start_col_idx_merge)
                    merged_cell_anchor.font = effective_row_font
                    merged_cell_anchor.alignment = effective_row_alignment
                    if row_specific_apply_border:
                        merged_cell_anchor.border = thin_border
                    else:
                        merged_cell_anchor.border = no_border
                except (ValueError, TypeError):
                    print(f"Warning: Invalid start column for merge rule on row {current_row_idx}: {start_col_str_merge}")
                except Exception as merge_style_err:
                    print(f"Warning: Error re-styling merged cell anchor at ({current_row_idx},{start_col_str_merge}): {merge_style_err}")

        # --- Apply Height for this entire row (using row-level height) ---
        if row_specific_height is not None:
            try:
                h_val = float(row_specific_height)
                if h_val > 0:
                    worksheet.row_dimensions[current_row_idx].height = h_val
            except (ValueError, TypeError):
                print(f"Warning: Invalid height value '{row_specific_height}' for row {current_row_idx}.")
            except Exception as height_err:
                print(f"Warning: Error setting height for row {current_row_idx}: {height_err}")

    print(f"--- Finished writing configured rows ---")

def apply_explicit_data_cell_merges_by_id(
    worksheet: Worksheet,
    row_num: int,
    column_id_map: Dict[str, int],  # Maps column ID to its 1-based column index
    num_total_columns: int,
    merge_rules_data_cells: Dict[str, Dict[str, Any]], # e.g., {'col_item': {'rowspan': 2}}
    sheet_styling_config: Optional[Dict[str, Any]] = None,
    DAF_mode: Optional[bool] = False
):
    """
    Applies horizontal merges to data cells in a specific row based on column IDs.
    """
    if not merge_rules_data_cells or row_num <= 0:
        return

    thin_side = Side(border_style="thin", color="000000")
    full_thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Loop through rules where the key is now the column ID
    for col_id, rule_details in merge_rules_data_cells.items():
        colspan_to_apply = rule_details.get("rowspan")

        if not isinstance(colspan_to_apply, int) or colspan_to_apply <= 1:
            continue
        
        # Get column index from the ID map
        start_col_idx = column_id_map.get(col_id)
        if not start_col_idx:
            print(f"Warning: Could not find column for merge rule with ID '{col_id}'.")
            continue
            
        end_col_idx = start_col_idx + colspan_to_apply - 1
        end_col_idx = min(end_col_idx, num_total_columns)

        if start_col_idx >= end_col_idx:
            continue

        try:
            # Unmerge any existing ranges in the target area
            for mc_range in list(worksheet.merged_cells.ranges):
                if mc_range.min_row == row_num and mc_range.max_row == row_num:
                    if mc_range.min_col <= end_col_idx and mc_range.max_col >= start_col_idx:
                        worksheet.unmerge_cells(str(mc_range))
            
            # Apply the new merge
            worksheet.merge_cells(start_row=row_num, start_column=start_col_idx,
                                  end_row=row_num, end_column=end_col_idx)
            
            # Style the anchor cell of the new merged range
            anchor_cell = worksheet.cell(row=row_num, column=start_col_idx)
            
            # Apply base styling for the column ID
            _apply_cell_style(anchor_cell, col_id, sheet_styling_config, DAF_mode)
            
            # Ensure the merged cell has the desired border and alignment
            anchor_cell.border = full_thin_border
            anchor_cell.alignment = center_alignment

        except Exception as e:
            print(f"Error applying explicit data cell merge for ID '{col_id}' on row {row_num}: {e}")

def write_summary_rows(
    worksheet: Worksheet,
    start_row: int,
    header_info: Dict[str, Any],
    all_tables_data: Dict[str, Any],
    table_keys: List[str],
    footer_config: Dict[str, Any],
    mapping_rules: Dict[str, Any],
    styling_config: Optional[Dict[str, Any]] = None,
    DAF_mode: Optional[bool] = False
) -> int:
    """
    Calculates and writes ID-driven summary rows, ensuring text cells are
    formatted as text and the final bold font style is applied correctly.
    """
    buffalo_summary_row = start_row
    leather_summary_row = start_row + 1
    next_available_row = start_row + 2

    try:
        # --- Get Styles from Footer Config ---
        style_config = footer_config.get('style', {})
        font_to_apply = Font(**style_config.get('font', {'bold': True}))
        align_to_apply = Alignment(**style_config.get('alignment', {'horizontal': 'center', 'vertical': 'center'}))

        # --- Calculation and Writing Logic (remains the same) ---
        # ... (all the calculation and cell writing logic is unchanged) ...
        column_id_map = header_info.get('column_id_map', {})
        idx_to_id_map = {v: k for k, v in column_id_map.items()}
        data_map = mapping_rules.get('data_map', {})
        numeric_ids_to_sum = ["col_qty_pcs", "col_qty_sf", "col_net", "col_gross", "col_cbm"]
        id_to_data_key_map = {v['id']: k for k, v in data_map.items() if v.get('id') in numeric_ids_to_sum}
        ids_to_sum = list(id_to_data_key_map.keys())
        buffalo_totals = {col_id: 0 for col_id in ids_to_sum}
        cow_totals = {col_id: 0 for col_id in ids_to_sum}
        buffalo_pallet_total = 0
        cow_pallet_total = 0
        for table_key in table_keys:
            table_data = all_tables_data.get(str(table_key), {})
            descriptions = table_data.get("description", [])
            pallet_counts = table_data.get("pallet_count", [])
            for i in range(len(descriptions)):
                raw_val = descriptions[i]
                desc_val = raw_val
                if isinstance(raw_val, list) and raw_val:
                    desc_val = raw_val[0]
                is_buffalo = desc_val and "BUFFALO" in str(desc_val).upper()
                target_dict = buffalo_totals if is_buffalo else cow_totals
                try:
                    pallet_val = int(pallet_counts[i]) if i < len(pallet_counts) else 0
                    if is_buffalo:
                        buffalo_pallet_total += pallet_val
                    else:
                        cow_pallet_total += pallet_val
                except (ValueError, TypeError): pass
                for col_id in ids_to_sum:
                    data_key = id_to_data_key_map.get(col_id)
                    if not data_key: continue
                    data_list = table_data.get(data_key, [])
                    if i < len(data_list):
                        try:
                            value_to_add = data_list[i]
                            if isinstance(value_to_add, (int, float)):
                                target_dict[col_id] += float(value_to_add)
                            elif isinstance(value_to_add, str) and value_to_add.strip():
                                target_dict[col_id] += float(value_to_add.replace(',', ''))
                        except (ValueError, TypeError, IndexError): pass
        num_columns = header_info['num_columns']
        desc_col_idx = column_id_map.get("col_desc")
        label_col_idx = column_id_map.get("col_pallet") or 2
        unmerge_row(worksheet, buffalo_summary_row, num_columns)
        worksheet.cell(row=buffalo_summary_row, column=label_col_idx, value="TOTAL OF:").number_format = FORMAT_TEXT
        worksheet.cell(row=buffalo_summary_row, column=label_col_idx + 1, value="BUFFALO LEATHER").number_format = FORMAT_TEXT
        if desc_col_idx:
            worksheet.cell(row=buffalo_summary_row, column=desc_col_idx, value=f"{buffalo_pallet_total} PALLETS").number_format = FORMAT_TEXT
        for col_id, total_value in buffalo_totals.items():
            col_idx = column_id_map.get(col_id)
            if col_idx:
                worksheet.cell(row=buffalo_summary_row, column=col_idx, value=total_value)
        unmerge_row(worksheet, leather_summary_row, num_columns)
        worksheet.cell(row=leather_summary_row, column=label_col_idx, value="TOTAL OF:").number_format = FORMAT_TEXT
        worksheet.cell(row=leather_summary_row, column=label_col_idx + 1, value="COW LEATHER").number_format = FORMAT_TEXT
        if desc_col_idx:
            worksheet.cell(row=leather_summary_row, column=desc_col_idx, value=f"{cow_pallet_total} PALLETS").number_format = FORMAT_TEXT
        for col_id, total_value in cow_totals.items():
            col_idx = column_id_map.get(col_id)
            if col_idx:
                worksheet.cell(row=leather_summary_row, column=col_idx, value=total_value)


        # --- Apply Styles to Both Rows with Correct Order ---
        for row_num in [buffalo_summary_row, leather_summary_row]:
            for c_idx in range(1, num_columns + 1):
                cell = worksheet.cell(row=row_num, column=c_idx)
                current_col_id = idx_to_id_map.get(c_idx)

                # Step 1: Apply general footer style first.
                cell.font = font_to_apply
                cell.alignment = align_to_apply
                cell.border = no_border

                # Step 2: Apply column-specific styles to override if present.
                if styling_config and current_col_id:
                    column_id_styles = styling_config.get("column_id_styles", {})
                    if current_col_id in column_id_styles:
                        col_style = column_id_styles[current_col_id]
                        if 'alignment' in col_style:
                            cell.alignment = Alignment(**{k: v for k, v in col_style['alignment'].items() if v is not None})
                        if 'font' in col_style:
                            cell.font = Font(**{k: v for k, v in col_style['font'].items() if v is not None})
        
        # --- Apply Row Height (remains the same) ---
        footer_height = None
        if styling_config:
            row_heights_cfg = styling_config.get("row_heights", {})
            footer_height = row_heights_cfg.get("footer", row_heights_cfg.get("header"))
        if footer_height is not None:
            try:
                h_val = float(footer_height)
                worksheet.row_dimensions[buffalo_summary_row].height = h_val
                worksheet.row_dimensions[leather_summary_row].height = h_val
            except (ValueError, TypeError): pass

        return next_available_row

    except Exception as summary_err:
        print(f"Warning: Failed processing summary rows: {summary_err}")
        traceback.print_exc()
        return start_row + 2

def _style_row_before_footer(
    worksheet: Worksheet,
    row_num: int,
    num_columns: int,
    sheet_styling_config: Optional[Dict[str, Any]],
    idx_to_id_map: Dict[int, str],
    col1_index: int, # The index of the first column to receive special border handling
    DAF_mode: bool
):
    """
    Applies column-specific styles, a full border, and a specific height
    to the static row before the footer. The first column will only have
    side borders.
    """
    if not sheet_styling_config or row_num <= 0:
        return

    # Set the row height using the 'header' value from the styling config.
    try:
        row_heights = sheet_styling_config.get("row_heights", {})
        header_height = row_heights.get("header")

        if header_height:
            worksheet.row_dimensions[row_num].height = header_height
    except Exception as e:
        print(f"Warning: Could not set row height for row {row_num}. Error: {e}")

    # --- START: Refactored Logic ---
    # Define the two border styles needed for this row
    thin_side = Side(border_style="thin", color="000000")
    
    # Style 1: Full border for all columns except the first
    full_thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Style 2: Side-only border for the first column
    side_only_border = Border(left=thin_side, right=thin_side)
    # --- END: Refactored Logic ---

    # Iterate through each column of the row to apply cell-level styles
    for c_idx in range(1, num_columns + 1):
        try:
            cell = worksheet.cell(row=row_num, column=c_idx)
            current_col_id = idx_to_id_map.get(c_idx)

            # 1. Apply font, alignment, and number formats based on the column ID.
            _apply_cell_style(cell, current_col_id, sheet_styling_config, DAF_mode)

            # --- START: Refactored Logic ---
            # 2. Apply a conditional border based on the column index.
            if c_idx == col1_index:
                # First column gets a border on the sides only
                cell.border = side_only_border
            else:
                # All other columns get a full border
                cell.border = full_thin_border
            # --- END: Refactored Logic ---

        except Exception as e:
            print(f"Warning: Could not style cell at ({row_num}, {c_idx}). Error: {e}")

def _apply_cell_style(cell, column_id: Optional[str], sheet_styling_config: Optional[Dict[str, Any]] = None, DAF_mode: Optional[bool] = False):
    """
    Applies font, alignment, and number format to a cell based on a column ID.
    """
    if not sheet_styling_config or not cell or not column_id:
        return

    try:
        # Get styling configurations using ID-based keys
        default_font_cfg = sheet_styling_config.get("default_font", {})
        default_align_cfg = sheet_styling_config.get("default_alignment", {})
        column_styles = sheet_styling_config.get("column_id_styles", {}) # <-- Uses "column_id_styles"

        # Find column-specific style rules if the ID matches
        col_specific_style = column_styles.get(column_id, {})

        # --- Apply Font ---
        final_font_cfg = default_font_cfg.copy()
        final_font_cfg.update(col_specific_style.get("font", {}))
        if final_font_cfg:
            cell.font = Font(**{k: v for k, v in final_font_cfg.items() if v is not None})

        # --- Apply Alignment ---
        final_align_cfg = default_align_cfg.copy()
        final_align_cfg.update(col_specific_style.get("alignment", {}))
        if final_align_cfg:
            cell.alignment = Alignment(**{k: v for k, v in final_align_cfg.items() if v is not None})
            
        # --- Apply Number Format ---
        number_format = col_specific_style.get("number_format")
        
        # PCS always uses config format, never forced format
        if column_id in ['col_pcs', 'col_qty_pcs']:
            if number_format and cell.number_format != FORMAT_TEXT:
                cell.number_format = number_format
        else:
            # Non-PCS columns follow DAF mode logic
            if number_format and cell.number_format != FORMAT_TEXT and not DAF_mode:
                cell.number_format = number_format
            elif number_format and cell.number_format != FORMAT_TEXT and DAF_mode:
                cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED2
            elif cell.number_format != FORMAT_TEXT and (cell.number_format == FORMAT_GENERAL or cell.number_format is None):
                if isinstance(cell.value, float): cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED2
                elif isinstance(cell.value, int): cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1

    except Exception as style_err:
        print(f"Error applying cell style for ID {column_id}: {style_err}")