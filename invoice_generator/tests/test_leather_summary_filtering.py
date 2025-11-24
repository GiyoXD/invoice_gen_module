import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))

import unittest
from unittest.mock import MagicMock, patch
from openpyxl import Workbook
from invoice_generator.builders.footer_builder import FooterBuilder

class TestLeatherSummaryFiltering(unittest.TestCase):
    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.style_config = {'styling_config': MagicMock()}
        self.context_config = {
            'header_info': {
                'num_columns': 5,
                'column_id_map': {'col_desc': 2, 'col_qty': 3, 'col_pallet': 4}
            },
            'pallet_count': 10,
            'sheet_name': 'Packing list'
        }
        self.data_config = {
            'footer_config': {
                'add_ons': {
                    'leather_summary': {'enabled': True}
                },
                'total_text_column_id': 'col_desc',
                'pallet_count_column_id': 'col_pallet',
                'sum_column_ids': ['col_qty']
            },
            'leather_summary': {}
        }

    def test_both_present(self):
        self.data_config['leather_summary'] = {
            'BUFFALO': {'pallet_count': 5, 'col_qty': 100},
            'COW': {'pallet_count': 5, 'col_qty': 100}
        }
        builder = FooterBuilder(self.ws, 10, self.style_config, self.context_config, self.data_config)
        # Mock internal methods to avoid complex setup
        builder._apply_footer_cell_style = MagicMock()
        builder._apply_footer_row_height = MagicMock()
        
        next_row = builder._build_summary_add_on(10)
        self.assertEqual(next_row, 12) # Should add 2 rows
        self.assertEqual(self.ws.cell(row=10, column=2).value, "TOTAL OF BUFFALO LEATHER")
        self.assertEqual(self.ws.cell(row=11, column=2).value, "TOTAL OF COW LEATHER")

    def test_buffalo_empty(self):
        self.data_config['leather_summary'] = {
            'BUFFALO': {'pallet_count': 0, 'col_qty': 0},
            'COW': {'pallet_count': 5, 'col_qty': 100}
        }
        builder = FooterBuilder(self.ws, 10, self.style_config, self.context_config, self.data_config)
        builder._apply_footer_cell_style = MagicMock()
        builder._apply_footer_row_height = MagicMock()
        
        next_row = builder._build_summary_add_on(10)
        self.assertEqual(next_row, 11) # Should add 1 row
        self.assertEqual(self.ws.cell(row=10, column=2).value, "TOTAL OF COW LEATHER")

    def test_cow_empty(self):
        self.data_config['leather_summary'] = {
            'BUFFALO': {'pallet_count': 5, 'col_qty': 100},
            'COW': {'pallet_count': 0, 'col_qty': 0}
        }
        builder = FooterBuilder(self.ws, 10, self.style_config, self.context_config, self.data_config)
        builder._apply_footer_cell_style = MagicMock()
        builder._apply_footer_row_height = MagicMock()
        
        next_row = builder._build_summary_add_on(10)
        self.assertEqual(next_row, 11) # Should add 1 row
        self.assertEqual(self.ws.cell(row=10, column=2).value, "TOTAL OF BUFFALO LEATHER")

    def test_both_empty(self):
        self.data_config['leather_summary'] = {
            'BUFFALO': {'pallet_count': 0, 'col_qty': 0},
            'COW': {'pallet_count': 0, 'col_qty': 0}
        }
        builder = FooterBuilder(self.ws, 10, self.style_config, self.context_config, self.data_config)
        builder._apply_footer_cell_style = MagicMock()
        builder._apply_footer_row_height = MagicMock()
        
        next_row = builder._build_summary_add_on(10)
        self.assertEqual(next_row, 10) # Should add 0 rows

if __name__ == '__main__':
    unittest.main()
