from logging import log
import logging
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Tuple
import copy

logger = logging.getLogger(__name__)

class TemplateStateBuilder:
    """
    A builder responsible for capturing and restoring the state of a template file.
    This includes the header, footer, and other static content.
    
    Captures the original template state during initialization, before any modifications.
    """
    
    DEBUG = False  # Set to True to enable debug printing

    def __init__(self, worksheet: Worksheet, num_header_cols: int, header_end_row: int, footer_start_row: int, debug: bool = False):
        """
        Initialize and immediately capture template state.
        
        Args:
            worksheet: The worksheet to capture state from
            num_header_cols: Number of header columns
            header_end_row: Last row of the header section
            footer_start_row: First row of the footer section (from template)
            debug: Enable debug printing (default: False)
        """
        self.worksheet = worksheet
        self.header_state: List[List[Dict[str, Any]]] = []
        self.footer_state: List[List[Dict[str, Any]]] = []
        self.header_merged_cells: List[str] = []
        self.footer_merged_cells: List[str] = []
        self.row_heights: Dict[int, float] = {}
        self.column_widths: Dict[int, float] = {}
        self.template_footer_start_row: int = footer_start_row
        self.template_footer_end_row: int = -1
        self.header_end_row = header_end_row
        self.min_row = 1
        self.max_row = self.worksheet.max_row
        self.min_col = 1
        self.num_header_cols = num_header_cols
        self.debug = debug or TemplateStateBuilder.DEBUG  # Use instance or class-level debug flag

        # Store default style objects for comparison
        default_workbook = openpyxl.Workbook()
        default_cell = default_workbook.active['A1']
        self.default_font = default_cell.font
        self.default_fill = default_cell.fill
        self.default_border = default_cell.border
        self.default_alignment = default_cell.alignment
        default_workbook.close() # Close the dummy workbook

        # Calculate max_col based on the maximum column with content in the entire worksheet
        max_col_with_content = 0
        max_row_with_content = 0 # Initialize max_row_with_content
        for r_idx in range(1, self.worksheet.max_row + 1):
            for c_idx in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=r_idx, column=c_idx)
                if self._has_content_or_style(cell):
                    max_col_with_content = max(max_col_with_content, c_idx)
                    max_row_with_content = max(max_row_with_content, r_idx) # Update max_row_with_content
        self.max_col = max(max_col_with_content, self.num_header_cols) # Ensure it's at least num_header_cols
        self.max_row = max(max_row_with_content, self.max_row) # Update self.max_row with max_row_with_content
        
        # Capture template state immediately during initialization
        if self.debug:
            logger.debug(f"Capturing template state during init")
            logger.debug(f"Header: rows 1-{header_end_row}, Footer: rows {footer_start_row}-{self.max_row}")
        self._capture_header(header_end_row)
        self._capture_footer(footer_start_row, self.max_row)
        if self.debug:
            logger.debug(f"State captured: {len(self.header_state)} header rows, {len(self.footer_state)} footer rows")


    def _has_content_or_style(self, cell) -> bool:
        if cell.value is not None and cell.value != '':
            return True
        # Check if any style is applied (not default)
        if cell.font and not self._is_default_style(cell.font, self.default_font): return True
        if cell.fill and not self._is_default_style(cell.fill, self.default_fill): return True
        if cell.border and not self._is_default_style(cell.border, self.default_border): return True
        if cell.alignment and not self._is_default_style(cell.alignment, self.default_alignment): return True
        return False

    def _is_default_style(self, style_obj, default_obj) -> bool:
        if style_obj is None:
            return True
        if default_obj is None: # Should not happen if default_obj is properly initialized
            return False
        
        # Compare relevant attributes for each style type
        if isinstance(style_obj, Font):
            return (
                style_obj.name == default_obj.name and
                style_obj.size == default_obj.size and
                style_obj.bold == default_obj.bold and
                style_obj.italic == default_obj.italic and
                style_obj.underline == default_obj.underline and
                style_obj.strike == default_obj.strike and
                style_obj.color == default_obj.color
            )
        elif isinstance(style_obj, PatternFill):
            return (
                style_obj.fill_type == default_obj.fill_type and
                style_obj.start_color == default_obj.start_color and
                style_obj.end_color == default_obj.end_color
            )
        elif isinstance(style_obj, Border):
            return (
                style_obj.left == default_obj.left and
                style_obj.right == default_obj.right and
                style_obj.top == default_obj.top and
                style_obj.bottom == default_obj.bottom and
                style_obj.diagonal == default_obj.diagonal
            )
        elif isinstance(style_obj, Alignment):
            return (
                style_obj.horizontal == default_obj.horizontal and
                style_obj.vertical == default_obj.vertical and
                style_obj.text_rotation == default_obj.text_rotation and
                style_obj.wrap_text == default_obj.wrap_text and
                style_obj.shrink_to_fit == default_obj.shrink_to_fit and
                style_obj.indent == default_obj.indent
            )
        
        return False # If type not recognized, assume not default

    def _format_cell_style_info(self, cell_info: Dict[str, Any], cell_coord: str) -> str:
        """Format cell styling information for debug logging."""
        parts = []
        
        if cell_info.get('value'):
            val_str = str(cell_info['value'])
            parts.append(f"value='{val_str[:30]}{'...' if len(val_str) > 30 else ''}'")
        
        if cell_info.get('font'):
            font = cell_info['font']
            font_parts = []
            if font.name: font_parts.append(f"name={font.name}")
            if font.size: font_parts.append(f"size={font.size}")
            if font.bold: font_parts.append("bold")
            if font.italic: font_parts.append("italic")
            if font.color and hasattr(font.color, 'rgb'):
                font_parts.append(f"color={font.color.rgb}")
            if font_parts:
                parts.append(f"font({', '.join(font_parts)})")
        
        if cell_info.get('fill'):
            fill = cell_info['fill']
            if fill.fill_type and fill.fill_type != 'none':
                fill_str = f"fill({fill.fill_type}"
                if hasattr(fill.start_color, 'rgb'):
                    fill_str += f", {fill.start_color.rgb}"
                fill_str += ")"
                parts.append(fill_str)
        
        if cell_info.get('border'):
            border = cell_info['border']
            border_parts = []
            if border.left and border.left.style: border_parts.append(f"L:{border.left.style}")
            if border.right and border.right.style: border_parts.append(f"R:{border.right.style}")
            if border.top and border.top.style: border_parts.append(f"T:{border.top.style}")
            if border.bottom and border.bottom.style: border_parts.append(f"B:{border.bottom.style}")
            if border_parts:
                parts.append(f"border({', '.join(border_parts)})")
        
        if cell_info.get('alignment'):
            align = cell_info['alignment']
            align_parts = []
            if align.horizontal: align_parts.append(f"h={align.horizontal}")
            if align.vertical: align_parts.append(f"v={align.vertical}")
            if align.wrap_text: align_parts.append("wrap")
            if align_parts:
                parts.append(f"align({', '.join(align_parts)})")
        
        return f"{cell_coord}: {', '.join(parts)}" if parts else None

    def _get_cell_info(self, worksheet, row, col) -> Dict[str, Any]:
        cell = worksheet.cell(row=row, column=col)
        top_left_cell = cell
        for merged_cell_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_cell_range:
                top_left_cell = worksheet.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col)
                break

        return {
            'value': cell.value,
            'font': copy.copy(top_left_cell.font) if top_left_cell.font and not self._is_default_style(top_left_cell.font, self.default_font) else None,
            'fill': copy.copy(top_left_cell.fill) if top_left_cell.fill and not self._is_default_style(top_left_cell.fill, self.default_fill) else None,
            'border': copy.copy(top_left_cell.border) if top_left_cell.border and not self._is_default_style(top_left_cell.border, self.default_border) else None,
            'alignment': copy.copy(top_left_cell.alignment) if top_left_cell.alignment and not self._is_default_style(top_left_cell.alignment, self.default_alignment) else None,
            'number_format': top_left_cell.number_format,
        }

    def _capture_header(self, end_row: int):
        """
        Captures the state of the header section.
        """
        logger.debug(f"=== CAPTURING HEADER (rows 1 to {end_row}) ===")
        
        # Determine the actual start row of the header by finding the first row with content
        header_start_row = 1
        for r_idx in range(1, end_row + 1):
            if any(self._has_content_or_style(self.worksheet.cell(row=r_idx, column=c_idx))
                   for c_idx in range(1, self.max_col + 1)):
                header_start_row = r_idx
                break

        logger.debug(f"  Header starts at row {header_start_row}, ends at row {end_row}")
        logger.debug(f"  Max columns: {self.max_col}")

        for r_idx in range(header_start_row, end_row + 1):
            row_data = []
            row_has_content = False
            styled_cells = []  # Track cells with interesting styling
            
            for c_idx in range(1, self.max_col + 1):
                cell_info = self._get_cell_info(self.worksheet, r_idx, c_idx)
                row_data.append(cell_info)
                
                # Check if this cell has content
                if cell_info['value'] is not None:
                    row_has_content = True
                
                # Track cells with styling
                if any([cell_info.get('font'), cell_info.get('fill'), cell_info.get('border')]):
                    col_letter = get_column_letter(c_idx)
                    style_str = self._format_cell_style_info(cell_info, f"{col_letter}{r_idx}")
                    if style_str:
                        styled_cells.append(style_str)
            
            self.header_state.append(row_data)
            self.row_heights[r_idx] = self.worksheet.row_dimensions[r_idx].height
            
            # Log row details
            if row_has_content:
                # Show non-empty cells in this row
                non_empty_cells = []
                for c_idx in range(1, self.max_col + 1):
                    cell_val = row_data[c_idx - 1]['value']
                    if cell_val is not None and cell_val != '':
                        col_letter = get_column_letter(c_idx)
                        non_empty_cells.append(f"{col_letter}{r_idx}='{cell_val}'")
                
                if non_empty_cells:
                    logger.debug(f"  Row {r_idx}: {', '.join(non_empty_cells[:5])}" + 
                               (f" ... ({len(non_empty_cells)-5} more)" if len(non_empty_cells) > 5 else ""))
                
                # Show styled cells (limit to first 2 to avoid log spam)
                if styled_cells[:2]:
                    for styled_cell in styled_cells[:2]:
                        logger.debug(f"    Style: {styled_cell}")

        # Capture merged cells within the header range
        header_merges = []
        for merged_cell_range in self.worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_cell_range.bounds
            if header_start_row <= min_row <= end_row and header_start_row <= max_row <= end_row:
                merge_str = str(merged_cell_range)
                self.header_merged_cells.append(merge_str)
                header_merges.append(merge_str)

        if header_merges:
            logger.debug(f"  Captured {len(header_merges)} merged cells: {', '.join(header_merges[:3])}" + 
                       (f" ... ({len(header_merges)-3} more)" if len(header_merges) > 3 else ""))

        # Capture column widths
        for c_idx in range(1, self.max_col + 1):
            self.column_widths[c_idx] = self.worksheet.column_dimensions[get_column_letter(c_idx)].width
        
        logger.debug(f"  ✓ Header capture complete: {len(self.header_state)} rows, {len(self.header_merged_cells)} merges")

    def _capture_footer(self, footer_start_row: int, max_possible_footer_row: int):
        """
        Captures the state of the footer section from the original template.
        
        Args:
            footer_start_row: First row of footer in the template
            max_possible_footer_row: Last row to check for footer content
        """
        logger.debug(f"=== CAPTURING FOOTER (starting from row {footer_start_row}) ===")
        logger.debug(f"  Max columns: {self.max_col}, max search row: {max_possible_footer_row}")
        
        # Footer start is already known from parameter
        self.template_footer_start_row = footer_start_row

        # Find the true max row with content by looking for contiguous ACTUAL content (values or merges)
        # Stop after finding N consecutive empty rows (indicates end of footer)
        # Only check for VALUES or MERGES, not just styling (to avoid capturing 180 styled-but-empty rows)
        MAX_EMPTY_ROWS_BEFORE_STOP = 3
        consecutive_empty_rows = 0
        footer_end_row = footer_start_row
        
        for r_idx in range(footer_start_row, min(footer_start_row + 50, max_possible_footer_row + 1)):  # Limit search to 50 rows
            # Check if row has actual content (values) or is part of a merge
            row_has_value = any(self.worksheet.cell(row=r_idx, column=c_idx).value is not None and 
                               self.worksheet.cell(row=r_idx, column=c_idx).value != ''
                               for c_idx in range(1, self.max_col + 1))
            
            row_has_merge = any(r_idx >= merged_range.min_row and r_idx <= merged_range.max_row
                               for merged_range in self.worksheet.merged_cells.ranges)
            
            if row_has_value or row_has_merge:
                footer_end_row = r_idx
                consecutive_empty_rows = 0
            else:
                consecutive_empty_rows += 1
                if consecutive_empty_rows >= MAX_EMPTY_ROWS_BEFORE_STOP:
                    # Found enough empty rows, footer ends here
                    break

        self.template_footer_end_row = footer_end_row
        logger.debug(f"  Footer ends at row {footer_end_row} ({footer_end_row - footer_start_row + 1} footer rows)")

        for r_idx in range(footer_start_row, footer_end_row + 1):
            row_data = []
            row_has_content = False
            styled_cells = []  # Track cells with interesting styling
            
            for c_idx in range(1, self.max_col + 1):
                cell_info = self._get_cell_info(self.worksheet, r_idx, c_idx)
                row_data.append(cell_info)
                
                # Check if this cell has content
                if cell_info['value'] is not None:
                    row_has_content = True
                
                # Track cells with styling
                if any([cell_info.get('font'), cell_info.get('fill'), cell_info.get('border')]):
                    col_letter = get_column_letter(c_idx)
                    style_str = self._format_cell_style_info(cell_info, f"{col_letter}{r_idx}")
                    if style_str:
                        styled_cells.append(style_str)
            
            self.footer_state.append(row_data)
            self.row_heights[r_idx] = self.worksheet.row_dimensions[r_idx].height
            
            # Log row details
            if row_has_content:
                # Show non-empty cells in this row
                non_empty_cells = []
                for c_idx in range(1, self.max_col + 1):
                    cell_val = row_data[c_idx - 1]['value']
                    if cell_val is not None and cell_val != '':
                        col_letter = get_column_letter(c_idx)
                        non_empty_cells.append(f"{col_letter}{r_idx}='{cell_val}'")
                
                if non_empty_cells:
                    logger.debug(f"  Row {r_idx}: {', '.join(non_empty_cells[:5])}" + 
                               (f" ... ({len(non_empty_cells)-5} more)" if len(non_empty_cells) > 5 else ""))
                
                # Show styled cells (limit to first 2 to avoid log spam)
                if styled_cells[:2]:
                    for styled_cell in styled_cells[:2]:
                        logger.debug(f"    Style: {styled_cell}")

        # Capture merged cells within the footer range
        footer_merges = []
        for merged_cell_range in self.worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_cell_range.bounds
            if footer_start_row <= min_row <= footer_end_row and footer_start_row <= max_row <= footer_end_row:
                merge_str = str(merged_cell_range)
                self.footer_merged_cells.append(merge_str)
                footer_merges.append(merge_str)

        if footer_merges:
            logger.debug(f"  Captured {len(footer_merges)} merged cells: {', '.join(footer_merges[:3])}" + 
                       (f" ... ({len(footer_merges)-3} more)" if len(footer_merges) > 3 else ""))

        # Capture column widths
        for c_idx in range(1, self.max_col + 1):
            self.column_widths[c_idx] = self.worksheet.column_dimensions[get_column_letter(c_idx)].width
        
        logger.debug(f"  ✓ Footer capture complete: {len(self.footer_state)} rows, {len(self.footer_merged_cells)} merges, template footer start: {self.template_footer_start_row}")

    def restore_header_only(self, target_worksheet: Worksheet):
        """
        Restores ONLY the header (structure, values, merges, formatting) to a new clean worksheet.
        This is used when creating a fresh worksheet to avoid template footer conflicts.
        """
        if self.debug:
            logger.debug(f"Restoring header to new worksheet")
            logger.debug(f"Header rows: {len(self.header_state)}, Header merges: {len(self.header_merged_cells)}")
        
        # Restore header cell values and formatting
        for row_idx, row_data in enumerate(self.header_state):
            actual_row = row_idx + self.min_row
            for col_idx, cell_info in enumerate(row_data):
                actual_col = col_idx + self.min_col
                target_cell = target_worksheet.cell(row=actual_row, column=actual_col)
                
                # Restore value
                if cell_info['value'] is not None:
                    target_cell.value = cell_info['value']
                
                # Restore formatting
                if cell_info['font']:
                    target_cell.font = copy.copy(cell_info['font'])
                if cell_info['fill']:
                    target_cell.fill = copy.copy(cell_info['fill'])
                if cell_info['border']:
                    target_cell.border = copy.copy(cell_info['border'])
                if cell_info['alignment']:
                    target_cell.alignment = copy.copy(cell_info['alignment'])
                if cell_info['number_format']:
                    target_cell.number_format = cell_info['number_format']
        
        # Restore header merged cells
        for merged_cell_range_str in self.header_merged_cells:
            try:
                target_worksheet.merge_cells(merged_cell_range_str)
                if self.debug:
                    logger.debug(f"Merged: {merged_cell_range_str}")
            except Exception as e:
                if self.debug:
                    logger.warning(f"Could not merge {merged_cell_range_str}: {e}")
        
        # Restore row heights
        for row_num, height in self.row_heights.items():
            if row_num <= self.header_end_row and height:
                target_worksheet.row_dimensions[row_num].height = height
        
        # Restore column widths
        for col_num, width in self.column_widths.items():
            if width:
                target_worksheet.column_dimensions[get_column_letter(col_num)].width = width
        
        if self.debug:
            logger.debug(f"Header restoration complete")

    def restore_footer_only(self, target_worksheet: Worksheet, footer_start_row: int):
        """
        Restores ONLY the footer (structure, values, merges, formatting) to the new worksheet.
        This places the template footer (static content) AFTER the dynamically created data footer.
        
        Args:
            target_worksheet: The worksheet to restore footer to
            footer_start_row: The row where the template footer should start (after data footer)
        """
        logger.debug(f"restore_footer_only called with:")
        logger.debug(f"footer_start_row parameter: {footer_start_row}")
        logger.debug(f"self.template_footer_start_row: {self.template_footer_start_row}")
        logger.debug(f"self.template_footer_end_row: {self.template_footer_end_row}")
        logger.debug(f"len(self.footer_state): {len(self.footer_state)}")
        
        if self.debug:
            logger.debug(f"Restoring template footer starting at row {footer_start_row}")
            logger.debug(f"Template footer rows: {len(self.footer_state)}, Footer merges: {len(self.footer_merged_cells)}")
        
        # Calculate offset: template footer was at self.template_footer_start_row, now goes to footer_start_row
        offset = footer_start_row - self.template_footer_start_row if self.template_footer_start_row > 0 else 0
        logger.debug(f"Calculated offset: {offset} (footer_start_row={footer_start_row} - template_footer_start_row={self.template_footer_start_row})")
        logger.debug(f"offset: {offset}")
        
        # Restore footer cell values and formatting with offset
        for row_idx, row_data in enumerate(self.footer_state):
            actual_row = self.template_footer_start_row + row_idx + offset
            for col_idx, cell_info in enumerate(row_data):
                actual_col = col_idx + self.min_col
                logger.debug(f"actual_row: {actual_row}, actual_col: {actual_col}")
                target_cell = target_worksheet.cell(row=actual_row, column=actual_col)
                
                # Restore value
                if cell_info['value'] is not None:
                    target_cell.value = cell_info['value']
                
                # Restore formatting
                if cell_info['font']:
                    target_cell.font = copy.copy(cell_info['font'])
                if cell_info['fill']:
                    target_cell.fill = copy.copy(cell_info['fill'])
                if cell_info['border']:
                    target_cell.border = copy.copy(cell_info['border'])
                if cell_info['alignment']:
                    target_cell.alignment = copy.copy(cell_info['alignment'])
                if cell_info['number_format']:
                    target_cell.number_format = cell_info['number_format']
        
        # Restore footer merged cells with offset
        for merged_cell_range_str in self.footer_merged_cells:
            try:
                from openpyxl.utils.cell import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(merged_cell_range_str)
                
                # Adjust row numbers with offset
                min_row += offset
                max_row += offset
                adjusted_range_str = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                target_worksheet.merge_cells(adjusted_range_str)
                if self.debug:
                    logger.debug(f"Merged: {merged_cell_range_str} -> {adjusted_range_str}")
            except Exception as e:
                if self.debug:
                    logger.warning(f"Could not merge {merged_cell_range_str}: {e}")
        
        # Restore row heights for footer rows
        for row_num, height in self.row_heights.items():
            if self.template_footer_start_row <= row_num <= self.template_footer_end_row and height:
                target_worksheet.row_dimensions[row_num + offset].height = height
        
        if self.debug:
            logger.debug(f"Template footer restoration complete ({len(self.footer_state)} rows restored)")

    def restore_state(self, target_worksheet: Worksheet, data_start_row: int, data_table_end_row: int, restore_footer_merges: bool = True):
        """
        Restores the captured FORMATTING (not values) to preserve template structure.
        Only restores merges, heights, widths - does NOT overwrite cell values.
        
        Args:
            target_worksheet: The worksheet to restore state to
            data_start_row: Starting row of data
            data_table_end_row: Ending row of data table
            restore_footer_merges: Whether to restore footer merges (False when FooterBuilder creates its own merges)
        """
        if self.debug:
            logger.debug(f"Restoring formatting (merges, heights, widths):")
            logger.debug(f"Header merges: {len(self.header_merged_cells)}")
            logger.debug(f"Footer merges: {len(self.footer_merged_cells)} (restore: {restore_footer_merges})")
            logger.debug(f"Template footer start row: {self.template_footer_start_row}")
            logger.debug(f"Data table end row: {data_table_end_row}")
        
        # Restore header merged cells without offset
        if self.debug:
            logger.debug(f"Restoring {len(self.header_merged_cells)} header merges...")
        for merged_cell_range_str in self.header_merged_cells:
            try:
                target_worksheet.merge_cells(merged_cell_range_str)
                if self.debug:
                    logger.debug(f"Merged: {merged_cell_range_str}")
            except Exception as e:
                if self.debug:
                    logger.warning(f"Could not merge {merged_cell_range_str}: {e}")

        # Calculate the offset for footer rows and merged cells
        footer_start_row_in_new_sheet = data_table_end_row + 1
        offset = footer_start_row_in_new_sheet - self.template_footer_start_row if self.template_footer_start_row != -1 else 0
        
        if self.debug:
            logger.debug(f"Footer offset: {offset} (template row {self.template_footer_start_row} -> new row {footer_start_row_in_new_sheet})")

        # Restore footer merged cells with offset (only if requested)
        if restore_footer_merges:
            if self.debug:
                logger.debug(f"Restoring {len(self.footer_merged_cells)} footer merges with offset {offset}...")
            for merged_cell_range_str in self.footer_merged_cells:
                try:
                    from openpyxl.utils.cell import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(merged_cell_range_str)
                    
                    # Adjust row numbers for all footer merged cells
                    min_row += offset
                    max_row += offset
                    adjusted_range_str = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                    target_worksheet.merge_cells(adjusted_range_str)
                    if self.debug:
                        logger.debug(f"Merged: {merged_cell_range_str} -> {adjusted_range_str}")
                except Exception as e:
                    if self.debug:
                        logger.warning(f"Could not merge {merged_cell_range_str}: {e}")
        else:
            if self.debug:
                logger.debug(f"Skipping footer merge restoration (FooterBuilder creates its own merges)")

        # Restore row heights for header
        if self.debug:
            logger.debug(f"Restoring row heights...")
        current_row = 1
        for row_data in self.header_state:
            target_worksheet.row_dimensions[current_row].height = self.row_heights.get(current_row, None)
            current_row += 1

        # Restore row heights for footer (with offset)
        for r_offset, row_data in enumerate(self.footer_state):
            r_idx = footer_start_row_in_new_sheet + r_offset
            original_footer_row_idx = self.template_footer_start_row + r_offset
            target_worksheet.row_dimensions[r_idx].height = self.row_heights.get(original_footer_row_idx, None)

        # Restore column widths
        if self.debug:
            logger.debug(f"Restoring column widths...")
        for c_idx, width in self.column_widths.items():
            target_worksheet.column_dimensions[get_column_letter(c_idx)].width = width
        
        if self.debug:
            logger.debug(f"Formatting restoration complete!")
