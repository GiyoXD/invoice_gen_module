
import logging
from typing import Any, Dict, List, Optional, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, Side, Border
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

from ..styling.models import StylingConfigModel
# Legacy apply_cell_style removed - using only StyleRegistry + CellStyler
from ..styling.style_registry import StyleRegistry
from ..styling.cell_styler import CellStyler
from .bundle_accessor import BundleAccessor

class FooterBuilderStyler(BundleAccessor):
    """
    Builds and styles footer sections using pure bundle architecture.
    
    This class handles BOTH structural building (rows, cells, formulas, merges)
    AND styling (fonts, borders, colors, alignment) in a single efficient pass.
    
    Styling logic is delegated to the style_applier module for separation of concerns.
    Uses config bundles for input and @property decorators for frequently accessed values.
    """
    
    def __init__(
        self,
        worksheet: Worksheet,
        footer_row_num: int,
        style_config: Dict[str, Any],
        context_config: Dict[str, Any],
        data_config: Dict[str, Any]
    ):
        """
        Initialize FooterBuilder with bundle configs.
        
        Args:
            worksheet: The worksheet to build in
            footer_row_num: The row number where footer should be placed
            style_config: Bundle containing styling_config
            context_config: Bundle containing header_info, pallet_count, sheet_name, is_last_table, dynamic_desc_used
            data_config: Bundle containing sum_ranges, footer_config, all_tables_data, table_keys, mapping_rules, DAF_mode, override_total_text
        """
        # Initialize base class with common bundles
        super().__init__(
            worksheet=worksheet,
            style_config=style_config,
            context_config=context_config,
            data_config=data_config  # Pass data_config to base via kwargs
        )
        
        # Store FooterBuilder-specific attributes
        self.footer_row_num = footer_row_num
        
        # Initialize StyleRegistry and CellStyler for ID-driven styling
        self.style_registry = None
        self.cell_styler = CellStyler()
        sheet_styling_config = style_config.get('styling_config')
        if sheet_styling_config:
            try:
                styling_dict = sheet_styling_config.model_dump() if hasattr(sheet_styling_config, 'model_dump') else sheet_styling_config
                if isinstance(styling_dict, dict) and 'columns' in styling_dict and 'row_contexts' in styling_dict:
                    self.style_registry = StyleRegistry(styling_dict)
                    logger.info("StyleRegistry initialized successfully for FooterBuilder")
                else:
                    logger.warning(f"⚠️  FooterBuilder: OLD format detected - StyleRegistry NOT initialized")
                    logger.warning(f"   Falling back to legacy style_applier methods")
            except Exception as e:
                logger.warning(f"Could not initialize StyleRegistry (falling back to legacy styling): {e}")
        
        # Track rows that have had height applied to avoid redundant operations
        self._rows_with_height_applied = set()
    
    # ========== Properties for Frequently Accessed Config Values ==========
    # Note: sheet_name, sheet_styling_config inherited from BundleAccessor
    
    @property
    def header_info(self) -> Dict[str, Any]:
        """Header information from context config."""
        return self.context_config.get('header_info', {})
    
    @property
    def sum_ranges(self) -> List[Tuple[int, int]]:
        """Sum ranges from data config."""
        return self.data_config.get('sum_ranges', [])
    
    @property
    def footer_config(self) -> Dict[str, Any]:
        """Footer configuration from data config."""
        return self.data_config.get('footer_config', {})
    
    @property
    def pallet_count(self) -> int:
        """Pallet count from context config."""
        return self.context_config.get('pallet_count', 0)
    
    @property
    def override_total_text(self) -> Optional[str]:
        """Override total text from data config."""
        return self.data_config.get('override_total_text')
    
    @property
    def DAF_mode(self) -> bool:
        """DAF mode flag from data config."""
        return self.data_config.get('DAF_mode', False)
    
    @property
    def all_tables_data(self) -> Optional[Dict[str, Any]]:
        """All tables data from data config."""
        return self.data_config.get('all_tables_data')
    
    @property
    def table_keys(self) -> Optional[List[str]]:
        """Table keys from data config."""
        return self.data_config.get('table_keys')
    
    @property
    def mapping_rules(self) -> Optional[Dict[str, Any]]:
        """Mapping rules from data config."""
        return self.data_config.get('mapping_rules')
    
    @property
    def is_last_table(self) -> bool:
        """Is last table flag from context config."""
        return self.context_config.get('is_last_table', False)
    
    @property
    def dynamic_desc_used(self) -> bool:
        """Dynamic description used flag from context config."""
        return self.context_config.get('dynamic_desc_used', False)

    def _apply_footer_cell_style(self, cell, col_id, row_context='footer'):
        """
        Apply footer cell style to a single cell using StyleRegistry (strict - no legacy fallback).
        
        Args:
            cell: The cell to apply styling to
            col_id: The column ID for this cell
            row_context: The row context to use (default 'footer', can be 'before_footer')
        """
        if not self.style_registry or not col_id:
            logger.error(f"❌ CRITICAL: StyleRegistry not initialized or no col_id for footer cell {cell.coordinate}")
            logger.error(f"   → Ensure config uses bundled format with 'columns' and 'row_contexts'")
            return
        
        # Use specified context for row styling
        style = self.style_registry.get_style(col_id, context=row_context)
        self.cell_styler.apply(cell, style)
        logger.debug(f"Applied StyleRegistry style to {row_context} cell {col_id}")
        
        # Apply row height ONCE per row (only on first column processed)
        row_num = cell.row
        if row_num not in self._rows_with_height_applied:
            row_height = self.style_registry.get_row_height(row_context)
            if row_height:
                self.cell_styler.apply_row_height(self.worksheet, row_num, row_height)
                logger.debug(f"Applied {row_context} row height {row_height} to row {row_num}")
            self._rows_with_height_applied.add(row_num)
    
    def _resolve_column_index(self, col_id, column_map_by_id: Dict[str, int]) -> Optional[int]:
        """
        Resolve a column ID to its actual column index.
        
        Handles both integer and string column IDs, with fallback to column_map_by_id lookup.
        
        Args:
            col_id: The column identifier (can be int, string representing int, or ID string)
            column_map_by_id: Map of column IDs to column indices
            
        Returns:
            The resolved column index (1-based), or None if not found
        """
        if col_id is None:
            return None
        
        # Handle integer column IDs
        if isinstance(col_id, int):
            return col_id + 1
        
        # Handle string column IDs
        if isinstance(col_id, str):
            try:
                # Try to parse as integer
                raw_index = int(col_id)
                return raw_index + 1
            except ValueError:
                # Look up in column map
                return column_map_by_id.get(col_id)
        
        return None

    def build(self) -> int:
        logger.info(f"[FooterBuilder] build() called - footer_row_num={self.footer_row_num}")
        logger.debug(f"[FooterBuilder] footer_config exists: {bool(self.footer_config)}")
        logger.debug(f"[FooterBuilder] footer_config keys: {list(self.footer_config.keys()) if self.footer_config else 'None'}")
        
        if not self.footer_config or self.footer_row_num <= 0:
            logger.error(f"[FooterBuilder] CANNOT BUILD FOOTER - Invalid config or row_num!")
            logger.error(f"   footer_config exists: {bool(self.footer_config)}")
            logger.error(f"   footer_row_num: {self.footer_row_num}")
            if self.footer_config:
                logger.error(f"   footer_config content: {self.footer_config}")
            return -1

        try:
            current_footer_row = self.footer_row_num
            initial_row = current_footer_row
            
            logger.info(f"[FooterBuilder] Starting footer generation at row {current_footer_row}")
            
            # Handle add_blank_before - insert blank row before footer
            add_blank_before = self.footer_config.get("add_blank_before", False)
            if add_blank_before:
                logger.debug(f"Adding blank row before footer at row {current_footer_row}")
                # Leave current_footer_row blank, move footer to next row
                current_footer_row += 1
            
            footer_type = self.footer_config.get("type", "regular")
            
            # Handle before_footer add-on - ONLY for regular footers, not grand_total
            add_ons = self.footer_config.get("add_ons", {})
            before_footer_addon = add_ons.get("before_footer", {})
            before_footer_enabled = before_footer_addon.get("enabled", False)
            
            if before_footer_enabled and footer_type == "regular":
                try:
                    logger.debug(f"Building before_footer row at row {current_footer_row}")
                    self._build_before_footer(current_footer_row, before_footer_addon)
                    logger.debug(f"before_footer row complete at row {current_footer_row}")
                    current_footer_row += 1
                except Exception as bf_err:
                    logger.error(f"Error building before_footer at row {current_footer_row}: {bf_err}")
                    # Non-fatal, continue
            elif before_footer_enabled and footer_type != "regular":
                logger.debug(f"Skipping before_footer for {footer_type} footer type")
            
            logger.info(f"[FooterBuilder] Building {footer_type} footer at row {current_footer_row}")

            try:
                if footer_type == "regular":
                    self._build_regular_footer(current_footer_row)
                    logger.info(f"[FooterBuilder] Regular footer built successfully at row {current_footer_row}")
                elif footer_type == "grand_total":
                    self._build_grand_total_footer(current_footer_row)
                    logger.info(f"[FooterBuilder] Grand total footer built successfully at row {current_footer_row}")
                else:
                    logger.warning(f"Unknown footer type '{footer_type}', using regular footer")
                    self._build_regular_footer(current_footer_row)
            except Exception as footer_build_err:
                logger.error(f"❌ [FooterBuilder] Error building {footer_type} footer at row {current_footer_row}: {footer_build_err}")
                import traceback
                logger.error(traceback.format_exc())
                raise

            # Apply row height to the footer row
            try:
                self._apply_footer_row_height(current_footer_row)
            except Exception as height_err:
                logger.error(f"Error applying footer row height at row {current_footer_row}: {height_err}")
                # Non-fatal, continue
            
            current_footer_row += 1

            # Handle add-ons
            add_ons = self.footer_config.get("add_ons", [])
            if "summary" in add_ons:
                try:
                    logger.debug(f"Building summary add-on starting at row {current_footer_row}")
                    next_row = self._build_summary_add_on(current_footer_row)
                    logger.debug(f"Summary add-on completed, next row: {next_row}")
                    current_footer_row = next_row
                except Exception as addon_err:
                    logger.error(f"Error building summary add-on at row {current_footer_row}: {addon_err}")
                    raise

            total_rows = current_footer_row - initial_row
            logger.info(f"[FooterBuilder] COMPLETE - Started at {initial_row}, ended at {current_footer_row - 1}, total rows: {total_rows}")

            return current_footer_row

        except Exception as e:
            logger.error(f"[FooterBuilder] FATAL ERROR during footer generation starting at row {self.footer_row_num}: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return -1

    def _build_regular_footer(self, current_footer_row: int):
        """Build regular footer with TOTAL: text."""
        logger.debug(f"[FooterBuilder._build_regular_footer] Starting at row {current_footer_row}")
        default_total_text = self.footer_config.get("total_text", "TOTAL:")
        self._build_footer_common(current_footer_row, default_total_text)
        logger.debug(f"[FooterBuilder._build_regular_footer] Complete")

    def _build_before_footer(self, row: int, before_footer_config: Dict[str, Any]):
        """
        Build before_footer row - a row with text that appears BEFORE the main footer.
        Example: "HS.CODE: 4107.12.00" or "LEATHER (HS.CODE: 4107.12.00)"
        
        Args:
            row: The row number to write to
            before_footer_config: Config dict with 'column_id', 'text', and optional 'merge'
        """
        logger.debug(f"[FooterBuilder._build_before_footer] Row {row}, config={before_footer_config}")
        
        num_columns = self.header_info.get('num_columns', 1)
        column_map_by_id = self.header_info.get('column_id_map', {})
        
        # Get target column and text
        column_id = before_footer_config.get('column_id')
        text = before_footer_config.get('text', '')
        merge_span = before_footer_config.get('merge', 0)  # Number of columns to merge
        
        if not column_id or not text:
            logger.warning(f"before_footer missing column_id or text: {before_footer_config}")
            return
        
        # Resolve column index
        col_idx = column_map_by_id.get(column_id)
        if not col_idx:
            logger.error(f"before_footer column_id '{column_id}' not found in column_map")
            return
        
        # Write text to cell
        cell = self.worksheet.cell(row=row, column=col_idx, value=text)
        logger.info(f"[FooterBuilder._build_before_footer] Wrote text to {cell.coordinate}: '{text}'")
        
        # Apply styling using footer row context (same as main footer)
        self._apply_footer_cell_style(cell, column_id, row_context='footer')
        
        # Apply merge if specified
        if merge_span > 0:
            from openpyxl.utils.cell import get_column_letter
            # merge_span is the TOTAL number of columns to span (including current cell)
            # So if merge_span=2, we merge current column + 1 more column
            end_col = col_idx + (merge_span - 1)
            merge_range = f"{get_column_letter(col_idx)}{row}:{get_column_letter(end_col)}{row}"
            try:
                self.worksheet.merge_cells(merge_range)
                logger.debug(f"[FooterBuilder._build_before_footer] Merged cells: {merge_range} (spanning {merge_span} columns)")
            except Exception as e:
                logger.warning(f"Could not merge cells {merge_range}: {e}")
        
        # Apply styling and borders to all cells in the row using footer row context
        idx_to_id_map = {v: k for k, v in column_map_by_id.items()}
        for c_idx in range(1, num_columns + 1):
            cell = self.worksheet.cell(row=row, column=c_idx)
            col_id = idx_to_id_map.get(c_idx)
            self._apply_footer_cell_style(cell, col_id, row_context='footer')
        
        logger.debug(f"[FooterBuilder._build_before_footer] Complete")

    def _build_grand_total_footer(self, current_footer_row: int):
        """Build grand total footer with TOTAL OF: text."""
        logger.debug(f"[FooterBuilder._build_grand_total_footer] Starting at row {current_footer_row}")
        self._build_footer_common(current_footer_row, "TOTAL OF:")
        logger.debug(f"[FooterBuilder._build_grand_total_footer] Complete")
    
    def _build_footer_common(self, current_footer_row: int, default_total_text: str):
        """
        Common footer building logic for both regular and grand total footers.
        
        Args:
            current_footer_row: The row to build the footer in
            default_total_text: Default text to use for total label
        """
        logger.debug(f"[FooterBuilder._build_footer_common] Row {current_footer_row}, text='{default_total_text}'")
        
        num_columns = self.header_info.get('num_columns', 1)
        column_map_by_id = self.header_info.get('column_id_map', {})
        
        logger.debug(f"[FooterBuilder._build_footer_common] num_columns={num_columns}, column_map has {len(column_map_by_id)} entries")

        # Write total text
        total_text = self.override_total_text if self.override_total_text is not None else default_total_text
        total_text_col_id = self.footer_config.get("total_text_column_id")
        total_text_col_idx = self._resolve_column_index(total_text_col_id, column_map_by_id)
        
        logger.info(f"[FooterBuilder._build_footer_common] TOTAL TEXT DEBUG:")
        logger.info(f"   total_text='{total_text}'")
        logger.info(f"   total_text_col_id='{total_text_col_id}'")
        logger.info(f"   total_text_col_idx={total_text_col_idx}")
        logger.info(f"   column_map_by_id={column_map_by_id}")
        
        if total_text_col_idx:
            cell = self.worksheet.cell(row=current_footer_row, column=total_text_col_idx, value=total_text)
            self._apply_footer_cell_style(cell, total_text_col_id)
            logger.info(f"[FooterBuilder._build_footer_common] WROTE TOTAL TEXT to {cell.coordinate} value='{cell.value}'")
        else:
            logger.error(f"[FooterBuilder._build_footer_common] MISSING total_text_column_id in footer config!")
            logger.error(f"   footer_config keys: {list(self.footer_config.keys())}")
            logger.error(f"   total_text_column_id value: {total_text_col_id}")
            logger.error(f"   This footer will have NO total text label!")

        # Write pallet count
        pallet_col_id = self.footer_config.get("pallet_count_column_id")
        pallet_col_idx = self._resolve_column_index(pallet_col_id, column_map_by_id)
        
        logger.debug(f"[FooterBuilder._build_footer_common] Pallet count: {self.pallet_count} at col_id={pallet_col_id}, col_idx={pallet_col_idx}")
        
        if pallet_col_idx and self.pallet_count > 0:
            pallet_text = f"{self.pallet_count} PALLET{'S' if self.pallet_count != 1 else ''}"
            cell = self.worksheet.cell(row=current_footer_row, column=pallet_col_idx, value=pallet_text)
            self._apply_footer_cell_style(cell, pallet_col_id)
            logger.debug(f"[FooterBuilder._build_footer_common] Wrote pallet text to {cell.coordinate}")

        # Write sum formulas
        sum_column_ids = self.footer_config.get("sum_column_ids", [])
        logger.debug(f"[FooterBuilder._build_footer_common] Sum columns: {sum_column_ids}, sum_ranges: {self.sum_ranges}")
        
        if self.sum_ranges:
            for col_id in sum_column_ids:
                col_idx = column_map_by_id.get(col_id)
                if col_idx:
                    col_letter = get_column_letter(col_idx)
                    sum_parts = [f"{col_letter}{start}:{col_letter}{end}" for start, end in self.sum_ranges]
                    formula = f"=SUM({','.join(sum_parts)})"
                    cell = self.worksheet.cell(row=current_footer_row, column=col_idx, value=formula)
                    self._apply_footer_cell_style(cell, col_id)
                    logger.debug(f"[FooterBuilder._build_footer_common] Wrote formula to {cell.coordinate}: {formula}")
        
        # Apply styling to all footer cells
        idx_to_id_map = {v: k for k, v in column_map_by_id.items()}
        cells_styled = 0
        for c_idx in range(1, num_columns + 1):
            cell = self.worksheet.cell(row=current_footer_row, column=c_idx)
            col_id = idx_to_id_map.get(c_idx)
            self._apply_footer_cell_style(cell, col_id)
            cells_styled += 1
        
        logger.debug(f"[FooterBuilder._build_footer_common] Applied styling to {cells_styled} cells")

        # Apply merge rules
        merge_rules = self.footer_config.get("merge_rules", [])
        for rule in merge_rules:
            start_column_id = rule.get("start_column_id")
            colspan = rule.get("colspan")
            resolved_start_col = self._resolve_column_index(start_column_id, column_map_by_id)
            
            if resolved_start_col and colspan:
                end_col = min(resolved_start_col + colspan - 1, num_columns)
                self.worksheet.merge_cells(start_row=current_footer_row, start_column=resolved_start_col, end_row=current_footer_row, end_column=end_col)

    def _build_summary_add_on(self, current_footer_row: int) -> int:
        from ..utils.layout import write_summary_rows # NEW IMPORT
        if self.DAF_mode and self.dynamic_desc_used and self.sheet_name == "Packing list" and self.is_last_table and self.all_tables_data and self.table_keys and self.mapping_rules:
            return write_summary_rows(
                worksheet=self.worksheet,
                start_row=current_footer_row,
                header_info=self.header_info,
                all_tables_data=self.all_tables_data,
                table_keys=self.table_keys,
                footer_config=self.footer_config,
                mapping_rules=self.mapping_rules,
                styling_config=self.sheet_styling_config,
                DAF_mode=self.DAF_mode,
                grand_total_pallets=self.pallet_count
            )
        return current_footer_row
