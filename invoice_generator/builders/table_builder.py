from typing import Any, Dict, List, Optional, Tuple, Union
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import traceback

from invoice_generator.data.data_preparer import prepare_data_rows, parse_mapping_rules
from invoice_generator.utils.layout import unmerge_row, unmerge_block, safe_unmerge_block, apply_column_widths, apply_row_heights
from invoice_generator.utils.writing import fill_static_row, apply_row_merges, write_grand_total_weight_summary, write_header, write_summary_rows, merge_contiguous_cells_by_id, apply_explicit_data_cell_merges_by_id
from invoice_generator.styling.style_applier import apply_cell_style
from .footer_builder import FooterBuilder

# --- Constants for Styling ---
thin_side = Side(border_style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side) # Full grid border
no_border = Border(left=None, right=None, top=None, bottom=None)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
bold_font = Font(bold=True)

# --- Constants for Number Formats ---
FORMAT_GENERAL = 'General'
FORMAT_TEXT = '@'
FORMAT_NUMBER_COMMA_SEPARATED1 = '#,##0'
FORMAT_NUMBER_COMMA_SEPARATED2 = '#,##0.00'

from invoice_generator.styling.models import StylingConfigModel

class TableBuilder:
    def __init__(self,
        worksheet: Worksheet,
        sheet_name: str,
        sheet_config: Dict[str, Any],
        all_sheet_configs: Dict[str, Any],
        data_source: Union[Dict[str, List[Any]], Dict[Tuple, Dict[str, Any]]],
        data_source_type: str,
        header_info: Dict[str, Any],
        mapping_rules: Dict[str, Any],
        sheet_styling_config: Optional[StylingConfigModel] = None,
        add_blank_after_header: bool = False,
        static_content_after_header: Optional[Dict[str, Any]] = None,
        add_blank_before_footer: bool = False,
        static_content_before_footer: Optional[Dict[str, Any]] = None,
        merge_rules_after_header: Optional[Dict[str, int]] = None,
        merge_rules_before_footer: Optional[Dict[str, int]] = None,
        merge_rules_footer: Optional[Dict[str, int]] = None,
        max_rows_to_fill: Optional[int] = None,
        grand_total_pallets: int = 0,
        custom_flag: bool = False,
        data_cell_merging_rules: Optional[Dict[str, Any]] = None,
        DAF_mode: Optional[bool] = False,
    ):
        self.worksheet = worksheet
        self.sheet_name = sheet_name
        self.sheet_config = sheet_config
        self.all_sheet_configs = all_sheet_configs
        self.data_source = data_source
        self.data_source_type = data_source_type
        self.header_info = header_info
        self.mapping_rules = mapping_rules
        self.sheet_styling_config = sheet_styling_config
        self.add_blank_after_header = add_blank_after_header
        self.static_content_after_header = static_content_after_header
        self.add_blank_before_footer = add_blank_before_footer
        self.static_content_before_footer = static_content_before_footer
        self.merge_rules_after_header = merge_rules_after_header
        self.merge_rules_before_footer = merge_rules_before_footer
        self.merge_rules_footer = merge_rules_footer
        self.max_rows_to_fill = max_rows_to_fill
        self.grand_total_pallets = grand_total_pallets
        self.custom_flag = custom_flag
        self.data_cell_merging_rules = data_cell_merging_rules
        self.DAF_mode = DAF_mode

        # Initialize variables that were previously in fill_invoice_data
        self.actual_rows_to_process = 0
        self.data_rows_prepared = []
        self.col1_index = 1
        self.num_static_labels = 0
        self.columns_to_grid = []
        self.desc_col_idx = None
        self.local_chunk_pallets = 0
        self.dynamic_desc_used = False

        self.row_after_header_idx = -1
        self.data_start_row = -1
        self.data_end_row = -1
        self.row_before_footer_idx = -1
        self.footer_row_final = -1

    def build(self) -> Tuple[bool, int, int, int, int]:
        # --- Initialize Variables --- (Keep existing initializations)
        # These are now instance variables

        # get data source pallet count and handle null/conversion errors
        for pallet_count in self.data_source.get("pallet_count", []):
            if pallet_count is not None:
                try:
                    # Convert to float first to handle decimal strings, then to int
                    numeric_pallet_count = float(str(pallet_count).strip())
                    self.local_chunk_pallets += int(numeric_pallet_count)
                except (ValueError, TypeError) as e:
                    # Log the conversion error but continue processing
                    print(f"Warning: Could not convert pallet_count '{pallet_count}' to number: {e}")
                    continue

        # --- Row Index Tracking --- (Keep existing)
        # These are now instance variables

        # Ensure dictionaries/lists are initialized (Keep existing)
        self.static_content_after_header = self.static_content_after_header or {}
        self.static_content_before_footer = self.static_content_before_footer or {}
        self.merge_rules_after_header = self.merge_rules_after_header or {}
        self.merge_rules_before_footer = self.merge_rules_before_footer or {}
        self.merge_rules_footer = self.merge_rules_footer or {} # Initialize footer merge rules
        self.mapping_rules = self.mapping_rules or {}
        col_id_map = self.header_info.get('column_id_map', {})
        column_map = self.header_info.get('column_map', {})
        idx_to_header_map = {v: k for k, v in column_map.items()}

        try:
            self.data_cell_merging_rules = self.data_cell_merging_rules or {}
            # --- Validate Header Info ---
            if not self.header_info or 'second_row_index' not in self.header_info or 'column_map' not in self.header_info or 'num_columns' not in self.header_info:
                print("Error: Invalid header_info provided.")
                return False, -1, -1, -1, 0

            # --- FIX: Extract num_columns and other values from header_info ---
            num_columns = self.header_info['num_columns']
            data_writing_start_row = self.header_info['second_row_index'] + 1
    
            # --- Find Description & Pallet Info Column Indices --- (Keep existing)
            self.desc_col_idx = col_id_map.get("col_desc")
            pallet_info_col_idx = col_id_map.get("col_pallet")
            if pallet_info_col_idx is None: print("Warning: Header 'Pallet Info' not found.")

            # --- ADD/MODIFY THIS PART FOR PALLET INFO INDEX ---
            if pallet_info_col_idx is None:
                print("Warning: Could not find a 'Pallet Info' (e.g., 'Pallet\\nNo') column header.")
            # --- END OF ADDITION/MODIFICATION FOR PALLET INFO INDEX ---

            # --- Get Styling Config --- (Keep existing)
            force_text_headers = []
            effective_header_font = bold_font # Start with default
            effective_header_align = center_alignment # Start with default

            if self.sheet_styling_config:
                if self.sheet_styling_config.header_font:
                    effective_header_font = Font(**self.sheet_styling_config.header_font.dict(exclude_none=True))
                if self.sheet_styling_config.header_alignment:
                    effective_header_align = Alignment(**self.sheet_styling_config.header_alignment.dict(exclude_none=True))
            parsed_rules = parse_mapping_rules(
                mapping_rules=self.mapping_rules,
                column_id_map=col_id_map,
                idx_to_header_map=idx_to_header_map
            )

            # Unpack the results into local variables for the rest of the function to use
            static_value_map = parsed_rules["static_value_map"]
            initial_static_col1_values = parsed_rules["initial_static_col1_values"]
            dynamic_mapping_rules = parsed_rules["dynamic_mapping_rules"]
            formula_rules = parsed_rules["formula_rules"]
            self.col1_index = parsed_rules["col1_index"]
            self.num_static_labels = parsed_rules["num_static_labels"]
            static_column_header_name = parsed_rules["static_column_header_name"]
            apply_special_border_rule = parsed_rules["apply_special_border_rule"]
            fallback_on_none = parsed_rules.get("dynamic_mapping_rules", {}).get("description", {}).get("fallback_on_none")

            # --- Prepare Data Rows for Writing (Determine number of rows needed from source) ---
            # This section remains largely the same, preparing the `data_rows_prepared` list
            # which holds the *input* data, not the calculated formulas.
            self.desc_col_idx = col_id_map.get("col_desc") # Get the description column index
            self.data_rows_prepared, pallet_counts_for_rows, self.dynamic_desc_used, num_data_rows_from_source = prepare_data_rows(
                data_source_type=self.data_source_type,
                data_source=self.data_source,
                dynamic_mapping_rules=dynamic_mapping_rules,
                column_id_map=col_id_map,
                idx_to_header_map=idx_to_header_map,
                desc_col_idx=self.desc_col_idx,
                num_static_labels=self.num_static_labels,
                static_value_map=static_value_map,
                DAF_mode=self.DAF_mode,
            )
    # --- Determine Final Number of Data Rows ---
    # The number of rows to process is the greater of the number of data rows or static labels.
            self.actual_rows_to_process = max(len(self.data_rows_prepared), self.num_static_labels)

            # Optional: Apply max_rows_to_fill constraint if it exists
            if self.max_rows_to_fill is not None and self.max_rows_to_fill >= 0:
                self.actual_rows_to_process = min(self.actual_rows_to_process, self.max_rows_to_fill)

            # Ensure pallet counts list matches the number of rows we intend to process
            if len(pallet_counts_for_rows) < self.actual_rows_to_process: pallet_counts_for_rows.extend([0] * (self.actual_rows_to_process - len(pallet_counts_for_rows)))
            elif len(pallet_counts_for_rows) > self.actual_rows_to_process: pallet_counts_for_rows = pallet_counts_for_rows[:self.actual_rows_to_process]

            # --- Calculate Total Rows to Insert and Row Indices ---
            total_rows_to_insert = 0
            current_row_offset = 0

            # Row after header (static/blank)
            if self.add_blank_after_header:
                self.row_after_header_idx = data_writing_start_row + current_row_offset
                total_rows_to_insert += 1
                current_row_offset += 1
            else:
                self.row_after_header_idx = -1 # Indicate no blank row

            # Data rows
            self.data_start_row = data_writing_start_row + current_row_offset
            if self.actual_rows_to_process > 0:
                self.data_end_row = self.data_start_row + self.actual_rows_to_process - 1
                total_rows_to_insert += self.actual_rows_to_process
                current_row_offset += self.actual_rows_to_process
            else:
                # No data rows to process (can happen if source is empty)
                self.data_end_row = self.data_start_row - 1 # Indicate no data rows

            # Row before footer (static/blank)
            if self.add_blank_before_footer:
                self.row_before_footer_idx = data_writing_start_row + current_row_offset
                total_rows_to_insert += 1
                current_row_offset += 1
            else:
                self.row_before_footer_idx = -1 # Indicate no blank row

            # Calculate final footer row index relative to where this chunk starts
            self.footer_row_final = data_writing_start_row + total_rows_to_insert
            total_rows_to_insert += 1 # Add 1 for the footer itself

            # --- Bulk Insert Rows --- # V11: Only insert if NOT pre-inserted by caller (i.e., for single-table modes)
            if self.data_source_type in ['aggregation', 'DAF_aggregation', "custom_aggregation"]:
                if total_rows_to_insert > 0:
                    try:
                        self.worksheet.insert_rows(data_writing_start_row, amount=total_rows_to_insert)
                        # Unmerge the block covering the inserted rows *before* the footer starts
                        safe_unmerge_block(self.worksheet, data_writing_start_row, self.footer_row_final - 1, num_columns)
                        print("Rows inserted and unmerged successfully.")
                    except Exception as bulk_insert_err:
                        print(f"Error during single-table bulk row insert/unmerge: {bulk_insert_err}")
                        # Adjust fallback row calculation
                        fallback_row = self.header_info.get('second_row_index', 0) + 1
                        return False, fallback_row, -1, -1, 0

            # --- Fill Row After Header (if applicable) --- 

            # --- Prepare DAF Data Dictionary (inside loop now, safer) ---
            # Removed the premature preparation block here.
            # DAF data dict will be prepared inside the loop if data_source_type is DAF_aggregation.

            # --- Fill Data Rows Loop ---
            if self.actual_rows_to_process > 0:
                print(f"--- DEBUG START LOOP (Sheet: {self.worksheet.title}) ---")
                print(f"  data_start_row: {self.data_start_row}")
                print(f"  actual_rows_to_process: {self.actual_rows_to_process}")
                print(f"  num_static_labels: {self.num_static_labels}")
                print(f"  col1_index: {self.col1_index}")
                print(f"  initial_static_col1_values: {initial_static_col1_values}")
                print(f"  data_source_type: {self.data_source_type}")
                # --- END DEBUG START LOOP ---
            try:
                # --- Create a reverse map from index to ID for easy lookups inside the loop ---
                idx_to_id_map = {v: k for k, v in col_id_map.items()}

                data_row_indices_written = []
                for i in range(self.actual_rows_to_process):
                    current_row_idx = self.data_start_row + i
                    data_row_indices_written.append(current_row_idx)
                    
                    row_data = self.data_rows_prepared[i] if i < len(self.data_rows_prepared) else {}

                    # Write initial static values for the first column
                    if i < self.num_static_labels and self.col1_index != -1:
                        cell = self.worksheet.cell(row=current_row_idx, column=self.col1_index)
                        cell.value = initial_static_col1_values[i]
                        apply_cell_style(cell, self.sheet_styling_config, {"col_id": idx_to_id_map.get(self.col1_index), "col_idx": self.col1_index, "static_col_idx": self.col1_index})

                    # Write dynamic data
                    for col_idx, value in row_data.items():
                        if isinstance(value, dict) and value.get("type") == "formula":
                            continue
                        cell = self.worksheet.cell(row=current_row_idx, column=col_idx)
                        cell.value = value
                        apply_cell_style(cell, self.sheet_styling_config, {"col_id": idx_to_id_map.get(col_idx), "col_idx": col_idx, "static_col_idx": self.col1_index})

                    # Write formulas
                    for col_idx, formula_info in formula_rules.items():
                        formula_template = formula_info["template"]
                        input_ids = formula_info["input_ids"]
                        
                        formula = formula_template
                        for input_id in input_ids:
                            input_col_idx = col_id_map.get(input_id)
                            if input_col_idx:
                                col_letter = get_column_letter(input_col_idx)
                                formula = formula.replace(f"{{col_ref_{input_ids.index(input_id)}}}", col_letter)
                        
                        formula = formula.replace("{row}", str(current_row_idx))
                        
                        cell = self.worksheet.cell(row=current_row_idx, column=col_idx)
                        cell.value = f"={formula}"
                        apply_cell_style(cell, self.sheet_styling_config, {"col_id": idx_to_id_map.get(col_idx), "col_idx": col_idx, "static_col_idx": self.col1_index})

                    # Apply data cell merging rules
                    if self.data_cell_merging_rules:
                        apply_explicit_data_cell_merges_by_id(
                            worksheet=self.worksheet,
                            row_num=current_row_idx,
                            column_id_map=col_id_map,
                            num_total_columns=num_columns,
                            merge_rules_data_cells=self.data_cell_merging_rules,
                            sheet_styling_config=self.sheet_styling_config,
                            DAF_mode=self.DAF_mode
                        )

            except Exception as fill_data_err:
                print(f"Error during data filling loop: {fill_data_err}\n{traceback.format_exc()}")
                return False, self.footer_row_final + 1, self.data_start_row, self.data_end_row, 0

        # Merge Description Column if the layout used fallback/static data
            if not self.dynamic_desc_used and self.data_start_row > 0 and self.data_end_row > self.data_start_row:
                desc_col_id = "col_desc" 
                if col_id_map.get(desc_col_id):
                    merge_contiguous_cells_by_id(
                        worksheet=self.worksheet,
                        start_row=self.data_start_row,
                        end_row=self.data_end_row,
                        col_id_to_merge=desc_col_id,
                        column_id_map=col_id_map
                    )

            # Always try to merge the Pallet Info Column if it exists
            if self.data_start_row > 0 and self.data_end_row > self.data_start_row:
                pallet_col_id = "col_pallet" 
                if col_id_map.get(pallet_col_id):
                    merge_contiguous_cells_by_id(
                        worksheet=self.worksheet,
                        start_row=self.data_start_row,
                        end_row=self.data_end_row,
                        col_id_to_merge=pallet_col_id,
                        column_id_map=col_id_map
                    )
            if self.data_start_row > 0 and self.data_end_row > self.data_start_row:
                pallet_col_id = "col_hs" 
                if col_id_map.get(pallet_col_id):
                    merge_contiguous_cells_by_id(
                        worksheet=self.worksheet,
                        start_row=self.data_start_row,
                        end_row=self.data_end_row,
                        col_id_to_merge=pallet_col_id,
                        column_id_map=col_id_map
                    )

    # --- Fill Row Before Footer ---
            if self.add_blank_before_footer and self.row_before_footer_idx > 0:
                try:
                    # Step 1: Fill the row with content (this applies default styles)
                    fill_static_row(self.worksheet, self.row_before_footer_idx, num_columns, self.static_content_before_footer)
                    
                    # Step 2: Apply the special styling and borders for this specific row
                    # _style_row_before_footer(
                    #     worksheet=self.worksheet,
                    #     row_num=self.row_before_footer_idx,
                    #     num_columns=num_columns,
                    #     sheet_styling_config=self.sheet_styling_config,
                    #     idx_to_id_map=idx_to_id_map, # Pass the ID map here
                    #     col1_index=self.col1_index,
                    #     DAF_mode=self.DAF_mode)
                except Exception as fill_bf_err:
                    print(f"Warning: Error filling/styling row before footer: {fill_bf_err}")
            

            # --- Fill Footer Row --- (Keep existing logic)
            # The SUM formulas here should correctly sum the results of the formulas
            # written in the data rows above.
            if self.footer_row_final > 0:
                # Get the footer configuration object from the main sheet config
                footer_config = self.sheet_config.get("footer_configurations", {})
                data_range_to_sum = [(self.data_start_row, self.data_end_row)]

                pallet_count = 0
                if self.data_source_type == "processed_tables":
                    pallet_count = self.local_chunk_pallets
                else:
                    pallet_count = self.grand_total_pallets

                footer_builder = FooterBuilder(
                    worksheet=self.worksheet,
                    footer_row_num=self.footer_row_final,
                    header_info=self.header_info,
                    sum_ranges=data_range_to_sum,
                    footer_config=footer_config,
                    pallet_count=pallet_count,
                    DAF_mode=self.data_source_type == "DAF_aggregation",
                    sheet_styling_config=self.sheet_styling_config
                )
                footer_builder.build()
        # No need to pass font, alignment, num_columns, etc. as the
        # function gets this info from header_info and footer_config.
            # --- Apply Merges ---
            # Apply merges to row after header (if applicable)
            if self.add_blank_after_header and self.row_after_header_idx > 0 and self.merge_rules_after_header:
                apply_row_merges(self.worksheet, self.row_after_header_idx, num_columns, self.merge_rules_after_header)

            # Apply merges to row before footer (if applicable)
            target_row_for_bf_merge = self.row_before_footer_idx if self.add_blank_before_footer and self.row_before_footer_idx > 0 else -1
            if target_row_for_bf_merge > 0 and self.merge_rules_before_footer:
                apply_row_merges(self.worksheet, target_row_for_bf_merge, num_columns, self.merge_rules_before_footer)

            # Apply merges to the footer row itself (if applicable)
            if self.footer_row_final > 0 and self.merge_rules_footer:
                print(f"Applying footer merges to row {self.footer_row_final} with rules: {self.merge_rules_footer}") # Optional Debug
                try:
                    apply_row_merges(self.worksheet, self.footer_row_final, num_columns, self.merge_rules_footer)
                except Exception as footer_merge_err:
                    print(f"Warning: Error applying footer merges: {footer_merge_err}")

            # --- Apply Row Heights --- (Keep existing)
            apply_row_heights(worksheet=self.worksheet, sheet_styling_config=self.sheet_styling_config, header_info=self.header_info, data_row_indices=data_row_indices_written, footer_row_index=self.footer_row_final, row_after_header_idx=self.row_after_header_idx, row_before_footer_idx=self.row_before_footer_idx)

            # --- Finalization --- (Keep existing)
            next_available_row_final = self.footer_row_final + 1
            if self.actual_rows_to_process == 0: self.data_start_row = -1; self.data_end_row = -1
            return True, next_available_row_final, self.data_start_row, self.data_end_row, self.local_chunk_pallets

        except Exception as e:
            # --- Error Handling --- (Keep existing)
            print(f"Critical error in fill_invoice_data: {e}\n{traceback.format_exc()}")
            fallback_row = self.header_info.get('second_row_index', 0) + 1; frf_local = locals().get('footer_row_final', -1)
            if frf_local > 0: fallback_row = max(fallback_row, frf_local + 1)
            else: est_footer = locals().get('initial_insert_point', fallback_row) + locals().get('total_rows_to_insert', 0); fallback_row = max(fallback_row, est_footer)
            return False, fallback_row, -1, -1, 0
