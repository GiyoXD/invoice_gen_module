import logging
from typing import Any, Dict, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook

from ..styling.models import StylingConfigModel
from .header_builder import HeaderBuilderStyler as HeaderBuilder
from .data_table_builder import DataTableBuilderStyler as DataTableBuilder
from .footer_builder import FooterBuilderStyler as FooterBuilder
from .text_replacement_builder import TextReplacementBuilder
from .template_state_builder import TemplateStateBuilder
from ..utils.text_replacement_rules import build_replacement_rules

# Initialize logger for this module
logger = logging.getLogger(__name__)

class LayoutBuilder:
    """
    The Director in the Builder pattern.
    Coordinates all builders to construct the complete document layout.
    
    RECOMMENDED USAGE (Modern Bundled Config Approach):
        Use BuilderConfigResolver to prepare configuration bundles, then pass them
        via style_config, context_config, and layout_config parameters. This approach
        centralizes config resolution logic and eliminates duplication.
        
        Example:
            from invoice_generator.config.builder_config_resolver import BuilderConfigResolver
            
            resolver = BuilderConfigResolver(
                config_loader=config_loader,
                sheet_name='Invoice',
                worksheet=worksheet,
                args=args,
                invoice_data=invoice_data,
                pallets=31
            )
            
            # Get bundles - resolver handles all data extraction
            style_config, context_config, layout_config, data_config = resolver.get_datatable_bundles()
            
            layout_builder = LayoutBuilder(
                workbook=workbook,
                worksheet=worksheet,
                template_worksheet=template,
                style_config=style_config,
                context_config=context_config,
                layout_config={**layout_config, **data_config}  # Merge layout + data
            )
    
    LEGACY USAGE (Individual Parameters):
        Direct parameter passing is still supported for backward compatibility
        but is deprecated. Consider migrating to the resolver approach above.
    """
    def __init__(
        self,
        workbook: Workbook,
        worksheet: Worksheet,
        template_worksheet: Worksheet,
        sheet_name: str = None,
        sheet_config: Dict[str, Any] = None,
        all_sheet_configs: Dict[str, Any] = None,
        invoice_data: Dict[str, Any] = None,
        styling_config: Optional[StylingConfigModel] = None,
        args: Optional[Any] = None,
        final_grand_total_pallets: int = 0,
        enable_text_replacement: bool = False,
        # Optional skip flags for custom processors
        skip_template_header_restoration: bool = False,
        skip_header_builder: bool = False,
        skip_data_table_builder: bool = False,
        skip_footer_builder: bool = False,
        skip_template_footer_restoration: bool = False,
        # Pre-captured template state (for multi-table efficiency)
        template_state_builder = None,
        # Bundled config support (RECOMMENDED - use BuilderConfigResolver)
        style_config: Dict[str, Any] = None,
        context_config: Dict[str, Any] = None,
        layout_config: Dict[str, Any] = None,
    ):
        # Support bundled config approach (unpack if provided)
        if style_config:
            styling_config = style_config.get('styling_config', styling_config)
        if context_config:
            sheet_name = context_config.get('sheet_name', sheet_name)
            invoice_data = context_config.get('invoice_data', invoice_data)
            all_sheet_configs = context_config.get('all_sheet_configs', all_sheet_configs)
            args = context_config.get('args', args)
            final_grand_total_pallets = context_config.get('final_grand_total_pallets', final_grand_total_pallets)
        if layout_config:
            sheet_config = layout_config.get('sheet_config', sheet_config)
            enable_text_replacement = layout_config.get('enable_text_replacement', enable_text_replacement)
            skip_header_builder = layout_config.get('skip_header_builder', skip_header_builder)
            skip_template_header_restoration = layout_config.get('skip_template_header_restoration', skip_template_header_restoration)
            skip_template_footer_restoration = layout_config.get('skip_template_footer_restoration', skip_template_footer_restoration)
            skip_data_table_builder = layout_config.get('skip_data_table_builder', skip_data_table_builder)
            skip_footer_builder = layout_config.get('skip_footer_builder', skip_footer_builder)
        
        # Store data source config from layout_config (if provided by resolver)
        self.provided_data_source = layout_config.get('data_source') if layout_config else None
        self.provided_data_source_type = layout_config.get('data_source_type') if layout_config else None
        self.provided_header_info = layout_config.get('header_info') if layout_config else None
        self.provided_mapping_rules = layout_config.get('mapping_rules') if layout_config else None
        self.provided_resolved_data = layout_config.get('resolved_data') if layout_config else None  # NEW: Support resolved data from TableDataAdapter
        
        self.workbook = workbook  # Output workbook (writable)
        self.worksheet = worksheet  # Output worksheet (writable)
        self.template_worksheet = template_worksheet  # Template worksheet (read-only usage)
        self.sheet_name = sheet_name
        self.sheet_config = sheet_config
        self.all_sheet_configs = all_sheet_configs
        self.invoice_data = invoice_data
        self.styling_config = styling_config
        self.args = args
        self.final_grand_total_pallets = final_grand_total_pallets
        self.enable_text_replacement = enable_text_replacement
        
        # Skip flags for flexible processor customization
        self.skip_template_header_restoration = skip_template_header_restoration
        self.skip_header_builder = skip_header_builder
        self.skip_data_table_builder = skip_data_table_builder
        self.skip_footer_builder = skip_footer_builder
        self.skip_template_footer_restoration = skip_template_footer_restoration
        
        # Pre-captured template state (for multi-table efficiency)
        self.pre_captured_template_state = template_state_builder
        
        logger.debug(f"LayoutBuilder initialized: skip_data_table_builder={self.skip_data_table_builder}")
        
        # Store results after build
        self.header_info = None
        self.next_row_after_footer = -1
        self.data_start_row = -1  # Expose data range for multi-table sum calculation
        self.data_end_row = -1    # Expose data range for multi-table sum calculation
        self.dynamic_desc_used = False  # Expose for summary add-on condition
        self.template_state_builder = None

    def build(self) -> bool:
        """
        Orchestrates all builders in the correct sequence.
        Reads template state from template_worksheet, writes to self.worksheet (output).
        This completely avoids merge conflicts since template and output are separate.
        """
        logger.info(f"Building layout for sheet '{self.sheet_name}'")
        logger.debug(f"Reading from template, writing to output worksheet")
        
        # 1. Text Replacement (if enabled) - Pre-processing
        # Note: This was already done at workbook level, skip here
        if self.enable_text_replacement:
            text_replacer = TextReplacementBuilder(
                workbook=self.workbook,
                invoice_data=self.invoice_data
            )
            if self.args and self.args.DAF:
                text_replacer.build()  # Run both placeholder and DAF replacements
            else:
                text_replacer._replace_placeholders()  # Only placeholders
        
        # 2. Calculate header boundaries for template state capture
        header_row = self.sheet_config.get('header_row', 1)

        header_to_write = self.sheet_config.get('header_to_write')
        num_header_cols = len(header_to_write) if header_to_write else 0
        
        # IMPORTANT: Clarify terminology - there are TWO types of headers:
        # 1. TEMPLATE HEADER: Decorative header section (company name, logo, etc.) - rows 1 to (table_header_row - 1)
        # 2. TABLE HEADER: Column headers for data table (e.g., "Item", "Quantity", "Price") - at table_header_row
        
        # Get table_header_row from config (where the data table column headers are)
        # For multi-table sheets, we use the ORIGINAL sheet_config header_row (from template),
        # not the dynamic header_row that changes for each table
        sheet_layout = self.all_sheet_configs.get(self.sheet_name, {}) if self.all_sheet_configs else {}
        table_header_row = sheet_layout.get('structure', {}).get('header_row', header_row)
        logger.debug(f"[LayoutBuilder DEBUG] sheet_name={self.sheet_name}, header_row={header_row}, table_header_row={table_header_row}")
        logger.debug(f"[LayoutBuilder DEBUG] all_sheet_configs keys: {list(self.all_sheet_configs.keys()) if self.all_sheet_configs else 'None'}")
        
        # Template decorative header spans from row 1 to the row BEFORE the table header
        template_header_start_row = 1
        template_header_end_row = table_header_row - 1  # Decorative header ends BEFORE table header
        
        # Calculate footer_start_row from template (estimate: table_header_row + 2-row table header + minimal data rows)
        # Table header is at table_header_row, second header row at table_header_row + 1
        # Data starts at table_header_row + 2, footer would be around data_start + 2 rows
        template_footer_start_row = table_header_row + 4  # table_header + 2-row header + ~2 data rows
        logger.debug(f"[LayoutBuilder DEBUG] template_header: rows {template_header_start_row}-{template_header_end_row}, table_header: row {table_header_row}, footer_start: row {template_footer_start_row}")
        
        # IMPORTANT: Use table_header_row for HeaderBuilder (where column headers go)
        # This is different from header_row which might be a local/dynamic value
        header_row_for_builder = table_header_row
        
        # 3. Template State Capture - Use pre-captured state if provided, otherwise capture new
        if self.pre_captured_template_state:
            logger.info(f"Using pre-captured template state (multi-table optimization)")
            self.template_state_builder = self.pre_captured_template_state
            logger.debug(f"Reusing template state: {len(self.template_state_builder.header_state)} header rows, {len(self.template_state_builder.footer_state)} footer rows")
        else:
            logger.info(f"Capturing template state from template worksheet")
            try:
                # Enable debug mode if args has debug flag
                debug_mode = getattr(self.args, 'debug', False) if self.args else False
                
                self.template_state_builder = TemplateStateBuilder(
                    worksheet=self.template_worksheet,  # Read from template
                    num_header_cols=num_header_cols,
                    header_end_row=template_header_end_row,  # Use template position, not output position
                    footer_start_row=template_footer_start_row,  # Use template position, not output position
                    debug=debug_mode  # Pass debug flag
                )
                logger.debug(f"Template state captured successfully: {len(self.template_state_builder.header_state)} header rows, {len(self.template_state_builder.footer_state)} footer rows")
            except Exception as e:
                logger.critical(f"CRITICAL: TemplateStateBuilder initialization failed for sheet '{self.sheet_name}'")
                logger.critical(f"Error: {e}", exc_info=True)
                logger.critical(f"Template header range: rows 1-{template_header_end_row}, Footer start: row {template_footer_start_row}")
                return False
            
            # Apply text replacements to captured state (once, before restoration)
            if self.args and self.invoice_data:
                logger.info(f"Applying text replacements to template state")
                try:
                    replacement_rules = build_replacement_rules(self.args)
                    changes = self.template_state_builder.apply_text_replacements(
                        replacement_rules=replacement_rules,
                        invoice_data=self.invoice_data
                    )
                    logger.info(f"Text replacements applied to template state: {changes} changes made")
                except Exception as e:
                    logger.error(f"Failed to apply text replacements: {e}")
                    import traceback
                    logger.error(traceback.format_exc())
        
        # 3b. Template header restoration DEFERRED - will be done AFTER table building
        # This ensures template content aligns with actual column count after filtering
        logger.debug(f"Deferring template header restoration until after table building")
        
        # 4. Header Builder - writes header data to NEW worksheet (unless skipped)
        if not self.skip_header_builder:
            # Convert styling_config dict to StylingConfigModel if needed
            # BUT: If it's already in NEW format (has 'columns' and 'row_contexts'), keep it as-is!
            styling_model = self.styling_config
            if styling_model and not isinstance(styling_model, StylingConfigModel):
                # Check if it's already in NEW format (columns + row_contexts)
                if isinstance(styling_model, dict) and 'columns' in styling_model and 'row_contexts' in styling_model:
                    # NEW format: keep as dict, don't convert to StylingConfigModel
                    logger.debug("Keeping NEW format styling (columns + row_contexts) as dict")
                else:
                    # OLD format: convert to StylingConfigModel
                    try:
                        styling_model = StylingConfigModel(**styling_model)
                    except Exception as e:
                        logger.warning(f"Could not create StylingConfigModel: {e}")
                        styling_model = None

            # Get bundled columns from sheet_config (bundled config v2.1 format)
            # These are in layout_config -> sheet_config -> 'structure' -> 'columns'
            bundled_columns = None
            column_mapping = None  # For template state column shifting
            
            if self.sheet_config:
                structure = self.sheet_config.get('structure', {})
                original_columns = structure.get('columns', [])
                bundled_columns = original_columns
                
                # Filter columns based on DAF/custom mode flags
                if bundled_columns:
                    DAF_mode = self.args.DAF if self.args and hasattr(self.args, 'DAF') else False
                    custom_mode = self.args.custom if self.args and hasattr(self.args, 'custom') else False
                    
                    # Build column mapping BEFORE filtering
                    # Map each template Excel column position to its output position (or None if removed)
                    if DAF_mode or custom_mode:
                        column_mapping = {}
                        template_excel_col = 1  # Current position in template
                        output_excel_col = 1    # Current position in output
                        
                        for col_def in original_columns:
                            col_id = col_def.get('id', '')
                            skip_in_daf = col_def.get('skip_in_daf', False)
                            skip_in_custom = col_def.get('skip_in_custom', False)
                            colspan = col_def.get('colspan', 1)
                            children = col_def.get('children', [])
                            
                            # Calculate actual Excel columns this definition occupies
                            if children:
                                # Parent with children: uses colspan number of Excel columns
                                num_excel_cols = len(children)
                            else:
                                # Simple column: uses colspan
                                num_excel_cols = colspan
                            
                            # Check if column should be skipped
                            should_skip = (DAF_mode and skip_in_daf) or (custom_mode and skip_in_custom)
                            
                            if should_skip:
                                # Mark all Excel columns occupied by this definition as removed
                                for i in range(num_excel_cols):
                                    column_mapping[template_excel_col + i] = None
                                logger.debug(f"Column '{col_id}' removed: template cols {template_excel_col}-{template_excel_col + num_excel_cols - 1} → None")
                            else:
                                # Map all Excel columns to their new positions
                                for i in range(num_excel_cols):
                                    column_mapping[template_excel_col + i] = output_excel_col + i
                                logger.debug(f"Column '{col_id}': template cols {template_excel_col}-{template_excel_col + num_excel_cols - 1} → output cols {output_excel_col}-{output_excel_col + num_excel_cols - 1}")
                                output_excel_col += num_excel_cols
                            
                            template_excel_col += num_excel_cols
                        
                        logger.info(f"Built column mapping for template shifting: {len([v for v in column_mapping.values() if v])} active Excel columns")
                    
                    # Now filter the columns list
                    original_count = len(bundled_columns)
                    bundled_columns = [
                        col for col in bundled_columns
                        if not (DAF_mode and col.get('skip_in_daf', False))
                        and not (custom_mode and col.get('skip_in_custom', False))
                    ]
                    if len(bundled_columns) < original_count:
                        logger.info(f"Filtered bundled_columns: {original_count} → {len(bundled_columns)} (DAF={DAF_mode}, custom={custom_mode})")
                
                if not bundled_columns:
                    logger.warning(f"No columns found in sheet_config.structure for sheet '{self.sheet_name}'")

            try:
                logger.debug(f"Creating HeaderBuilder at row {header_row_for_builder}")
                logger.debug(f"HeaderBuilder input - bundled_columns: {len(bundled_columns) if bundled_columns else 0}, legacy_layout: {len(header_to_write) if header_to_write else 0}")
                header_builder = HeaderBuilder(
                    worksheet=self.worksheet,
                    start_row=header_row_for_builder,  # Use table_header_row (row 21), NOT header_row (row 1)
                    header_layout_config=header_to_write,  # Legacy format (if provided)
                    bundled_columns=bundled_columns,  # Bundled format (preferred)
                    sheet_styling_config=styling_model,
                )
                logger.debug(f"Calling HeaderBuilder.build() starting at row {header_row_for_builder}")
                self.header_info = header_builder.build()
                
                if not self.header_info or not self.header_info.get('column_map'):
                    logger.error(f"HeaderBuilder failed for sheet '{self.sheet_name}'")
                    logger.error(f"header_info or column_map is missing - HALTING EXECUTION")
                    logger.error(f"start_row: {header_row_for_builder}, bundled_columns: {len(bundled_columns) if bundled_columns else 0}")
                    return False
                
                header_end_row = self.header_info.get('second_row_index', header_row_for_builder)
                logger.debug(f"HeaderBuilder completed - rows {header_row_for_builder}-{header_end_row}, {len(self.header_info.get('column_map', {}))} columns")
                
                # DEBUG: Check if font is still set after HeaderBuilder
                if self.worksheet:
                    debug_cell = self.worksheet.cell(row=header_row_for_builder, column=1)
                    logger.debug(f"POST-HeaderBuilder - Cell({header_row_for_builder},1) font: name={debug_cell.font.name}, size={debug_cell.font.size}, bold={debug_cell.font.bold}")
            except Exception as e:
                logger.error(f"HeaderBuilder crashed for sheet '{self.sheet_name}'")
                logger.error(f"Error: {e}", exc_info=True)
                logger.error(f"header_row_for_builder={header_row_for_builder}, header_to_write={header_to_write}, bundled_columns={len(bundled_columns) if bundled_columns else 0}")
                return False
        else:
            logger.info(f"Skipping header builder (skip_header_builder=True)")
            # Check if header_info was pre-provided in layout_config (bundled config pattern)
            if self.sheet_config and 'header_info' in self.sheet_config:
                self.header_info = self.sheet_config['header_info']
                logger.debug(f"Using pre-provided header_info from layout_config")
            else:
                # Must provide dummy header_info for downstream builders
                self.header_info = {'column_map': {}, 'first_row_index': header_row, 'second_row_index': header_row + 1}
            styling_model = self.styling_config

        # 5. Data Table Builder (writes data rows, returns footer position) (unless skipped)
        logger.debug(f"skip_data_table_builder = {self.skip_data_table_builder}")
        if not self.skip_data_table_builder:
            logger.info(f"Entering data table builder block")
            sheet_inner_mapping_rules_dict = self.sheet_config.get('mappings', {})
            add_blank_after_hdr_flag = self.sheet_config.get("add_blank_after_header", False)
            static_content_after_hdr_dict = self.sheet_config.get("static_content_after_header", {})
            add_blank_before_ftr_flag = self.sheet_config.get("add_blank_before_footer", False)
            static_content_before_ftr_dict = self.sheet_config.get("static_content_before_footer", {})
            merge_rules_after_hdr = self.sheet_config.get("merge_rules_after_header", {})
            merge_rules_before_ftr = self.sheet_config.get("merge_rules_before_footer", {})
            merge_rules_footer = self.sheet_config.get("merge_rules_footer", {})
            data_cell_merging_rules = self.sheet_config.get("data_cell_merging_rule", None)
            
            # ========== Data Source Resolution ==========
            # Initialize data_source_type for Python scoping (used in legacy paths)
            data_source_type = None
            
            logger.debug(f"[LayoutBuilder DEBUG] self.provided_resolved_data = {self.provided_resolved_data is not None}")
            logger.debug(f"[LayoutBuilder DEBUG] self.provided_data_source_type = {self.provided_data_source_type}")
            logger.debug(f"[LayoutBuilder DEBUG] self.provided_data_source = {type(self.provided_data_source) if self.provided_data_source is not None else None}")
            
            # Primary path: Use TableDataAdapter-provided resolved_data (modern approach)
            # This is the RECOMMENDED method - data is already prepared
            if self.provided_resolved_data:
                logger.info(f"Using resolver-provided resolved_data (modern approach)")
                # DataTableBuilder expects resolved_data directly, not wrapped in dtb_data_config
                dtb_data_config = self.provided_resolved_data
            # Secondary path: Use BuilderConfigResolver-provided data source (legacy bundled config approach)
            # This still requires DataTableBuilder to call prepare_data_rows internally
            elif (self.provided_data_source_type is not None and 
                self.provided_data_source is not None and
                (not isinstance(self.provided_data_source, dict) or self.provided_data_source)):
                logger.info(f"Using resolver-provided data source: {self.provided_data_source_type}")
                data_to_fill = self.provided_data_source
                data_source_type = self.provided_data_source_type
                sheet_inner_mapping_rules_dict = self.provided_mapping_rules or {}
                
                dtb_data_config = {
                    'data_source': data_to_fill,
                    'data_source_type': data_source_type,
                    'header_info': self.header_info,
                    'mapping_rules': sheet_inner_mapping_rules_dict
                }
            else:
                # LEGACY PATH - DEPRECATED
                # This logic is maintained for backward compatibility but should be replaced
                # by using BuilderConfigResolver + TableDataAdapter in all calling code
                logger.warning(f"Using legacy data source resolution. Consider using BuilderConfigResolver + TableDataAdapter instead")
                data_source_indicator = self.sheet_config.get("data_source")
                data_to_fill = None
                data_source_type = None

                # Handle custom aggregation mode
                if self.args and self.args.custom and data_source_indicator == 'aggregation':
                    data_to_fill = self.invoice_data.get('custom_aggregation_results')
                    data_source_type = 'custom_aggregation'
                    logger.debug(f"Legacy: Using custom_aggregation")

                # Handle DAF and standard aggregation modes
                if data_to_fill is None:
                    # Auto-switch to DAF mode for Invoice/Contract sheets if DAF flag is set
                    if self.args and self.args.DAF and self.sheet_name in ["Invoice", "Contract"]:
                        data_source_indicator = 'DAF_aggregation'

                    if data_source_indicator == 'DAF_aggregation':
                        data_to_fill = self.invoice_data.get('final_DAF_compounded_result')
                        data_source_type = 'DAF_aggregation'
                        logger.debug(f"Legacy: Using DAF_aggregation")
                    elif data_source_indicator == 'aggregation':
                        data_to_fill = self.invoice_data.get('standard_aggregation_results')
                        data_source_type = 'aggregation'
                        logger.debug(f"Legacy: Using standard aggregation")
                    elif 'processed_tables_data' in self.invoice_data and data_source_indicator in self.invoice_data.get('processed_tables_data', {}):
                        data_to_fill = self.invoice_data['processed_tables_data'].get(data_source_indicator)
                        data_source_type = 'processed_tables'
                        logger.debug(f"Legacy: Using processed_tables with key '{data_source_indicator}'")
                        
                # Check if data was found
                if data_to_fill is None:
                    data_source_indicator = self.sheet_config.get("data_source") if not self.provided_data_source_type else self.provided_data_source_type
                    logger.warning(f"Data source '{data_source_indicator}' unknown or data empty. Skipping fill.")
                    # Set next_row_after_footer to a valid value (right after header) for multi-table processors
                    logger.debug(f"[LayoutBuilder DEBUG] Before early return:")
                    logger.debug(f"  header_row={header_row}")
                    logger.debug(f"  self.sheet_config.get('header_row')={self.sheet_config.get('header_row', 'NOT FOUND')}")
                    logger.debug(f"  Calculating next_row_after_footer = header_row + 2 = {header_row + 2}")
                    self.next_row_after_footer = header_row + 2  # After two-row header
                    self.data_start_row = 0
                    self.data_end_row = 0
                    logger.debug(f"[LayoutBuilder DEBUG] Early return: header_row={header_row}, next_row_after_footer={self.next_row_after_footer}")
                    return True
                
                # Legacy path: data_config with raw data
                dtb_data_config = {
                    'data_source': data_to_fill,
                    'data_source_type': data_source_type,
                    'header_info': self.header_info,
                    'mapping_rules': sheet_inner_mapping_rules_dict
                }
            # ========== End Data Source Resolution ==========

            # DataTableBuilder uses the new simplified interface
            try:
                expected_row_start = self.header_info.get('second_row_index', 0) + 1
                logger.debug(f"Creating DataTableBuilder - Expected to start at row {expected_row_start}")
                logger.debug(f"DataTableBuilder input - data_source_type: {dtb_data_config.get('data_source_type')}, has_data: {bool(dtb_data_config.get('data_source'))}")
                
                data_table_builder = DataTableBuilder(
                    worksheet=self.worksheet,
                    header_info=self.header_info,
                    resolved_data=dtb_data_config,
                    sheet_styling_config=styling_model,
                    vertical_merge_columns=['col_desc']  # Always merge col_desc if all values are identical
                )

                logger.debug(f"Calling DataTableBuilder.build()")
                fill_success, footer_row_position, data_start_row, data_end_row, local_chunk_pallets = data_table_builder.build()

                # Store data range for multi-table processors to access
                self.data_start_row = data_start_row
                self.data_end_row = data_end_row
                self.dynamic_desc_used = data_table_builder.dynamic_desc_used  # Track for summary add-on

                if not fill_success:
                    logger.error(f"DataTableBuilder failed for sheet '{self.sheet_name}'")
                    logger.error(f"Failed to fill table data - HALTING EXECUTION")
                    logger.error(f"footer_row_position={footer_row_position}, data_start_row={data_start_row}, data_end_row={data_end_row}")
                    logger.error(f"expected_row_start={expected_row_start}, data_source_type={dtb_data_config.get('data_source_type')}")
                    return False
                
                rows_written = data_end_row - data_start_row + 1 if data_end_row >= data_start_row else 0
                logger.debug(f"DataTableBuilder completed - rows {data_start_row}-{data_end_row} ({rows_written} rows), footer at row {footer_row_position}")
                
                # 5b. NOW restore template header - AFTER table is built
                # This ensures template content aligns with actual number of columns used
                # CRITICAL: This should only restore decorative header (rows 1 to table_header_row-1)
                # It must NOT overwrite the table header row that HeaderBuilder styled
                if not self.skip_template_header_restoration:
                    logger.info(f"Restoring template header AFTER table build (correct column alignment)")
                    try:
                        # Get actual column count from header_info (this reflects filtered columns)
                        actual_num_cols = self.header_info.get('num_columns', None)
                        table_header_row_num = self.header_info.get('second_row_index', 0)
                        logger.debug(f"Template header will use actual column count: {actual_num_cols}")
                        logger.debug(f"Template header ends at row {self.template_state_builder.header_end_row}")
                        logger.debug(f"Table header row is at: {table_header_row_num}")
                        logger.debug(f"These should NOT overlap! (template_end < table_header)")
                        
                        # Set column mapping if columns were filtered
                        if column_mapping:
                            self.template_state_builder.set_column_mapping(column_mapping)
                            logger.info(f"Applied column mapping to template state for filtered columns")
                        
                        self.template_state_builder.restore_header_only(
                            target_worksheet=self.worksheet,
                            actual_num_cols=actual_num_cols
                        )
                        logger.info(f"Template header restored successfully with {actual_num_cols} columns (rows 1-{self.template_state_builder.header_end_row})")
                    except Exception as e:
                        logger.error(f"Failed to restore template header after table build")
                        logger.error(f"Error: {e}", exc_info=True)
                        return False
                else:
                    logger.debug(f"Skipping template header restoration (skip_template_header_restoration=True)")
                
            except Exception as e:
                logger.error(f"DataTableBuilder crashed for sheet '{self.sheet_name}'")
                logger.error(f"Error: {e}", exc_info=True)
                logger.error(f"header_info={self.header_info}")
                logger.error(f"data_config keys: {list(dtb_data_config.keys())}")
                return False
        else:
            logger.info(f"Skipping data table builder (skip_data_table_builder=True)")
            # Provide dummy values for downstream builders
            footer_row_position = header_row + 2  # After header
            data_start_row = 0
            data_end_row = 0
            local_chunk_pallets = 0
            data_source_type = None
        
        # 6. Footer Builder (proper Director pattern - called explicitly by LayoutBuilder) (unless skipped)
        logger.debug(f"Checking FooterBuilder - skip_footer_builder={self.skip_footer_builder}")
        if not self.skip_footer_builder:
            # Prepare footer parameters
            # Use local_chunk_pallets from data if available, otherwise use grand total
            # For multi-table sheets, local_chunk_pallets will be specific to this table
            # For single-table sheets, use the final_grand_total_pallets
            if local_chunk_pallets > 0:
                pallet_count = local_chunk_pallets
            else:
                pallet_count = self.final_grand_total_pallets

            # Get footer config and sum ranges
            # Support both bundled config format ('footer') and legacy format ('footer_configurations')
            footer_config = self.sheet_config.get('footer', self.sheet_config.get('footer_configurations', {}))
            # Support both bundled config format ('data_flow.mappings') and legacy format ('mappings')
            data_flow = self.sheet_config.get('data_flow', {})
            sheet_inner_mapping_rules_dict = data_flow.get('mappings', self.sheet_config.get('mappings', {}))
            data_range_to_sum = []
            if data_start_row > 0 and data_end_row >= data_start_row:
                data_range_to_sum = [(data_start_row, data_end_row)]

            # Bundle configs for FooterBuilder
            fb_style_config = {
                'styling_config': styling_model
            }
            
            fb_context_config = {
                'header_info': self.header_info,
                'pallet_count': pallet_count,
                'sheet_name': self.sheet_name,
                'is_last_table': True,
                'dynamic_desc_used': False  # TODO: Track this if needed
            }
            
            fb_data_config = {
                'sum_ranges': data_range_to_sum,
                'footer_config': footer_config,
                'all_tables_data': None,  # TODO: Pass if multi-table support needed
                'table_keys': None,
                'mapping_rules': sheet_inner_mapping_rules_dict,
                'DAF_mode': self.args.DAF if self.args and hasattr(self.args, 'DAF') else False,
                'override_total_text': None
            }

            logger.debug(f"Creating FooterBuilder at row {footer_row_position}")
            logger.debug(f"FooterBuilder input - footer_type: {footer_config.get('type', 'regular')}, add_blank_before: {footer_config.get('add_blank_before', False)}, pallet_count: {pallet_count}")
            
            try:
                footer_builder = FooterBuilder(
                    worksheet=self.worksheet,
                    footer_row_num=footer_row_position,
                    style_config=fb_style_config,
                    context_config=fb_context_config,
                    data_config=fb_data_config
                )
                
                logger.debug(f"Calling FooterBuilder.build() with footer_row_position={footer_row_position}")
                footer_start = footer_row_position
                self.next_row_after_footer = footer_builder.build()
                
                # Validate footer builder result
                if self.next_row_after_footer is None or self.next_row_after_footer <= 0:
                    logger.error(f"FooterBuilder failed for sheet '{self.sheet_name}'")
                    logger.error(f"Invalid next_row_after_footer={self.next_row_after_footer} - HALTING EXECUTION")
                    logger.error(f"footer_row_position={footer_row_position}, sum_ranges={data_range_to_sum}")
                    logger.error(f"footer_config: {footer_config}")
                    return False
                
                footer_rows_written = self.next_row_after_footer - footer_start
                logger.debug(f"FooterBuilder completed - rows {footer_start}-{self.next_row_after_footer - 1} ({footer_rows_written} rows), next available: {self.next_row_after_footer}")
            except Exception as e:
                logger.error(f"FooterBuilder crashed for sheet '{self.sheet_name}'")
                logger.error(f"Error: {e}", exc_info=True)
                logger.error(f"footer_row_position={footer_row_position}, pallet_count={pallet_count}")
                logger.error(f"footer_config: {footer_config}")
                return False
            
            # Apply footer height to all footer rows (including add-ons like grand total)
            if self.next_row_after_footer > footer_row_position:
                # Multiple footer rows were created (e.g., regular footer + grand total)
                for footer_row in range(footer_row_position, self.next_row_after_footer):
                    self._apply_footer_row_height(footer_row, styling_model)
            else:
                # Single footer row
                self._apply_footer_row_height(footer_row_position, styling_model)
        else:
            logger.info(f"Skipping footer builder (skip_footer_builder=True)")
            # No footer, so next row is right after data (or header if no data)
            self.next_row_after_footer = footer_row_position
        
        # 7. Template Footer Restoration (unless skipped)
        # Restore the template footer (static content like "Manufacture:", etc.) AFTER the dynamic footer
        # This places the template footer below the data footer
        if not self.skip_template_footer_restoration:
            write_pointer_row = self.next_row_after_footer  # Next available row after dynamic footer
            
            # Validate that we have a valid row position before attempting restoration
            if write_pointer_row is None or write_pointer_row <= 0:
                logger.error(f"Cannot restore template footer - invalid write_pointer_row={write_pointer_row}")
                logger.error(f"This indicates a previous builder failed - HALTING EXECUTION")
                return False
            
            template_footer_rows = self.template_state_builder.max_row - self.template_state_builder.template_footer_start_row + 1
            logger.info(f"Restoring template footer after row {write_pointer_row}")
            logger.debug(f"Template footer restoration - Source rows: {self.template_state_builder.template_footer_start_row}-{self.template_state_builder.max_row} ({template_footer_rows} rows), Target start: {write_pointer_row}")
            
            # Calculate actual number of columns from bundled config
            actual_num_cols = None
            if self.sheet_config and 'structure' in self.sheet_config:
                bundled_columns = self.sheet_config['structure'].get('columns', [])
                if bundled_columns:
                    actual_num_cols = len(bundled_columns)
                    logger.debug(f"Using actual column count from config: {actual_num_cols}")
            
            # Set column mapping if columns were filtered
            if column_mapping:
                self.template_state_builder.set_column_mapping(column_mapping)
                logger.info(f"Applied column mapping to template state for footer restoration")
            
            try:
                self.template_state_builder.restore_footer_only(
                    target_worksheet=self.worksheet,  # Write to output worksheet
                    footer_start_row=write_pointer_row,
                    actual_num_cols=actual_num_cols
                )
                logger.debug(f"Template footer restored successfully - rows {write_pointer_row}-{write_pointer_row + template_footer_rows - 1}")
            except Exception as e:
                logger.error(f"Failed to restore footer from template for sheet '{self.sheet_name}'")
                logger.error(f"Error: {e}", exc_info=True)
                logger.error(f"Attempted to restore footer at row {write_pointer_row}")
                logger.error(f"Template footer range: {self.template_state_builder.template_footer_start_row}-{self.template_state_builder.max_row}")
                return False
        else:
            logger.debug(f"Skipping template footer restoration (skip_template_footer_restoration=True)")
        
        logger.info(f"Layout built successfully for sheet '{self.sheet_name}'")
        
        return True
    
    def _apply_footer_row_height(self, footer_row: int, styling_config):
        """Helper method to apply footer height to a single footer row."""
        if not styling_config:
            return
        
        # Handle both NEW format (dict with 'row_contexts') and OLD format (StylingConfigModel with rowHeights)
        row_heights_cfg = None
        if isinstance(styling_config, dict):
            # NEW format: row heights are in row_contexts.footer.row_height
            if 'row_contexts' in styling_config:
                footer_context = styling_config['row_contexts'].get('footer', {})
                if 'row_height' in footer_context:
                    # NEW format stores height directly in context
                    height = footer_context['row_height']
                    if height:
                        self.worksheet.row_dimensions[footer_row].height = height
                        logger.debug(f"Applied footer height {height} to row {footer_row} (NEW format)")
                return
        elif hasattr(styling_config, 'rowHeights'):
            # OLD format: StylingConfigModel with rowHeights attribute
            row_heights_cfg = styling_config.rowHeights
        
        if not row_heights_cfg:
            return
        
        footer_height_config = row_heights_cfg.get("footer")
        match_header_height_flag = row_heights_cfg.get("footer_matches_header_height", True)
        
        # Determine the footer height
        final_footer_height = None
        if match_header_height_flag:
            # Get header height from config
            header_height = row_heights_cfg.get("header")
            if header_height is not None:
                final_footer_height = header_height
        if final_footer_height is None and footer_height_config is not None:
            final_footer_height = footer_height_config
        
        # Apply the height
        if final_footer_height is not None and footer_row > 0:
            try:
                h_val = float(final_footer_height)
                if h_val > 0:
                    self.worksheet.row_dimensions[footer_row].height = h_val
            except (ValueError, TypeError):
                pass
