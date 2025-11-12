import logging
from typing import Any, Dict, List, Optional
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

from ..styling.models import StylingConfigModel
from ..styling.style_applier import apply_header_style, apply_cell_style
from ..utils.layout import calculate_header_dimensions
from openpyxl.utils import get_column_letter

class HeaderBuilderStyler:
    def __init__(
        self,
        worksheet: Worksheet,
        start_row: int,
        header_layout_config: Optional[List[Dict[str, Any]]] = None,
        bundled_columns: Optional[List[Dict[str, Any]]] = None,
        sheet_styling_config: Optional[StylingConfigModel] = None,
    ):
        """
        Initialize HeaderBuilder with either legacy or bundled config.
        
        Args:
            worksheet: The worksheet to write to
            start_row: Starting row for header
            header_layout_config: Legacy format (list with row/col/text/id/rowspan/colspan)
            bundled_columns: Bundled format (list with id/header/format/rowspan/colspan/children)
            sheet_styling_config: Styling configuration
        """
        self.worksheet = worksheet
        self.start_row = start_row
        self.sheet_styling_config = sheet_styling_config
        
        # Convert bundled columns to internal format if provided
        if bundled_columns:
            logger.info(f"Using BUNDLED config (columns={len(bundled_columns)})")
            # Show first column as example
            if bundled_columns:
                sample = bundled_columns[0]
                sample_header = str(sample.get('header', '')).encode('ascii', 'replace').decode('ascii')
                logger.debug(f"Sample bundled column: id='{sample.get('id')}', header='{sample_header}', format='{sample.get('format')}'")
            self.header_layout_config = self._convert_bundled_columns(bundled_columns)
            logger.debug(f"Converted to {len(self.header_layout_config)} header cells")
        else:
            logger.warning(f"Using legacy config")
            self.header_layout_config = header_layout_config or []

    def build(self) -> Optional[Dict[str, Any]]:
        if not self.header_layout_config or self.start_row <= 0:
            return None

        num_header_rows, num_header_cols = calculate_header_dimensions(self.header_layout_config)

        first_row_index = self.start_row
        last_row_index = self.start_row
        max_col = 0
        column_map = {}
        column_id_map = {}

        for cell_config in self.header_layout_config:
            row_offset = cell_config.get('row', 0)
            col_offset = cell_config.get('col', 0)
            text = cell_config.get('text', '')
            cell_id = cell_config.get('id')
            rowspan = cell_config.get('rowspan', 1)
            colspan = cell_config.get('colspan', 1)

            cell_row = self.start_row + row_offset
            cell_col = 1 + col_offset

            last_row_index = max(last_row_index, cell_row + rowspan - 1)
            max_col = max(max_col, cell_col + colspan - 1)

            cell = self.worksheet.cell(row=cell_row, column=cell_col, value=text)
            apply_header_style(cell, self.sheet_styling_config)
            
            context = {
                "col_id": cell_id,
                "col_idx": cell_col,
                "is_header": True
            }
            apply_cell_style(cell, self.sheet_styling_config, context)

            if cell_id:
                column_map[text] = get_column_letter(cell_col)
                column_id_map[cell_id] = cell_col

            if rowspan > 1 or colspan > 1:
                self.worksheet.merge_cells(start_row=cell_row, start_column=cell_col,
                                      end_row=cell_row + rowspan - 1, end_column=cell_col + colspan - 1)

        return {
            'first_row_index': first_row_index,
            'second_row_index': last_row_index,
            'column_map': column_map,
            'column_id_map': column_id_map,
            'num_columns': max_col
        }
    
    def _convert_bundled_columns(self, columns: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Convert bundled columns format to internal header_layout_config format.
        
        Bundled format:
            {"id": "col_po", "header": "P.O. №", "format": "@", "rowspan": 2}
        
        Internal format:
            {"row": 0, "col": 1, "text": "P.O. №", "id": "col_po", "rowspan": 2, "colspan": 1}
        """
        headers = []
        col_index = 0
        
        for col in columns:
            col_id = col.get('id', '')
            header_text = col.get('header', '')
            rowspan = col.get('rowspan', 1)
            colspan = col.get('colspan', 1)
            
            # Handle parent column with children (e.g., Quantity with PCS/SF)
            if 'children' in col:
                # Add parent header
                headers.append({
                    'row': 0,
                    'col': col_index,
                    'text': header_text,
                    'id': col_id,
                    'rowspan': 1,
                    'colspan': len(col['children'])
                })
                
                # Add children headers
                for child in col['children']:
                    headers.append({
                        'row': 1,
                        'col': col_index,
                        'text': child.get('header', ''),
                        'id': child.get('id', ''),
                        'rowspan': 1,
                        'colspan': 1
                    })
                    col_index += 1
            else:
                headers.append({
                    'row': 0,
                    'col': col_index,
                    'text': header_text,
                    'id': col_id,
                    'rowspan': rowspan,
                    'colspan': colspan
                })
                col_index += 1
        
        return headers
