from openpyxl import Workbook
from openpyxl.styles import Font

# Create a test workbook
wb = Workbook()
ws = wb.active

# Test 1: Create font explicitly with all parameters
cell1 = ws['A1']
cell1.value = "Test 1"
cell1.font = Font(name='Times New Roman', size=12, bold=True, italic=False, color=None)

# Test 2: Create font with only non-None parameters
cell2 = ws['A2']
cell2.value = "Test 2"
cell2.font = Font(name='Times New Roman', size=12, bold=True)

# Test 3: Create font without explicit None
cell3 = ws['A3']
cell3.value = "Test 3"
font3 = Font()
font3.name = 'Times New Roman'
font3.size = 12
font3.bold = True
cell3.font = font3

print("Before save:")
print(f"A1: name={cell1.font.name}, size={cell1.font.size}, bold={cell1.font.bold}")
print(f"A2: name={cell2.font.name}, size={cell2.font.size}, bold={cell2.font.bold}")
print(f"A3: name={cell3.font.name}, size={cell3.font.size}, bold={cell3.font.bold}")

wb.save('test_font_methods.xlsx')

# Load and check
from openpyxl import load_workbook
wb2 = load_workbook('test_font_methods.xlsx')
ws2 = wb2.active

print("\nAfter load:")
print(f"A1: name={ws2['A1'].font.name}, size={ws2['A1'].font.size}, bold={ws2['A1'].font.bold}")
print(f"A2: name={ws2['A2'].font.name}, size={ws2['A2'].font.size}, bold={ws2['A2'].font.bold}")
print(f"A3: name={ws2['A3'].font.name}, size={ws2['A3'].font.size}, bold={ws2['A3'].font.bold}")
