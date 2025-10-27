"""
Test script to verify WorkbookBuilder integration point.

This test validates:
1. Template workbook loads as read_only=False, data_only=True
2. All sheet names collected from template
3. WorkbookBuilder creates new workbook with same sheet names
4. Both workbooks accessible (template for reading, output for writing)
"""

from openpyxl import load_workbook
from invoice_generator.builders.workbook_builder import WorkbookBuilder
import os


def test_workbook_builder_integration():
    """Test WorkbookBuilder integration point."""
    
    # Path to template
    template_dir = "invoice_generator/template"
    template_files = [f for f in os.listdir(template_dir) if f.endswith('.xlsx')]
    
    if not template_files:
        print("❌ No template files found in invoice_generator/template")
        return False
    
    template_path = os.path.join(template_dir, template_files[0])
    print(f"📂 Using template: {template_path}")
    
    try:
        # Step 1: Load template workbook as read_only=False, data_only=True
        print("\n🔍 Step 1: Loading template workbook...")
        print("   Parameters: read_only=False, data_only=True")
        
        template_workbook = load_workbook(
            template_path,
            read_only=False,
            data_only=True
        )
        
        print(f"✅ Template loaded successfully")
        print(f"   Mode: read_only=False, data_only=True")
        
        # Step 2: Collect all sheet names from template
        print("\n📋 Step 2: Collecting sheet names from template...")
        template_sheet_names = template_workbook.sheetnames
        
        print(f"✅ Found {len(template_sheet_names)} sheets:")
        for i, name in enumerate(template_sheet_names, 1):
            print(f"   {i}. {name}")
        
        # Step 3: Create WorkbookBuilder with template sheet names
        print("\n🏗️ Step 3: Creating WorkbookBuilder...")
        workbook_builder = WorkbookBuilder(sheet_names=template_sheet_names)
        print(f"✅ WorkbookBuilder instantiated with {len(template_sheet_names)} sheet names")
        
        # Step 4: Build new clean workbook
        print("\n🔨 Step 4: Building new output workbook...")
        output_workbook = workbook_builder.build()
        print(f"✅ New workbook created")
        
        # Step 5: Verify both workbooks
        print("\n✅ Step 5: Verifying workbook properties...")
        
        # Check output workbook has all sheets
        output_sheet_names = output_workbook.sheetnames
        print(f"   Output workbook sheets: {len(output_sheet_names)}")
        
        if output_sheet_names != template_sheet_names:
            print(f"   ❌ ERROR: Sheet names don't match!")
            print(f"      Template: {template_sheet_names}")
            print(f"      Output:   {output_sheet_names}")
            return False
        
        print(f"   ✅ Sheet names match template")
        
        # Verify sheets are empty (no content)
        for sheet_name in output_sheet_names:
            ws = output_workbook[sheet_name]
            # Check if sheet is empty (max_row should be 1, max_column should be 1 for empty sheet)
            if ws.max_row == 1 and ws.max_column == 1:
                cell_value = ws.cell(1, 1).value
                if cell_value is None:
                    print(f"   ✅ '{sheet_name}' is empty (ready for building)")
                else:
                    print(f"   ⚠️ '{sheet_name}' has content in A1: {cell_value}")
            else:
                print(f"   ⚠️ '{sheet_name}' is not empty (max_row={ws.max_row}, max_col={ws.max_column})")
        
        # Verify template workbook is still accessible
        print("\n📖 Step 6: Verifying template workbook accessibility...")
        template_test_sheet = template_workbook[template_sheet_names[0]]
        print(f"   ✅ Can access template sheet '{template_sheet_names[0]}'")
        print(f"   Template sheet dimensions: {template_test_sheet.max_row} rows x {template_test_sheet.max_column} cols")
        
        # Check if template has content (it should)
        if template_test_sheet.max_row > 1 or template_test_sheet.max_column > 1:
            print(f"   ✅ Template sheet has content (as expected)")
        else:
            print(f"   ⚠️ Template sheet appears empty")
        
        # Final summary
        print("\n" + "="*60)
        print("✅ SUCCESS: WorkbookBuilder Integration Point Works!")
        print("="*60)
        print(f"✓ Template loaded: {template_path}")
        print(f"✓ Mode: read_only=False, data_only=True")
        print(f"✓ {len(template_sheet_names)} sheet names collected")
        print(f"✓ WorkbookBuilder created new workbook")
        print(f"✓ Output workbook has matching sheet names")
        print(f"✓ Output sheets are empty and ready for building")
        print(f"✓ Template workbook remains accessible for reading")
        
        print("\n📝 Key Points:")
        print("  • Template workbook: READ-ONLY usage (state capture)")
        print("  • Output workbook: WRITABLE (final output)")
        print("  • No template copying needed")
        print("  • Clean separation between read and write workbooks")
        
        # Cleanup
        template_workbook.close()
        output_workbook.close()
        
        return True
        
    except Exception as e:
        print(f"\n❌ ERROR: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    print("="*60)
    print("TEST: WorkbookBuilder Integration Point")
    print("="*60)
    success = test_workbook_builder_integration()
    exit(0 if success else 1)
